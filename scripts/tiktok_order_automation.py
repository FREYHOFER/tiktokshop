#!/usr/bin/env python3
"""Fetch TikTok Shop orders and prepare Libri customer-order packages.

Safe defaults:
- TikTok orders are only read.
- Libri files are prepared locally.
- No final Libri checkout is submitted.

The script can also read a manually exported TikTok "Versandbereit Bestellung"
CSV so the workflow can be tested before TikTok API credentials are available.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import hashlib
import hmac
import html
import json
import os
import re
import sys
import time
import urllib.error
import subprocess
import urllib.parse
import urllib.request
from dataclasses import asdict, dataclass, field
from pathlib import Path
from zoneinfo import ZoneInfo

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from fetch_libri_product_pages import load_env_file  # noqa: E402


LIBRI_IMPORT_HEADERS = ["Kundennummer", "Artikel Nr", "Autor/Titel", "Menge", "Vormerkkennzeichen", "Bestellzeichen"]
DEFAULT_BASE_URL = "https://open-api.tiktokglobalshop.com"
DEFAULT_OUTPUT_ROOT = Path("outputs") / "order_automation"
DEFAULT_STATE_PATH = DEFAULT_OUTPUT_ROOT / "state.json"
DEFAULT_TIMEZONE = "Europe/Berlin"
TIKTOK_STATUS_READY = {"AWAITING_SHIPMENT", "VERSANDBEREIT", "READY_TO_SHIP"}


@dataclass
class CustomerAddress:
    name: str = ""
    phone: str = ""
    email: str = ""
    country: str = ""
    state: str = ""
    district: str = ""
    city: str = ""
    zipcode: str = ""
    street: str = ""
    house: str = ""
    full_address: str = ""
    raw: dict = field(default_factory=dict)


@dataclass
class OrderLine:
    ean: str
    seller_sku: str
    product_name: str
    quantity: int
    sku_id: str = ""
    warnings: list[str] = field(default_factory=list)


@dataclass
class AutomationOrder:
    order_id: str
    package_id: str
    status: str
    fulfillment_type: str
    delivery_option: str
    buyer_username: str
    created_time: str
    paid_time: str
    address: CustomerAddress
    lines: list[OrderLine]
    source: str
    raw: dict = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)


class TikTokApiError(RuntimeError):
    pass


def clean(value: object) -> str:
    return str(value or "").replace("\ufeff", "").strip()


def bool_env(value: str, default: bool = False) -> bool:
    if not clean(value):
        return default
    return clean(value).casefold() in {"1", "true", "yes", "y", "on"}


def env_value(env: dict[str, str], key: str, default: str = "") -> str:
    return clean(os.environ.get(key) or env.get(key) or default)


def normalize_order_id(value: object) -> str:
    return re.sub(r"\s+", "", clean(value))


def extract_ean(*values: object) -> str:
    for value in values:
        text = clean(value)
        if not text:
            continue
        for match in re.findall(r"(?:97[89])[\d\-\s]{10,20}", text):
            digits = re.sub(r"\D", "", match)
            if len(digits) == 13:
                return digits
    return ""


def parse_int(value: object, default: int = 0) -> int:
    text = clean(value).replace(",", ".")
    if not text:
        return default
    try:
        return max(int(float(text)), 0)
    except ValueError:
        return default


def safe_filename(value: str) -> str:
    text = re.sub(r"[^A-Za-z0-9_.-]+", "_", clean(value))
    return text[:80] or "unknown"


def first_value(source: dict, *keys: str) -> str:
    for key in keys:
        if key in source and clean(source.get(key)):
            return clean(source.get(key))
    return ""


def nested_value(source: dict, *paths: str) -> object:
    for path in paths:
        current: object = source
        ok = True
        for part in path.split("."):
            if isinstance(current, dict) and part in current:
                current = current[part]
            else:
                ok = False
                break
        if ok and current not in (None, ""):
            return current
    return ""


def compact_json(data: object) -> str:
    return json.dumps(data, ensure_ascii=False, separators=(",", ":"))


class TikTokShopClient:
    def __init__(self, env: dict[str, str], env_path: Path):
        self.env = env
        self.env_path = env_path
        self.app_key = env_value(env, "TIKTOK_APP_KEY")
        self.app_secret = env_value(env, "TIKTOK_APP_SECRET")
        self.access_token = env_value(env, "TIKTOK_ACCESS_TOKEN")
        self.shop_cipher = env_value(env, "TIKTOK_SHOP_CIPHER")
        self.base_url = env_value(env, "TIKTOK_BASE_URL", DEFAULT_BASE_URL).rstrip("/")
        self.version = env_value(env, "TIKTOK_API_VERSION", "202309")
        self.include_version = bool_env(env_value(env, "TIKTOK_INCLUDE_VERSION", "true"), True)
        self.access_token_in_query = bool_env(env_value(env, "TIKTOK_ACCESS_TOKEN_IN_QUERY", "false"), False)
        self.sign_include_body = bool_env(env_value(env, "TIKTOK_SIGN_INCLUDE_BODY", "true"), True)
        missing = [
            key
            for key, value in [
                ("TIKTOK_APP_KEY", self.app_key),
                ("TIKTOK_APP_SECRET", self.app_secret),
                ("TIKTOK_ACCESS_TOKEN", self.access_token),
            ]
            if not value
        ]
        if missing:
            raise TikTokApiError(
                "Missing TikTok API values in "
                + str(env_path)
                + ": "
                + ", ".join(missing)
                + ". Fill them locally, never in chat."
            )

    def sign(self, path: str, params: dict[str, str], body_text: str = "", include_body: bool | None = None) -> str:
        include_body = self.sign_include_body if include_body is None else include_body
        sign_params = {
            key: str(value)
            for key, value in params.items()
            if key not in {"sign", "access_token"} and value is not None and value != ""
        }
        sign_input = path + "".join(f"{key}{sign_params[key]}" for key in sorted(sign_params))
        if include_body and body_text:
            sign_input += body_text
        wrapped = self.app_secret + sign_input + self.app_secret
        return hmac.new(self.app_secret.encode("utf-8"), wrapped.encode("utf-8"), hashlib.sha256).hexdigest()

    def request(
        self,
        method: str,
        path: str,
        params: dict[str, object] | None = None,
        body: dict | None = None,
        retry_alt_signature: bool = True,
    ) -> dict:
        query: dict[str, str] = {
            "app_key": self.app_key,
            "timestamp": str(int(time.time())),
        }
        if self.include_version:
            query["version"] = self.version
        for key, value in (params or {}).items():
            if value not in (None, ""):
                query[key] = str(value)
        if self.access_token_in_query:
            query["access_token"] = self.access_token

        body_text = compact_json(body) if body is not None else ""
        query["sign"] = self.sign(path, query, body_text)
        url = self.base_url + path + "?" + urllib.parse.urlencode(query)
        data = body_text.encode("utf-8") if body is not None else None
        request = urllib.request.Request(
            url,
            data=data,
            method=method.upper(),
            headers={
                "Content-Type": "application/json",
                "x-tts-access-token": self.access_token,
                "User-Agent": "TikTokShop-Libri-Automation/1.0",
            },
        )
        try:
            with urllib.request.urlopen(request, timeout=60) as response:
                payload = response.read().decode("utf-8", errors="replace")
        except urllib.error.HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace")
            if retry_alt_signature and body_text and ("sign" in detail.casefold() or "signature" in detail.casefold()):
                return self._retry_with_alternate_signature(method, path, params or {}, body or {})
            raise TikTokApiError(f"TikTok API HTTP {exc.code}: {detail[:1000]}") from exc
        except urllib.error.URLError as exc:
            raise TikTokApiError(f"TikTok API connection failed: {exc}") from exc

        try:
            result = json.loads(payload)
        except json.JSONDecodeError as exc:
            raise TikTokApiError(f"TikTok API returned non-JSON: {payload[:1000]}") from exc
        code = clean(result.get("code"))
        message = clean(result.get("message") or result.get("msg"))
        if code and code not in {"0", "OK", "SUCCESS"}:
            if retry_alt_signature and body_text and ("sign" in message.casefold() or "signature" in message.casefold()):
                return self._retry_with_alternate_signature(method, path, params or {}, body or {})
            raise TikTokApiError(f"TikTok API error {code}: {message or result}")
        return result

    def _retry_with_alternate_signature(self, method: str, path: str, params: dict[str, object], body: dict) -> dict:
        old = self.sign_include_body
        self.sign_include_body = not old
        try:
            return self.request(method, path, params, body, retry_alt_signature=False)
        finally:
            self.sign_include_body = old

    def ensure_shop_cipher(self) -> str:
        if self.shop_cipher:
            return self.shop_cipher
        response = self.request("GET", "/authorization/202309/shops")
        shops = list(nested_value(response, "data.shops", "data.authorized_shops") or [])
        if len(shops) == 1:
            cipher = clean(nested_value(shops[0], "cipher", "shop_cipher"))
            if cipher:
                self.shop_cipher = cipher
                return cipher
        if not shops:
            raise TikTokApiError("No authorized TikTok shops found for this app token.")
        names = [clean(shop.get("name") or shop.get("shop_name") or shop.get("id")) for shop in shops]
        raise TikTokApiError(
            "Multiple authorized TikTok shops found. Set TIKTOK_SHOP_CIPHER in .env. Shops: " + ", ".join(names)
        )

    def search_awaiting_orders(self, order_status: str, page_size: int, hours_back: int) -> list[dict]:
        shop_cipher = self.ensure_shop_cipher()
        all_orders: list[dict] = []
        page_token = ""
        body: dict[str, object] = {}
        if order_status:
            body["order_status"] = order_status
        if hours_back > 0:
            since = dt.datetime.now(dt.timezone.utc) - dt.timedelta(hours=hours_back)
            body["update_time_ge"] = int(since.timestamp())
        while True:
            params: dict[str, object] = {"shop_cipher": shop_cipher, "page_size": page_size}
            if page_token:
                params["page_token"] = page_token
            response = self.request("POST", "/order/202309/orders/search", params=params, body=body)
            data = response.get("data") or {}
            orders = data.get("orders") or data.get("order_list") or []
            all_orders.extend(orders)
            page_token = clean(data.get("next_page_token") or data.get("page_token"))
            more = data.get("more")
            if not page_token or more is False:
                break
        return all_orders

    def get_order_details(self, order_ids: list[str]) -> list[dict]:
        shop_cipher = self.ensure_shop_cipher()
        details: list[dict] = []
        for start in range(0, len(order_ids), 50):
            chunk = order_ids[start : start + 50]
            if not chunk:
                continue
            response = self.request(
                "GET",
                "/order/202309/orders",
                params={"shop_cipher": shop_cipher, "ids": ",".join(chunk)},
                body=None,
            )
            data = response.get("data") or {}
            details.extend(data.get("orders") or data.get("order_list") or [])
        return details


def customer_reference(order: AutomationOrder) -> str:
    base = order.buyer_username or order.address.name or order.order_id
    return (base + "; tiktokshop")[:80]


def row_address(row: dict[str, str]) -> CustomerAddress:
    street = first_value(row, "Street Name", "Street", "Address Line 1")
    house = first_value(row, "House Name or Number", "House Number", "Address Line 2")
    full = " ".join(part for part in [street, house, first_value(row, "Zipcode", "Zip"), first_value(row, "City")] if part)
    return CustomerAddress(
        name=first_value(row, "Recipient", "Name"),
        phone=first_value(row, "Phone #", "Phone", "Phone Number"),
        email=first_value(row, "Email"),
        country=first_value(row, "Country"),
        state=first_value(row, "State"),
        district=first_value(row, "District"),
        city=first_value(row, "City"),
        zipcode=first_value(row, "Zipcode", "Zip"),
        street=street,
        house=house,
        full_address=full,
        raw=dict(row),
    )


def read_tiktok_export_csv(path: Path) -> list[AutomationOrder]:
    if str(path).casefold() == "latest":
        path = latest_ready_order_csv()
    text = path.read_text(encoding="utf-8-sig", errors="replace")
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    rows = list(csv.DictReader(text.splitlines(), dialect=dialect))
    grouped: dict[str, list[dict[str, str]]] = {}
    for row in rows:
        order_id = normalize_order_id(row.get("Order ID"))
        if order_id:
            grouped.setdefault(order_id, []).append(row)

    orders: list[AutomationOrder] = []
    for order_id, order_rows in grouped.items():
        first = order_rows[0]
        address = row_address(first)
        lines: list[OrderLine] = []
        warnings: list[str] = []
        for row in order_rows:
            ean = extract_ean(row.get("Seller SKU"), row.get("Product Name"), row.get("Seller Note"))
            line_warnings = []
            if not ean:
                line_warnings.append("missing_ean_from_seller_sku_or_product_name")
            qty = parse_int(row.get("Quantity"), 1)
            if qty <= 0:
                line_warnings.append("quantity_zero")
            lines.append(
                OrderLine(
                    ean=ean,
                    seller_sku=clean(row.get("Seller SKU")),
                    product_name=clean(row.get("Product Name")),
                    quantity=max(qty, 1),
                    sku_id=clean(row.get("SKU ID")),
                    warnings=line_warnings,
                )
            )
        status = clean(first.get("Order Status"))
        normalized_status = status.upper().replace(" ", "_")
        if normalized_status not in TIKTOK_STATUS_READY:
            warnings.append(f"order_status_is_{status or 'empty'}")
        if not address.name or not address.city or not address.zipcode:
            warnings.append("customer_address_incomplete")
        orders.append(
            AutomationOrder(
                order_id=order_id,
                package_id=normalize_order_id(first.get("Package ID")),
                status=status,
                fulfillment_type=clean(first.get("Fulfillment Type")),
                delivery_option=clean(first.get("Delivery Option")),
                buyer_username=clean(first.get("Buyer Username")),
                created_time=clean(first.get("Created Time")),
                paid_time=clean(first.get("Paid Time")),
                address=address,
                lines=lines,
                source=str(path),
                raw={"rows": order_rows},
                warnings=warnings,
            )
        )
    return orders


def normalize_api_address(order: dict) -> CustomerAddress:
    address = nested_value(order, "recipient_address", "shipping_address", "delivery_address") or {}
    if not isinstance(address, dict):
        address = {}
    first = clean(address.get("first_name"))
    last = clean(address.get("last_name"))
    lines = [
        clean(address.get(key))
        for key in ["address_line1", "address_line2", "address_line3", "address_line4"]
        if clean(address.get(key))
    ]
    name = first_value(address, "name", "full_name", "recipient_name")
    if not name and (first or last):
        name = f"{first} {last}".strip()
    city = first_value(address, "city", "city_name")
    zipcode = first_value(address, "postal_code", "zipcode", "zip_code")
    full_address = first_value(address, "full_address", "address")
    if not full_address:
        full_address = " ".join(part for part in [*lines, zipcode, city] if part)
    return CustomerAddress(
        name=name,
        phone=first_value(address, "phone_number", "phone", "mobile"),
        email=first_value(address, "email"),
        country=first_value(address, "region_code", "country", "country_code"),
        state=first_value(address, "state", "state_name", "province"),
        district=first_value(address, "district", "district_info", "district_name"),
        city=city,
        zipcode=zipcode,
        street=lines[0] if lines else "",
        house=lines[1] if len(lines) > 1 else "",
        full_address=full_address,
        raw=address,
    )


def normalize_api_order(order: dict, source: str) -> AutomationOrder:
    order_id = normalize_order_id(nested_value(order, "id", "order_id"))
    packages = nested_value(order, "packages", "package_list") or []
    package_id = ""
    if isinstance(packages, list) and packages:
        first_package = packages[0] if isinstance(packages[0], dict) else {}
        package_id = normalize_order_id(nested_value(first_package, "id", "package_id"))
    package_id = package_id or normalize_order_id(nested_value(order, "package_id"))
    raw_items = nested_value(order, "line_items", "items", "skus") or []
    if not isinstance(raw_items, list):
        raw_items = []
    lines: list[OrderLine] = []
    for item in raw_items:
        if not isinstance(item, dict):
            continue
        seller_sku = first_value(item, "seller_sku", "seller_sku_id", "sku_seller_id")
        product_name = first_value(item, "product_name", "product_title", "sku_name", "item_name")
        ean = extract_ean(seller_sku, product_name, item.get("seller_note"))
        warnings = []
        if not ean:
            warnings.append("missing_ean_from_seller_sku_or_product_name")
        lines.append(
            OrderLine(
                ean=ean,
                seller_sku=seller_sku,
                product_name=product_name,
                quantity=parse_int(nested_value(item, "quantity", "sku_quantity"), 1) or 1,
                sku_id=clean(nested_value(item, "sku_id", "id")),
                warnings=warnings,
            )
        )
    address = normalize_api_address(order)
    warnings = []
    if not lines:
        warnings.append("no_line_items_in_api_response")
    if not address.name or not address.city or not address.zipcode:
        warnings.append("customer_address_incomplete")
    return AutomationOrder(
        order_id=order_id,
        package_id=package_id,
        status=clean(nested_value(order, "status", "order_status")),
        fulfillment_type=clean(nested_value(order, "fulfillment_type", "delivery_type")),
        delivery_option=clean(nested_value(order, "delivery_option", "shipping_type")),
        buyer_username=clean(nested_value(order, "buyer_username", "buyer_user_name")),
        created_time=clean(nested_value(order, "create_time", "created_time")),
        paid_time=clean(nested_value(order, "paid_time", "payment.paid_time")),
        address=address,
        lines=lines,
        source=source,
        raw=order,
        warnings=warnings,
    )


def latest_ready_order_csv() -> Path:
    downloads = Path.home() / "Downloads"
    matches = sorted(downloads.glob("Versandbereit Bestellung*.csv"), key=lambda path: path.stat().st_mtime, reverse=True)
    if not matches:
        raise SystemExit("No 'Versandbereit Bestellung*.csv' found in Downloads. Pass --input-csv <path>.")
    return matches[0]


def load_state(path: Path) -> dict:
    if not path.exists():
        return {"prepared_orders": {}}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {"prepared_orders": {}}


def save_state(path: Path, state: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(state, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")


def write_import_xlsx(path: Path, order: AutomationOrder, customer_number: str) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Bestellung"
    sheet.append(LIBRI_IMPORT_HEADERS)
    widths = [16, 18, 72, 10, 22, 32]
    for col_idx, header in enumerate(LIBRI_IMPORT_HEADERS, start=1):
        cell = sheet.cell(1, col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center")
        sheet.column_dimensions[get_column_letter(col_idx)].width = widths[col_idx - 1]
    ref = customer_reference(order)
    for line in order.lines:
        if not line.ean:
            continue
        sheet.append([customer_number, line.ean, line.product_name[:240], line.quantity, "", ref])
    sheet.freeze_panes = "A2"
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def write_import_csv(path: Path, order: AutomationOrder, customer_number: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh, delimiter=";")
        writer.writerow(LIBRI_IMPORT_HEADERS)
        ref = customer_reference(order)
        for line in order.lines:
            if line.ean:
                writer.writerow([customer_number, line.ean, line.product_name, line.quantity, "", ref])


def write_address_csv(path: Path, order: AutomationOrder) -> None:
    fieldnames = [
        "order_id",
        "package_id",
        "name",
        "phone",
        "email",
        "country",
        "state",
        "district",
        "city",
        "zipcode",
        "street",
        "house",
        "full_address",
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        row = {key: getattr(order.address, key, "") for key in fieldnames if hasattr(order.address, key)}
        row["order_id"] = order.order_id
        row["package_id"] = order.package_id
        writer.writerow(row)


def write_order_readme(path: Path, order: AutomationOrder) -> None:
    warnings = sorted(set(order.warnings + [warning for line in order.lines for warning in line.warnings]))
    lines = [
        f"# TikTok Order {order.order_id}",
        "",
        "Prepared files:",
        "- libri_kundenbestellung_import.xlsx: upload/add items in Mein.Libri Auftragserfassung",
        "- kundenadresse.csv: customer address for Libri customer order step 2",
        "- tiktok_order.json: raw/normalized order snapshot for audit",
        "",
        "Libri flow:",
        "1. In Mein.Libri, use Kundenbestellung.",
        "2. Choose Direktversand zum Kunden in step 2.",
        "3. Copy the address from kundenadresse.csv.",
        "4. Submit only after the title, quantity, recipient, and delivery mode match.",
        "",
    ]
    if warnings:
        lines.extend(["Warnings:", *[f"- {warning}" for warning in warnings], ""])
    path.write_text("\n".join(lines), encoding="utf-8")


def write_summary(path: Path, orders: list[AutomationOrder], statuses: dict[str, str]) -> None:
    fieldnames = [
        "order_id",
        "package_id",
        "status",
        "fulfillment_type",
        "delivery_option",
        "line_count",
        "item_count",
        "recipient_city",
        "recipient_zipcode",
        "automation_status",
        "warnings",
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        for order in orders:
            warnings = sorted(set(order.warnings + [warning for line in order.lines for warning in line.warnings]))
            writer.writerow(
                {
                    "order_id": order.order_id,
                    "package_id": order.package_id,
                    "status": order.status,
                    "fulfillment_type": order.fulfillment_type,
                    "delivery_option": order.delivery_option,
                    "line_count": len(order.lines),
                    "item_count": sum(line.quantity for line in order.lines),
                    "recipient_city": order.address.city,
                    "recipient_zipcode": order.address.zipcode,
                    "automation_status": statuses.get(order.order_id, ""),
                    "warnings": " | ".join(warnings),
                }
            )


def prepare_orders(
    orders: list[AutomationOrder],
    output_root: Path,
    state_path: Path,
    customer_number: str,
    rebuild: bool,
    ignore_state: bool,
    skip_empty_runs: bool,
) -> tuple[Path | None, dict[str, str]]:
    state = load_state(state_path)
    prepared = state.setdefault("prepared_orders", {})
    statuses: dict[str, str] = {}
    orders_to_write: list[tuple[AutomationOrder, list[OrderLine]]] = []

    for order in orders:
        order_key = order.order_id
        if not order_key:
            statuses[order_key] = "skipped_missing_order_id"
            continue
        if not ignore_state and order_key in prepared and not rebuild:
            statuses[order_key] = "skipped_already_prepared"
            continue
        valid_lines = [line for line in order.lines if line.ean and line.quantity > 0]
        if not valid_lines:
            statuses[order_key] = "review_no_valid_ean_lines"
        elif order.warnings or any(line.warnings for line in order.lines):
            statuses[order_key] = "prepared_with_warnings"
        else:
            statuses[order_key] = "prepared"
        orders_to_write.append((order, valid_lines))

    if skip_empty_runs and not orders_to_write:
        return None, statuses

    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    run_dir = output_root / stamp
    run_dir.mkdir(parents=True, exist_ok=True)

    for order, valid_lines in orders_to_write:
        order_key = order.order_id
        order_dir = run_dir / safe_filename(order_key)
        order_dir.mkdir(parents=True, exist_ok=True)
        order_for_file = AutomationOrder(**{**asdict(order), "address": order.address, "lines": valid_lines})
        write_import_xlsx(order_dir / "libri_kundenbestellung_import.xlsx", order_for_file, customer_number)
        write_import_csv(order_dir / "libri_kundenbestellung_import.csv", order_for_file, customer_number)
        write_address_csv(order_dir / "kundenadresse.csv", order)
        write_order_readme(order_dir / "README.md", order)
        (order_dir / "tiktok_order.json").write_text(
            json.dumps(asdict(order), ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        prepared[order_key] = {
            "prepared_at": dt.datetime.now(dt.timezone.utc).isoformat(),
            "status": statuses[order_key],
            "run_dir": str(run_dir.resolve()),
            "package_id": order.package_id,
        }

    write_summary(run_dir / "orders_summary.csv", orders, statuses)
    save_state(state_path, state)
    return run_dir, statuses


def fetch_from_api(args: argparse.Namespace, env: dict[str, str], env_path: Path) -> list[AutomationOrder]:
    client = TikTokShopClient(env, env_path)
    order_status = args.order_status or env_value(env, "TIKTOK_ORDER_STATUS", "AWAITING_SHIPMENT")
    search_orders = client.search_awaiting_orders(order_status, args.page_size, args.hours_back)
    ids = [normalize_order_id(nested_value(order, "id", "order_id")) for order in search_orders]
    ids = [order_id for order_id in ids if order_id]
    details = client.get_order_details(list(dict.fromkeys(ids))) if ids else []
    if details:
        return [normalize_api_order(order, "tiktok_api") for order in details]
    return [normalize_api_order(order, "tiktok_api_search_only") for order in search_orders]


def run_once(args: argparse.Namespace) -> tuple[Path | None, dict[str, str], int]:
    env_path = Path(args.env)
    env = load_env_file(env_path)
    customer_number = env_value(env, "LIBRI_CUSTOMER_NUMBER")
    if args.input_csv:
        orders = read_tiktok_export_csv(Path(args.input_csv))
    else:
        orders = fetch_from_api(args, env, env_path)
    run_dir, statuses = prepare_orders(
        orders=orders,
        output_root=Path(args.output_root),
        state_path=Path(args.state),
        customer_number=customer_number,
        rebuild=args.rebuild,
        ignore_state=args.ignore_state,
        skip_empty_runs=args.skip_empty_runs,
    )

    # Auto-submit orders to Libri if requested
    if args.auto_submit_libri and run_dir is not None:
        auto_submit_prepared_orders(run_dir, statuses, env_path)

    return run_dir, statuses, len(orders)


def auto_submit_prepared_orders(run_dir: Path, statuses: dict[str, str], env_path: Path) -> None:
    """Automatically submit prepared orders to Libri."""
    submitted = 0
    failed = 0
    for order_id, status in statuses.items():
        if status.startswith("prepared"):
            order_dir = run_dir / safe_filename(order_id)
            if not order_dir.exists():
                print(f"⚠ Order directory not found: {order_dir}")
                failed += 1
                continue

            print(f"Auto-submitting order {order_id} to Libri...")
            script_path = Path(__file__).resolve().parent / "libri_customer_submit.py"
            try:
                result = subprocess.run(
                    [sys.executable, str(script_path), "--order-dir", str(order_dir), "--env", str(env_path)],
                    capture_output=True,
                    text=True,
                    timeout=120,
                )
                if result.returncode == 0:
                    print(f"✓ Order {order_id} submitted successfully")
                    submitted += 1
                else:
                    print(f"✗ Order {order_id} submission failed: {result.stderr}")
                    failed += 1
            except Exception as e:
                print(f"✗ Error submitting order {order_id}: {e}")
                failed += 1

    print(f"Libri submissions: {submitted} success, {failed} failed")


def seconds_until_run_at(run_at: str, timezone_name: str) -> float:
    tz = ZoneInfo(timezone_name)
    now = dt.datetime.now(tz)
    hour, minute = [int(part) for part in run_at.split(":", 1)]
    target = now.replace(hour=hour, minute=minute, second=0, microsecond=0)
    if target <= now:
        target += dt.timedelta(days=1)
    return max((target - now).total_seconds(), 1.0)


def watch(args: argparse.Namespace) -> int:
    while True:
        if args.run_at:
            wait_seconds = seconds_until_run_at(args.run_at, args.timezone)
            print(f"Waiting until next {args.run_at} {args.timezone} run ({int(wait_seconds)} seconds).", flush=True)
            time.sleep(wait_seconds)
        run_dir, statuses, order_count = run_once(args)
        prepared_count = sum(1 for status in statuses.values() if status.startswith("prepared"))
        output_text = str(run_dir.resolve()) if run_dir else "no new output"
        print(f"{dt.datetime.now().isoformat(timespec='seconds')} - found {order_count}, prepared {prepared_count}: {output_text}", flush=True)
        if not args.run_at:
            time.sleep(max(args.poll_minutes, 1) * 60)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Fetch TikTok Shop orders and prepare Libri customer-order files.")
    parser.add_argument("--env", default=".env")
    parser.add_argument(
        "--input-csv",
        default="",
        help="TikTok Seller Center ready-to-ship CSV, or 'latest'. If omitted, the TikTok API is used.",
    )
    parser.add_argument("--output-root", default=str(DEFAULT_OUTPUT_ROOT))
    parser.add_argument("--state", default=str(DEFAULT_STATE_PATH))
    parser.add_argument("--rebuild", action="store_true", help="Rebuild packages even if an order is already in state.")
    parser.add_argument("--ignore-state", action="store_true", help="Do not skip previously prepared orders.")
    parser.add_argument("--order-status", default="", help="Defaults to TIKTOK_ORDER_STATUS or AWAITING_SHIPMENT.")
    parser.add_argument("--hours-back", type=int, default=72)
    parser.add_argument("--page-size", type=int, default=50)
    parser.add_argument("--watch", action="store_true", help="Keep running. Use --run-at for daily mode or --poll-minutes.")
    parser.add_argument("--run-at", default="", help="Daily run time like 17:00. Only used with --watch.")
    parser.add_argument("--timezone", default=DEFAULT_TIMEZONE)
    parser.add_argument("--poll-minutes", type=int, default=15)
    parser.add_argument("--skip-empty-runs", action="store_true", help="Do not create an output folder when there are no new orders.")
    parser.add_argument(
        "--auto-submit-libri",
        action="store_true",
        help="Submit prepared orders to Libri after file generation. This places real Libri orders.",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    if args.watch:
        return watch(args)
    run_dir, statuses, order_count = run_once(args)
    prepared_count = sum(1 for status in statuses.values() if status.startswith("prepared"))
    review_count = sum(1 for status in statuses.values() if "review" in status or "warning" in status)
    skipped_count = sum(1 for status in statuses.values() if status.startswith("skipped"))
    print(f"Orders found: {order_count}")
    print(f"Prepared: {prepared_count}")
    print(f"Needs review/warnings: {review_count}")
    print(f"Skipped: {skipped_count}")
    if run_dir:
        print(f"Output: {run_dir.resolve()}")
    else:
        print("Output: none (no new orders)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
