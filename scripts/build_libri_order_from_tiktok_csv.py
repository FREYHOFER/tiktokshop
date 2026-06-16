#!/usr/bin/env python3
"""Create a Libri customer-order import file from a TikTok ready-to-ship CSV.

Default behavior is safe: parse, validate, and write local import files only.
Use --upload-to-libri explicitly to upload the generated workbook into Mein.Libri
Auftragserfassung. The script does not click/submit the final order checkout.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import html
import mimetypes
import os
import re
import sys
import urllib.parse
import urllib.request
import uuid
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from fetch_libri_product_pages import fetch, load_env_file, login  # noqa: E402


LIBRI_IMPORT_HEADERS = ["Kundennummer", "Artikel Nr", "Autor/Titel", "Menge", "Vormerkkennzeichen", "Bestellzeichen"]
REQUIRED_TIKTOK_COLUMNS = ["Order ID", "Order Status", "Seller SKU", "Product Name", "Quantity"]
MASKED_RE = re.compile(r"\*{2,}")
ORDER_PAGE_URL = "https://mein.libri.de/Bestellen/Auftragserfassung.html"


@dataclass
class ParsedOrderLine:
    order_id: str
    package_id: str
    order_status: str
    substatus: str
    seller_sku: str
    ean: str
    product_name: str
    quantity: int
    buyer_username: str
    recipient: str
    customer_reference: str
    masked_recipient: bool
    warnings: list[str]


def clean(value: object) -> str:
    return str(value or "").replace("\ufeff", "").strip()


def normalize_order_id(value: str) -> str:
    return re.sub(r"\s+", "", clean(value))


def extract_ean(row: dict[str, str]) -> str:
    candidates = [row.get("Seller SKU", ""), row.get("Product Name", ""), row.get("Seller Note", "")]
    for value in candidates:
        for match in re.findall(r"(?:97[89])[\d\-\s]{10,20}", value or ""):
            digits = re.sub(r"\D", "", match)
            if len(digits) == 13:
                return digits
    return ""


def parse_quantity(value: str) -> int:
    text = clean(value).replace(",", ".")
    try:
        qty = int(float(text))
    except ValueError:
        return 0
    return max(qty, 0)


def is_masked(value: str) -> bool:
    text = clean(value)
    return not text or bool(MASKED_RE.search(text))


def customer_reference(row: dict[str, str], mode: str) -> tuple[str, bool]:
    recipient = clean(row.get("Recipient", ""))
    buyer = clean(row.get("Buyer Username", ""))
    order_id = normalize_order_id(row.get("Order ID", ""))
    masked = is_masked(recipient)
    if mode == "recipient" and not masked:
        name = recipient
    elif mode == "recipient":
        name = buyer or order_id
    elif mode == "buyer_username":
        name = buyer or (recipient if not masked else "") or order_id
    elif mode == "order_id":
        name = order_id
    else:
        name = recipient if recipient and not masked else buyer or order_id
    reference = f"{name}; tiktokshop" if name else "tiktokshop"
    return reference[:80], masked


def read_tiktok_csv(path: Path, reference_mode: str) -> tuple[list[ParsedOrderLine], list[str]]:
    text = path.read_text(encoding="utf-8-sig", errors="replace")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    rows = list(csv.DictReader(text.splitlines(), dialect=dialect))
    if not rows:
        return [], ["CSV enthält keine Bestellzeilen."]
    missing = [col for col in REQUIRED_TIKTOK_COLUMNS if col not in rows[0]]
    if missing:
        return [], ["CSV-Spalten fehlen: " + ", ".join(missing)]

    parsed: list[ParsedOrderLine] = []
    global_warnings: list[str] = []
    for row in rows:
        warnings: list[str] = []
        status = clean(row.get("Order Status", ""))
        substatus = clean(row.get("Order Substatus", ""))
        if status and status.casefold() != "versandbereit":
            warnings.append(f"Status ist {status!r}, nicht 'Versandbereit'.")
        ean = extract_ean(row)
        if not ean:
            warnings.append("Keine ISBN/EAN gefunden.")
        qty = parse_quantity(row.get("Quantity", ""))
        if qty <= 0:
            warnings.append("Menge ist leer oder 0.")
        ref, masked = customer_reference(row, reference_mode)
        if masked:
            warnings.append("Empfängername ist im TikTok-Export maskiert; Referenz nutzt Fallback.")
        parsed.append(
            ParsedOrderLine(
                order_id=normalize_order_id(row.get("Order ID", "")),
                package_id=normalize_order_id(row.get("Package ID", "")),
                order_status=status,
                substatus=substatus,
                seller_sku=clean(row.get("Seller SKU", "")),
                ean=ean,
                product_name=clean(row.get("Product Name", "")),
                quantity=qty,
                buyer_username=clean(row.get("Buyer Username", "")),
                recipient=clean(row.get("Recipient", "")),
                customer_reference=ref,
                masked_recipient=masked,
                warnings=warnings,
            )
        )
    duplicate_keys = {}
    for line in parsed:
        key = (line.order_id, line.ean, line.customer_reference)
        duplicate_keys[key] = duplicate_keys.get(key, 0) + 1
    if any(count > 1 for count in duplicate_keys.values()):
        global_warnings.append("Hinweis: doppelte Order/EAN-Zeilen gefunden; sie werden in der Libri-Datei aggregiert.")
    return parsed, global_warnings


def aggregate_lines(lines: list[ParsedOrderLine]) -> list[ParsedOrderLine]:
    grouped: dict[tuple[str, str, str], ParsedOrderLine] = {}
    for line in lines:
        key = (line.order_id, line.ean, line.customer_reference)
        if key not in grouped:
            grouped[key] = line
        else:
            grouped[key].quantity += line.quantity
            grouped[key].warnings = sorted(set(grouped[key].warnings + line.warnings + ["Mehrere TikTok-Zeilen aggregiert."]))
    return list(grouped.values())


def load_customer_number(env_path: Path) -> str:
    return clean(load_env_file(env_path).get("LIBRI_CUSTOMER_NUMBER", ""))


def write_import_xlsx(path: Path, lines: list[ParsedOrderLine], customer_number: str) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Bestellung"
    sheet.append(LIBRI_IMPORT_HEADERS)
    for col_idx, header in enumerate(LIBRI_IMPORT_HEADERS, start=1):
        cell = sheet.cell(1, col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center")
        sheet.column_dimensions[get_column_letter(col_idx)].width = [16, 18, 72, 10, 22, 32][col_idx - 1]
    for line in lines:
        sheet.append([customer_number, line.ean, line.product_name[:240], line.quantity, "", line.customer_reference])
    sheet.freeze_panes = "A2"
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def write_import_csv(path: Path, lines: list[ParsedOrderLine], customer_number: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh, delimiter=";")
        writer.writerow(LIBRI_IMPORT_HEADERS)
        for line in lines:
            writer.writerow([customer_number, line.ean, line.product_name, line.quantity, "", line.customer_reference])


def write_mapping(path: Path, lines: list[ParsedOrderLine], global_warnings: list[str]) -> None:
    fieldnames = [
        "order_id",
        "package_id",
        "order_status",
        "order_substatus",
        "seller_sku",
        "ean",
        "quantity",
        "buyer_username",
        "recipient",
        "masked_recipient",
        "customer_reference",
        "product_name",
        "warnings",
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        for warning in global_warnings:
            writer.writerow({"warnings": warning})
        for line in lines:
            writer.writerow(
                {
                    "order_id": line.order_id,
                    "package_id": line.package_id,
                    "order_status": line.order_status,
                    "order_substatus": line.substatus,
                    "seller_sku": line.seller_sku,
                    "ean": line.ean,
                    "quantity": line.quantity,
                    "buyer_username": line.buyer_username,
                    "recipient": line.recipient,
                    "masked_recipient": "yes" if line.masked_recipient else "no",
                    "customer_reference": line.customer_reference,
                    "product_name": line.product_name,
                    "warnings": " | ".join(line.warnings),
                }
            )


def multipart_upload_body(fields: dict[str, str], file_field: str, file_path: Path) -> tuple[bytes, str]:
    boundary = f"----codex-libri-{uuid.uuid4().hex}"
    chunks: list[bytes] = []
    for key, value in fields.items():
        chunks.append(f"--{boundary}\r\n".encode("utf-8"))
        chunks.append(f'Content-Disposition: form-data; name="{key}"\r\n\r\n'.encode("utf-8"))
        chunks.append(value.encode("utf-8"))
        chunks.append(b"\r\n")
    content_type = mimetypes.guess_type(file_path.name)[0] or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    chunks.append(f"--{boundary}\r\n".encode("utf-8"))
    chunks.append(
        (
            f'Content-Disposition: form-data; name="{file_field}"; filename="{file_path.name}"\r\n'
            f"Content-Type: {content_type}\r\n\r\n"
        ).encode("utf-8")
    )
    chunks.append(file_path.read_bytes())
    chunks.append(b"\r\n")
    chunks.append(f"--{boundary}--\r\n".encode("utf-8"))
    return b"".join(chunks), f"multipart/form-data; boundary={boundary}"


def upload_to_libri(env_path: Path, import_path: Path, output_dir: Path) -> Path:
    opener = login(env_path)
    page_url, page_html = fetch(opener, ORDER_PAGE_URL)
    token_match = re.search(r'name="cmsauthenticitytoken"\s+value="([^"]+)"', page_html)
    if not token_match:
        raise RuntimeError("Kein Libri-cmsauthenticitytoken auf der Auftragserfassungsseite gefunden.")
    fields = {
        "module_fnc[primary]": "AddEanListToBasketFromExcel",
        "cmsauthenticitytoken": html.unescape(token_match.group(1)),
    }
    body, content_type = multipart_upload_body(fields, "excelFile", import_path)
    request = urllib.request.Request(
        page_url,
        data=body,
        headers={"User-Agent": "Mozilla/5.0", "Content-Type": content_type},
        method="POST",
    )
    with opener.open(request, timeout=60) as response:
        response_html = response.read().decode("utf-8", errors="replace")
    response_path = output_dir / "libri_upload_response.html"
    response_path.write_text(response_html, encoding="utf-8")
    return response_path


def upload_text_to_libri(env_path: Path, lines: list[ParsedOrderLine], output_dir: Path) -> Path:
    opener = login(env_path)
    page_url, page_html = fetch(opener, ORDER_PAGE_URL)
    token_match = re.search(r'name="cmsauthenticitytoken"\s+value="([^"]+)"', page_html)
    if not token_match:
        raise RuntimeError("Kein Libri-cmsauthenticitytoken auf der Auftragserfassungsseite gefunden.")

    ean_lines: list[str] = []
    for line in lines:
        for _ in range(max(line.quantity, 1)):
            ean_lines.append(line.ean)
    payload = {
        "module_fnc[primary]": "AddEanListToBasketFromTextField",
        "eanList": "\n".join(ean_lines),
        "cmsauthenticitytoken": html.unescape(token_match.group(1)),
    }
    _, response_html = fetch(opener, page_url, payload)
    response_path = output_dir / "libri_text_upload_response.html"
    response_path.write_text(response_html, encoding="utf-8")
    return response_path


def default_output_dir() -> Path:
    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    return Path("outputs") / "libri_order_imports" / stamp


def latest_ready_order_csv() -> Path:
    downloads = Path.home() / "Downloads"
    matches = sorted(downloads.glob("Versandbereit Bestellung*.csv"), key=lambda path: path.stat().st_mtime, reverse=True)
    if not matches:
        raise SystemExit("Keine 'Versandbereit Bestellung*.csv' im Downloads-Ordner gefunden. Bitte --input angeben.")
    return matches[0]


def build(args: argparse.Namespace) -> int:
    input_path = Path(args.input) if args.input else latest_ready_order_csv()
    output_dir = Path(args.output_dir) if args.output_dir else default_output_dir()
    output_dir.mkdir(parents=True, exist_ok=True)
    lines, global_warnings = read_tiktok_csv(input_path, args.reference)
    lines = aggregate_lines(lines)
    valid_lines = [line for line in lines if line.ean and line.quantity > 0]
    customer_number = "" if args.blank_customer_number else load_customer_number(Path(args.env))

    xlsx_path = output_dir / "libri_kundenbestellung_import.xlsx"
    csv_path = output_dir / "libri_kundenbestellung_import.csv"
    mapping_path = output_dir / "tiktok_to_libri_order_mapping.csv"
    write_import_xlsx(xlsx_path, valid_lines, customer_number)
    write_import_csv(csv_path, valid_lines, customer_number)
    write_mapping(mapping_path, lines, global_warnings)

    response_path = None
    if args.upload_to_libri:
        if args.upload_mode == "text":
            response_path = upload_text_to_libri(Path(args.env), valid_lines, output_dir)
        else:
            response_path = upload_to_libri(Path(args.env), xlsx_path, output_dir)

    print(f"Input: {input_path.resolve()}")
    print(f"TikTok-Zeilen: {len(lines)}")
    print(f"Libri-Importzeilen: {len(valid_lines)}")
    print(f"XLSX: {xlsx_path.resolve()}")
    print(f"CSV: {csv_path.resolve()}")
    print(f"Mapping: {mapping_path.resolve()}")
    if response_path:
        print(f"Libri upload response: {response_path.resolve()}")
        print("Hinweis: Upload zur Auftragserfassung ist erfolgt; finaler Bestellabschluss wird von diesem Script nicht geklickt.")
    if any(line.warnings for line in lines) or global_warnings:
        print("Warnungen im Mapping prüfen.")
    return 0 if valid_lines else 1


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Build a Libri customer-order import from TikTok ready-to-ship CSV.")
    parser.add_argument("--input", default="", help="TikTok CSV export. Defaults to newest 'Versandbereit Bestellung*.csv' in Downloads.")
    parser.add_argument("--output-dir", default="")
    parser.add_argument("--env", default=".env")
    parser.add_argument(
        "--reference",
        choices=["recipient_then_username", "recipient", "buyer_username", "order_id"],
        default="recipient_then_username",
        help="Bestellzeichen source before '; tiktokshop'. Masked recipients automatically fall back unless order_id is chosen.",
    )
    parser.add_argument("--blank-customer-number", action="store_true", help="Leave Kundennummer blank in the Libri import file.")
    parser.add_argument(
        "--upload-to-libri",
        action="store_true",
        help="Upload to Mein.Libri Auftragserfassung. This can add items to the Libri basket; it does not finalize checkout.",
    )
    parser.add_argument(
        "--upload-mode",
        choices=["excel", "text"],
        default="excel",
        help="Libri upload route. 'excel' uses the generated workbook; 'text' posts the EAN list field.",
    )
    return parser


def main() -> int:
    return build(build_parser().parse_args())


if __name__ == "__main__":
    raise SystemExit(main())
