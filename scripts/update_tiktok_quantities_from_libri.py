#!/usr/bin/env python3
"""Sync TikTok Shop SKU quantities from current Mein.Libri stock.

The job treats Libri as the stock source of truth:
1. find TikTok Shop SKUs that use the local LIBRI-{EAN} seller SKU convention,
2. refresh each matching Mein.Libri product detail page,
3. parse Libri "Bestand",
4. update TikTok Shop inventory for the matching TikTok SKU.

All API writes are logged to outputs/inventory_updates/<timestamp>/.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import os
import re
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from fetch_libri_product_pages import PRODUCT_URL, fetch, load_env_file, login  # noqa: E402
from tiktok_libri_pipeline import parse_libri_detail_html  # noqa: E402
from tiktok_order_automation import TikTokShopClient, clean, compact_json  # noqa: E402


DEFAULT_OUTPUT_ROOT = Path("outputs") / "inventory_updates"
DEFAULT_WAREHOUSE_ID = "7630853807695791895"
DEFAULT_SELLER_SKU_PREFIX = "LIBRI-"
DEFAULT_UPLOAD_LOG_STATUSES = {"created", "skipped_existing_sku"}
DEFAULT_PRODUCT_STATUSES = "ACTIVATE,SELLER_DEACTIVATED,DRAFT,PENDING,FAILED"
LOG_FIELDS = [
    "timestamp_utc",
    "ean",
    "seller_sku",
    "product_id",
    "sku_id",
    "warehouse_id",
    "tiktok_quantity_before",
    "libri_quantity",
    "quantity_sent",
    "status",
    "message",
    "libri_page",
    "product_status",
    "title",
]


@dataclass
class TikTokSku:
    seller_sku: str
    ean: str
    product_id: str
    sku_id: str
    warehouse_id: str
    current_quantity: int | None = None
    title: str = ""
    product_status: str = ""


@dataclass
class LibriStock:
    ean: str
    quantity: int | None
    page_path: Path
    title: str = ""
    message: str = ""


class TikTokInventoryClient(TikTokShopClient):
    def search_products(
        self,
        seller_skus: list[str] | None = None,
        page_size: int = 50,
        max_pages: int = 0,
    ) -> list[dict[str, Any]]:
        products: list[dict[str, Any]] = []
        page_token = ""
        page_count = 0
        body: dict[str, Any] = {}
        if seller_skus:
            body["seller_skus"] = seller_skus

        while True:
            params: dict[str, Any] = {
                "shop_cipher": self.ensure_shop_cipher(),
                "page_size": page_size,
            }
            if page_token:
                params["page_token"] = page_token
            response = self.request("POST", "/product/202309/products/search", params=params, body=body)
            data = response.get("data") or {}
            page_products = data.get("products") or data.get("product_list") or []
            if isinstance(page_products, list):
                products.extend(product for product in page_products if isinstance(product, dict))
            page_count += 1
            page_token = clean(data.get("next_page_token") or data.get("page_token"))
            more = data.get("more")
            if not page_token or more is False:
                break
            if max_pages and page_count >= max_pages:
                break
        return products

    def update_inventory(self, product_id: str, sku_id: str, warehouse_id: str, quantity: int) -> dict[str, Any]:
        path = f"/product/202309/products/{product_id}/inventory/update"
        return self.request(
            "POST",
            path,
            params={"shop_cipher": self.ensure_shop_cipher()},
            body={
                "skus": [
                    {
                        "id": sku_id,
                        "inventory": [
                            {
                                "warehouse_id": warehouse_id,
                                "quantity": quantity,
                            }
                        ],
                    }
                ]
            },
        )


def now_utc() -> str:
    return dt.datetime.now(dt.timezone.utc).isoformat(timespec="seconds")


def csv_values(value: str) -> set[str]:
    return {part.strip() for part in clean(value).split(",") if part.strip()}


def extract_ean(value: object, seller_sku_prefix: str = DEFAULT_SELLER_SKU_PREFIX) -> str:
    text = clean(value)
    if seller_sku_prefix and text.casefold().startswith(seller_sku_prefix.casefold()):
        text = text[len(seller_sku_prefix) :]
    for match in re.findall(r"(?:97[89])[\d\-\s]{10,20}", text):
        digits = re.sub(r"\D", "", match)
        if len(digits) == 13:
            return digits
    digits = re.sub(r"\D", "", text)
    return digits if len(digits) == 13 and digits.startswith(("978", "979")) else ""


def seller_sku_for_ean(ean: str, seller_sku_prefix: str) -> str:
    return f"{seller_sku_prefix}{ean}"


def dedupe(values: Iterable[str]) -> list[str]:
    return [value for value in dict.fromkeys(clean(value) for value in values if clean(value))]


def sniff_csv(text: str) -> csv.Dialect:
    try:
        return csv.Sniffer().sniff(text[:4096], delimiters=",;\t")
    except csv.Error:
        return csv.excel


def read_upload_log(path: Path, seller_sku_prefix: str, statuses: set[str]) -> list[str]:
    text = path.read_text(encoding="utf-8-sig", errors="replace")
    rows = csv.DictReader(text.splitlines(), dialect=sniff_csv(text))
    seller_skus: list[str] = []
    for row in rows:
        status = clean(row.get("status"))
        if statuses and status not in statuses:
            continue
        seller_sku = clean(row.get("seller_sku"))
        ean = extract_ean(row.get("ean") or seller_sku, seller_sku_prefix)
        if not seller_sku and ean:
            seller_sku = seller_sku_for_ean(ean, seller_sku_prefix)
        if seller_sku:
            seller_skus.append(seller_sku)
    return dedupe(seller_skus)


def read_sku_csv(path: Path, seller_sku_prefix: str) -> list[str]:
    text = path.read_text(encoding="utf-8-sig", errors="replace")
    rows = csv.DictReader(text.splitlines(), dialect=sniff_csv(text))
    seller_skus: list[str] = []
    for row in rows:
        seller_sku = clean(row.get("seller_sku") or row.get("Seller SKU") or row.get("sku") or row.get("SKU"))
        ean = extract_ean(row.get("ean") or row.get("isbn") or row.get("gtin_code") or seller_sku, seller_sku_prefix)
        if not seller_sku and ean:
            seller_sku = seller_sku_for_ean(ean, seller_sku_prefix)
        if seller_sku:
            seller_skus.append(seller_sku)
    return dedupe(seller_skus)


def workbook_headers(sheet) -> dict[str, int]:
    return {
        clean(sheet.cell(row=1, column=col).value): col
        for col in range(1, sheet.max_column + 1)
        if clean(sheet.cell(row=1, column=col).value)
    }


def read_workbook(path: Path, seller_sku_prefix: str, start_row: int) -> list[str]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Template"] if "Template" in workbook.sheetnames else workbook.active
    headers = workbook_headers(sheet)
    seller_sku_col = headers.get("seller_sku")
    gtin_col = headers.get("gtin_code")
    seller_skus: list[str] = []
    for row_idx in range(start_row, min(sheet.max_row, 5000) + 1):
        seller_sku = clean(sheet.cell(row=row_idx, column=seller_sku_col).value) if seller_sku_col else ""
        ean = ""
        if gtin_col:
            ean = extract_ean(sheet.cell(row=row_idx, column=gtin_col).value, seller_sku_prefix)
        ean = ean or extract_ean(seller_sku, seller_sku_prefix)
        if not seller_sku and ean:
            seller_sku = seller_sku_for_ean(ean, seller_sku_prefix)
        if seller_sku:
            seller_skus.append(seller_sku)
    workbook.close()
    return dedupe(seller_skus)


def collect_requested_seller_skus(args: argparse.Namespace) -> list[str]:
    seller_skus: list[str] = []
    seller_skus.extend(args.sku)
    seller_skus.extend(seller_sku_for_ean(extract_ean(ean, args.seller_sku_prefix), args.seller_sku_prefix) for ean in args.ean)

    upload_log_statuses = csv_values(args.source_statuses)
    for raw_path in args.upload_log:
        seller_skus.extend(read_upload_log(Path(raw_path), args.seller_sku_prefix, upload_log_statuses))
    for raw_path in args.sku_csv:
        seller_skus.extend(read_sku_csv(Path(raw_path), args.seller_sku_prefix))
    for raw_path in args.workbook:
        seller_skus.extend(read_workbook(Path(raw_path), args.seller_sku_prefix, args.start_row))

    cleaned = []
    for seller_sku in seller_skus:
        seller_sku = clean(seller_sku)
        ean = extract_ean(seller_sku, args.seller_sku_prefix)
        if seller_sku and ean:
            cleaned.append(seller_sku)
    return dedupe(cleaned)


def first_int(*values: object) -> int | None:
    for value in values:
        text = clean(value).replace(",", ".")
        if not text:
            continue
        try:
            return max(int(float(text)), 0)
        except ValueError:
            continue
    return None


def inventory_entries(sku: dict[str, Any]) -> list[dict[str, Any]]:
    for key in ["inventory", "inventories", "stock_infos", "warehouse_stock"]:
        value = sku.get(key)
        if isinstance(value, list):
            return [entry for entry in value if isinstance(entry, dict)]
    return []


def inventory_for_sku(sku: dict[str, Any], preferred_warehouse_id: str) -> tuple[str, int | None]:
    entries = inventory_entries(sku)
    chosen: dict[str, Any] | None = None
    if preferred_warehouse_id:
        chosen = next(
            (
                entry
                for entry in entries
                if clean(entry.get("warehouse_id") or entry.get("id") or nested_dict_value(entry, "warehouse", "id"))
                == preferred_warehouse_id
            ),
            None,
        )
    if chosen is None and entries:
        chosen = entries[0]
    if chosen is None:
        return preferred_warehouse_id, None

    warehouse_id = clean(chosen.get("warehouse_id") or chosen.get("id") or nested_dict_value(chosen, "warehouse", "id"))
    quantity = first_int(
        chosen.get("quantity"),
        chosen.get("available_quantity"),
        chosen.get("available_stock"),
        chosen.get("stock"),
    )
    return warehouse_id or preferred_warehouse_id, quantity


def nested_dict_value(source: dict[str, Any], parent_key: str, child_key: str) -> object:
    value = source.get(parent_key)
    if isinstance(value, dict):
        return value.get(child_key)
    return ""


def product_skus(product: dict[str, Any], seller_sku_prefix: str, default_warehouse_id: str) -> list[TikTokSku]:
    product_id = clean(product.get("id") or product.get("product_id"))
    title = clean(product.get("title") or product.get("name") or product.get("product_name"))
    product_status = clean(product.get("status") or product.get("product_status"))
    raw_skus = product.get("skus") or product.get("sku_list") or []
    if not isinstance(raw_skus, list):
        return []

    output: list[TikTokSku] = []
    for sku in raw_skus:
        if not isinstance(sku, dict):
            continue
        seller_sku = clean(sku.get("seller_sku") or sku.get("seller_sku_id") or sku.get("external_sku_id"))
        if seller_sku_prefix and not seller_sku.casefold().startswith(seller_sku_prefix.casefold()):
            continue
        ean = extract_ean(seller_sku, seller_sku_prefix)
        sku_id = clean(sku.get("id") or sku.get("sku_id"))
        warehouse_id, current_quantity = inventory_for_sku(sku, default_warehouse_id)
        if seller_sku and ean:
            output.append(
                TikTokSku(
                    seller_sku=seller_sku,
                    ean=ean,
                    product_id=product_id,
                    sku_id=sku_id,
                    warehouse_id=warehouse_id,
                    current_quantity=current_quantity,
                    title=title,
                    product_status=product_status,
                )
            )
    return output


def resolve_tiktok_skus(
    client: TikTokInventoryClient,
    args: argparse.Namespace,
    requested_seller_skus: list[str],
    default_warehouse_id: str,
) -> tuple[list[TikTokSku], list[dict[str, Any]]]:
    found: dict[str, TikTokSku] = {}
    log_rows: list[dict[str, Any]] = []
    allowed_statuses = csv_values(args.product_statuses)

    if requested_seller_skus:
        batch_size = min(max(args.search_batch_size, 1), 10)
        for start in range(0, len(requested_seller_skus), batch_size):
            batch = requested_seller_skus[start : start + batch_size]
            products = client.search_products(seller_skus=batch, page_size=max(args.page_size, len(batch)))
            for product in products:
                for sku in product_skus(product, args.seller_sku_prefix, default_warehouse_id):
                    found.setdefault(sku.seller_sku, sku)
            if args.tiktok_delay:
                time.sleep(args.tiktok_delay)

        for seller_sku in requested_seller_skus:
            if seller_sku not in found:
                log_rows.append(
                    base_log_row(
                        TikTokSku(
                            seller_sku=seller_sku,
                            ean=extract_ean(seller_sku, args.seller_sku_prefix),
                            product_id="",
                            sku_id="",
                            warehouse_id=default_warehouse_id,
                        ),
                        status="skipped_tiktok_sku_not_found",
                        message="Seller SKU was not found by TikTok product search.",
                    )
                )
    elif args.discover_tiktok_products:
        products = client.search_products(page_size=args.page_size, max_pages=args.max_pages)
        for product in products:
            for sku in product_skus(product, args.seller_sku_prefix, default_warehouse_id):
                found.setdefault(sku.seller_sku, sku)
    else:
        raise SystemExit("No source SKUs supplied. Use --discover-tiktok-products or pass --ean/--sku/--workbook/--upload-log.")

    resolved = list(found.values())
    if allowed_statuses:
        kept: list[TikTokSku] = []
        for sku in resolved:
            if sku.product_status in allowed_statuses:
                kept.append(sku)
            else:
                log_rows.append(
                    base_log_row(
                        sku,
                        status="skipped_product_status",
                        message=f"Product status {sku.product_status or '<empty>'} is not in --product-statuses.",
                    )
                )
        resolved = kept

    resolved.sort(key=lambda sku: sku.seller_sku)
    if args.limit:
        resolved = resolved[: args.limit]
    return resolved, log_rows


def local_libri_page(ean: str, dirs: list[str]) -> Path | None:
    for raw_dir in dirs:
        path = Path(raw_dir) / f"{ean}.html"
        if path.exists() and path.stat().st_size > 0:
            return path
    return None


def fetch_libri_stock(ean: str, opener, output_dir: Path) -> LibriStock:
    output_dir.mkdir(parents=True, exist_ok=True)
    final_url, body = fetch(opener, PRODUCT_URL.format(ean=ean))
    if "Login.html" in final_url or "<title>Mein.Libri - Login</title>" in body:
        raise RuntimeError("Libri redirected to login while fetching product page.")
    page_path = output_dir / f"{ean}.html"
    page_path.write_text(body, encoding="utf-8")
    product = parse_libri_detail_html(page_path)
    message = ""
    if product.ean and product.ean != ean:
        message = f"Libri page EAN {product.ean} did not match requested EAN {ean}."
    return LibriStock(ean=ean, quantity=product.stock, page_path=page_path, title=product.title, message=message)


def read_local_libri_stock(ean: str, dirs: list[str]) -> LibriStock:
    page_path = local_libri_page(ean, dirs)
    if page_path is None:
        raise RuntimeError("No local Libri page found. Enable --refresh-libri or add the page to --local-libri-dir.")
    product = parse_libri_detail_html(page_path)
    return LibriStock(ean=ean, quantity=product.stock, page_path=page_path, title=product.title)


def base_log_row(sku: TikTokSku, status: str = "", message: str = "") -> dict[str, Any]:
    return {
        "timestamp_utc": now_utc(),
        "ean": sku.ean,
        "seller_sku": sku.seller_sku,
        "product_id": sku.product_id,
        "sku_id": sku.sku_id,
        "warehouse_id": sku.warehouse_id,
        "tiktok_quantity_before": "" if sku.current_quantity is None else sku.current_quantity,
        "libri_quantity": "",
        "quantity_sent": "",
        "status": status,
        "message": message,
        "libri_page": "",
        "product_status": sku.product_status,
        "title": sku.title,
    }


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=LOG_FIELDS, delimiter=";")
        writer.writeheader()
        writer.writerows(rows)


def write_summary(path: Path, rows: list[dict[str, Any]], dry_run: bool, log_path: Path) -> None:
    counts: dict[str, int] = {}
    for row in rows:
        status = clean(row.get("status")) or "unknown"
        counts[status] = counts.get(status, 0) + 1
    lines = [
        "# Libri Inventory Update",
        "",
        f"Generated: {dt.datetime.now().isoformat(timespec='seconds')}",
        f"Dry run: {'yes' if dry_run else 'no'}",
        f"Log: {log_path}",
        "",
        "## Counts",
        "",
    ]
    for status, count in sorted(counts.items()):
        lines.append(f"- {status}: {count}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def effective_warehouse_id(args: argparse.Namespace, env: dict[str, str]) -> str:
    return clean(
        args.warehouse_id
        or os.environ.get("TIKTOK_WAREHOUSE_ID")
        or env.get("TIKTOK_WAREHOUSE_ID")
        or env.get("TIKTOK_DEFAULT_WAREHOUSE_ID")
        or DEFAULT_WAREHOUSE_ID
    )


def target_quantity(libri_quantity: int, max_quantity: int) -> int:
    if max_quantity > 0:
        return min(libri_quantity, max_quantity)
    return libri_quantity


def response_errors(response: dict[str, Any]) -> str:
    data = response.get("data") or {}
    candidates: list[Any] = []
    for key in ["errors", "extra_errors", "failed_skus"]:
        value = data.get(key)
        if value:
            candidates.append(value)
    for sku in data.get("skus") or []:
        if isinstance(sku, dict) and sku.get("extra_errors"):
            candidates.append(sku)
    return compact_json(candidates) if candidates else ""


def run(args: argparse.Namespace) -> int:
    env_path = Path(args.env)
    env = load_env_file(env_path)
    warehouse_id = effective_warehouse_id(args, env)
    client = TikTokInventoryClient(env, env_path)
    requested_seller_skus = collect_requested_seller_skus(args)

    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    run_dir = Path(args.output_root) / stamp
    libri_page_dir = run_dir / "libri_pages"
    log_path = run_dir / "inventory_update_log.csv"
    summary_path = run_dir / "summary.md"

    tiktok_skus, log_rows = resolve_tiktok_skus(client, args, requested_seller_skus, warehouse_id)
    write_csv(log_path, log_rows)

    if not tiktok_skus:
        write_summary(summary_path, log_rows, args.dry_run, log_path)
        print("Resolved TikTok SKUs: 0")
        print(f"Log: {log_path.resolve()}")
        return 1 if not log_rows else 0

    opener = login(env_path) if args.refresh_libri else None

    for sku in tiktok_skus:
        row = base_log_row(sku)
        try:
            if not sku.product_id or not sku.sku_id:
                row.update(status="skipped_missing_tiktok_ids", message="TikTok product_id or sku_id was empty.")
                log_rows.append(row)
                write_csv(log_path, log_rows)
                continue
            if not sku.warehouse_id:
                row.update(status="skipped_missing_warehouse", message="No warehouse ID found or configured.")
                log_rows.append(row)
                write_csv(log_path, log_rows)
                continue

            stock = fetch_libri_stock(sku.ean, opener, libri_page_dir) if args.refresh_libri else read_local_libri_stock(sku.ean, args.local_libri_dir)
            row["libri_page"] = str(stock.page_path)
            if stock.title and not row["title"]:
                row["title"] = stock.title
            if stock.quantity is None:
                row.update(status="skipped_missing_libri_stock", message=stock.message or "Libri stock field was empty.")
                log_rows.append(row)
                write_csv(log_path, log_rows)
                continue

            quantity = target_quantity(stock.quantity, args.max_quantity)
            row["libri_quantity"] = stock.quantity
            row["quantity_sent"] = quantity
            if sku.current_quantity == quantity and not args.update_unchanged:
                row.update(status="unchanged", message=stock.message or "TikTok quantity already matches Libri.")
            elif args.dry_run:
                row.update(status="dry_run_update", message=stock.message or "Would update TikTok inventory.")
            else:
                response = client.update_inventory(sku.product_id, sku.sku_id, sku.warehouse_id, quantity)
                errors = response_errors(response)
                if errors:
                    row.update(status="updated_with_response_errors", message=errors[:1500])
                else:
                    request_id = clean(response.get("request_id"))
                    row.update(status="updated", message=stock.message or (f"TikTok request_id={request_id}" if request_id else "TikTok inventory updated."))
                if args.tiktok_delay:
                    time.sleep(args.tiktok_delay)
        except Exception as exc:
            row.update(status="failed", message=str(exc)[:1500])
            log_rows.append(row)
            write_csv(log_path, log_rows)
            if args.stop_on_error:
                raise
            continue

        log_rows.append(row)
        write_csv(log_path, log_rows)
        if args.libri_delay:
            time.sleep(args.libri_delay)

    write_summary(summary_path, log_rows, args.dry_run, log_path)
    updated_count = sum(1 for row in log_rows if row["status"] == "updated")
    failed_count = sum(1 for row in log_rows if row["status"] == "failed")
    dry_run_count = sum(1 for row in log_rows if row["status"] == "dry_run_update")
    unchanged_count = sum(1 for row in log_rows if row["status"] == "unchanged")
    print(f"Resolved TikTok SKUs: {len(tiktok_skus)}")
    print(f"Updated: {updated_count}")
    print(f"Dry-run updates: {dry_run_count}")
    print(f"Unchanged: {unchanged_count}")
    print(f"Failed: {failed_count}")
    print(f"Log: {log_path.resolve()}")
    return 1 if failed_count else 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Update TikTok Shop quantities from current Mein.Libri stock.")
    parser.add_argument("--env", default=".env")
    parser.add_argument("--output-root", default=str(DEFAULT_OUTPUT_ROOT))
    parser.add_argument("--seller-sku-prefix", default=DEFAULT_SELLER_SKU_PREFIX)
    parser.add_argument("--warehouse-id", default="", help="TikTok warehouse ID. Defaults to TIKTOK_WAREHOUSE_ID or the current project default.")
    parser.add_argument("--sku", action="append", default=[], help="Seller SKU to update. Can be repeated.")
    parser.add_argument("--ean", action="append", default=[], help="Libri ISBN/EAN to update. Can be repeated.")
    parser.add_argument("--sku-csv", action="append", default=[], help="CSV containing seller_sku, ean, isbn, or gtin_code. Can be repeated.")
    parser.add_argument("--upload-log", action="append", default=[], help="api_product_upload_log.csv source. Can be repeated.")
    parser.add_argument("--source-statuses", default=",".join(sorted(DEFAULT_UPLOAD_LOG_STATUSES)), help="Upload-log statuses to include.")
    parser.add_argument("--workbook", action="append", default=[], help="TikTok upload workbook source. Can be repeated.")
    parser.add_argument("--start-row", type=int, default=7, help="First workbook data row.")
    parser.add_argument("--discover-tiktok-products", action=argparse.BooleanOptionalAction, default=True)
    parser.add_argument(
        "--product-statuses",
        default=DEFAULT_PRODUCT_STATUSES,
        help="Comma-separated TikTok product statuses to include. Defaults skip deleted listings.",
    )
    parser.add_argument("--page-size", type=int, default=50)
    parser.add_argument("--max-pages", type=int, default=0, help="Optional cap when discovering TikTok products.")
    parser.add_argument("--search-batch-size", type=int, default=10)
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--refresh-libri", action=argparse.BooleanOptionalAction, default=True)
    parser.add_argument("--local-libri-dir", action="append", default=["libri_product_pages", "libri_bulk_pages"])
    parser.add_argument("--max-quantity", type=int, default=0, help="Optional cap before sending quantity to TikTok.")
    parser.add_argument("--dry-run", action="store_true", help="Fetch and log quantities without writing to TikTok.")
    parser.add_argument("--update-unchanged", action="store_true", help="Send TikTok updates even when quantities already match.")
    parser.add_argument("--libri-delay", type=float, default=0.3)
    parser.add_argument("--tiktok-delay", type=float, default=0.2)
    parser.add_argument("--stop-on-error", action=argparse.BooleanOptionalAction, default=False)
    return parser


def main(argv: list[str] | None = None) -> int:
    return run(build_parser().parse_args(argv))


if __name__ == "__main__":
    raise SystemExit(main())
