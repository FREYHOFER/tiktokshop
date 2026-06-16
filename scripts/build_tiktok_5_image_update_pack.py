#!/usr/bin/env python3
"""Build a TikTok bulk-edit workbook with at least five product images.

The script preserves TikTok's workbook structure. It replaces the primary image
with a single front-cover asset and fills image_2..image_5 from existing product
photos, official publisher/Libri secondary media where safe, or neutral
cover-derived product views.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import html
import io
import json
import mimetypes
import os
import re
import time
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
import uuid
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from PIL import Image, ImageDraw, ImageFilter, ImageOps

import sys

sys.path.insert(0, str(Path(__file__).resolve().parent))
from build_tiktok_quality_update_pack import (  # noqa: E402
    create_white_cover_image_from_image,
    download_image,
    extract_reference_cover_image,
    image_is_single_cover_like,
    librinet_media_role,
    trim_plain_border,
)
from tiktok_libri_pipeline import parse_libri_detail_html  # noqa: E402
from upload_tiktok_images_cloudinary import cloudinary_signature, load_env, multipart_body, require_env  # noqa: E402


DEFAULT_DATA_START_ROW = 6
IMAGE_KEYS = ["main_image"] + [f"image_{idx}" for idx in range(2, 10)]
PRIMARY_GALLERY_KEYS = ["main_image", "image_2", "image_3", "image_4", "image_5"]
PRH_PUBLISHER_TERMS = [
    "penguin",
    "heyne",
    "goldmann",
    "btb",
    "blanvalet",
    "luchterhand",
    "bertelsmann",
    "dva",
    "siedler",
    "random house",
]
USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"


@dataclass
class LocalAsset:
    path: Path
    source: str
    role: str
    hosted_url: str = ""
    upload_status: str = "pending"
    error: str = ""


def slugify(value: str) -> str:
    value = (
        value.lower()
        .replace("ä", "ae")
        .replace("ö", "oe")
        .replace("ü", "ue")
        .replace("ß", "ss")
    )
    value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", "-", value).strip("-")


def author_slug(author: str) -> str:
    parts = [part.strip() for part in re.split(r"[,;/]", author or "") if part.strip()]
    if len(parts) >= 2:
        return slugify(f"{parts[1]} {parts[0]}")
    return slugify(author)


def title_variants(title: str) -> list[str]:
    raw = html.unescape(title or "")
    variants = [
        raw,
        re.split(r"\s+-\s+|\s+–\s+|\s+:", raw)[0],
        raw.split(".")[0],
    ]
    words = re.findall(r"[\wÄÖÜäöüß]+", raw)
    if len(words) > 7:
        variants.append(" ".join(words[:7]))
    deduped: list[str] = []
    for variant in variants:
        normalized = " ".join(variant.split())
        if normalized and normalized not in deduped:
            deduped.append(normalized)
    return deduped


def is_prh_publisher(publisher: str) -> bool:
    text = (publisher or "").casefold()
    return any(term in text for term in PRH_PUBLISHER_TERMS)


def fetch_text(url: str, timeout: int = 5) -> tuple[str, str]:
    request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(request, timeout=timeout) as response:
        return response.read().decode("utf-8", "replace"), response.geturl()


def try_prh_product_page(product) -> tuple[str, str]:
    if not product or not is_prh_publisher(product.publisher):
        return "", ""
    author = author_slug(product.author)
    if not author:
        return "", ""
    formats = ["taschenbuch", "paperback", "gebundenes-buch"]
    for title in title_variants(product.title)[:2]:
        base = f"{author}-{slugify(title)}"
        for book_format in formats:
            url = f"https://www.penguin.de/buecher/{base}/{book_format}/{product.ean}"
            try:
                text, final_url = fetch_text(url)
            except (OSError, urllib.error.URLError, urllib.error.HTTPError):
                continue
            if product.ean in text or product.ean.replace("", "-").strip("-") in text:
                return text, final_url
    return "", ""


def responsive_candidates(url: str) -> list[str]:
    url = html.unescape(url)
    if url.startswith("/"):
        url = "https://www.penguin.de" + url
    match = re.search(r"(/resource/responsive-image/([^/]+)/)([^/]+)(/.+)$", urllib.parse.urlparse(url).path)
    if not match:
        return [url]
    prefix, _asset_id, _size, suffix = match.groups()
    suffix = re.sub(r"\.(jpg|jpeg|png|webp)$", ".webp", suffix, flags=re.I)
    sizes = ["930", "800", "600", "h1200", "h900", "h700", "h600", "350", "280"]
    return [f"https://www.penguin.de{prefix}{size}{suffix}" for size in sizes]


def download_best_image(urls: list[str]) -> tuple[Image.Image | None, str]:
    best_image = None
    best_url = ""
    best_area = 0
    for url in urls:
        try:
            request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT, "Accept": "image/webp,image/*,*/*"})
            with urllib.request.urlopen(request, timeout=8) as response:
                data = response.read()
            image = Image.open(io.BytesIO(data))
            image.load()
            image = ImageOps.exif_transpose(image).convert("RGB")
        except Exception:  # noqa: BLE001 - try the next responsive size
            continue
        area = image.width * image.height
        if area > best_area:
            best_image = image
            best_url = url
            best_area = area
    return best_image, best_url


def parse_prh_images(page_text: str) -> dict[str, list[tuple[str, str]]]:
    result = {"front": [], "secondary": []}
    for match in re.finditer(r"image_0[12]:(https://www\.penguin\.de/resource/responsive-image/[^;\"']+)", page_text):
        result["front"].append(("Verlagscover Penguin Random House", match.group(1)))
    gallery_match = re.search(r'm04-product-details-thumbnails-slider.+?</div>\s*</div>', page_text, flags=re.I | re.S)
    if gallery_match:
        gallery = gallery_match.group(0)
        for img_match in re.finditer(r'<img[^>]+src="([^"]+)"[^>]+alt="([^"]*)"', gallery, flags=re.I | re.S):
            url, alt = img_match.group(1), html.unescape(img_match.group(2) or "")
            alt_norm = alt.casefold()
            if any(term in alt_norm for term in ["reihenübersicht", "reihe", "bundle", "set"]):
                continue
            if alt.strip():
                result["secondary"].append((f"Verlagsgalerie Penguin Random House: {alt}", url))
    return result


def canvas_with_image(image: Image.Image, output_path: Path, canvas_size: int, fill_ratio: float = 0.84) -> None:
    image = trim_plain_border(ImageOps.exif_transpose(image)).convert("RGBA")
    canvas = Image.new("RGBA", (canvas_size, canvas_size), "white")
    max_w = int(canvas_size * fill_ratio)
    max_h = int(canvas_size * fill_ratio)
    scale = min(max_w / image.width, max_h / image.height)
    resized = image.resize((max(1, int(image.width * scale)), max(1, int(image.height * scale))), Image.Resampling.LANCZOS)
    shadow = Image.new("RGBA", (resized.width + 34, resized.height + 34), (0, 0, 0, 0))
    shadow_draw = ImageDraw.Draw(shadow)
    shadow_draw.rounded_rectangle((16, 16, resized.width + 18, resized.height + 18), radius=2, fill=(0, 0, 0, 38))
    shadow = shadow.filter(ImageFilter.GaussianBlur(10))
    x = (canvas_size - resized.width) // 2
    y = (canvas_size - resized.height) // 2
    canvas.alpha_composite(shadow, (x - 8, y - 8))
    canvas.alpha_composite(resized, (x, y))
    output_path.parent.mkdir(parents=True, exist_ok=True)
    canvas.convert("RGB").save(output_path, quality=92, optimize=True)


def make_angle_view(image: Image.Image, output_path: Path, canvas_size: int) -> None:
    cover = trim_plain_border(ImageOps.exif_transpose(image)).convert("RGBA")
    target_h = int(canvas_size * 0.76)
    target_w = max(1, int(cover.width * (target_h / cover.height)))
    cover = cover.resize((target_w, target_h), Image.Resampling.LANCZOS)
    spine_w = max(30, int(target_w * 0.085))
    avg = cover.resize((1, 1), Image.Resampling.BILINEAR).getpixel((0, 0))
    canvas = Image.new("RGBA", (canvas_size, canvas_size), "white")
    x = (canvas_size - target_w - spine_w) // 2 + 20
    y = (canvas_size - target_h) // 2
    shadow = Image.new("RGBA", (target_w + spine_w + 70, target_h + 70), (0, 0, 0, 0))
    draw = ImageDraw.Draw(shadow)
    draw.rounded_rectangle((26, 26, target_w + spine_w + 38, target_h + 38), radius=3, fill=(0, 0, 0, 44))
    shadow = shadow.filter(ImageFilter.GaussianBlur(14))
    canvas.alpha_composite(shadow, (x - 30, y - 24))
    spine = Image.new("RGBA", (spine_w, target_h), avg[:3] + (255,))
    spine_draw = ImageDraw.Draw(spine)
    spine_draw.rectangle((0, 0, spine_w - 1, target_h), fill=tuple(max(0, int(v * 0.7)) for v in avg[:3]) + (255,))
    spine_draw.rectangle((spine_w - 4, 0, spine_w - 1, target_h), fill=(255, 255, 255, 45))
    canvas.alpha_composite(spine, (x, y))
    canvas.alpha_composite(cover, (x + spine_w, y))
    output_path.parent.mkdir(parents=True, exist_ok=True)
    canvas.convert("RGB").save(output_path, quality=92, optimize=True)


def make_detail_crop(image: Image.Image, output_path: Path, canvas_size: int, part: str) -> None:
    cover = trim_plain_border(ImageOps.exif_transpose(image)).convert("RGB")
    if part == "top":
        crop = cover.crop((0, 0, cover.width, int(cover.height * 0.58)))
    else:
        crop = cover.crop((0, int(cover.height * 0.42), cover.width, cover.height))
    canvas_with_image(crop, output_path, canvas_size, fill_ratio=0.88)


def image_fingerprint(value: str) -> str:
    normalized = re.sub(r"[?&](?:dr|from|height|width|t|idc|ps|shcp|shp)=[^&]+", "", value or "")
    return hashlib.sha1(normalized.encode("utf-8", "ignore")).hexdigest()


def upload_asset(asset: LocalAsset, public_id: str, folder: str, cloud_name: str, api_key: str, api_secret: str) -> LocalAsset:
    timestamp = str(int(time.time()))
    fields = {
        "api_key": api_key,
        "timestamp": timestamp,
        "folder": folder,
        "public_id": public_id,
        "overwrite": "true",
        "unique_filename": "false",
    }
    fields["signature"] = cloudinary_signature(fields, api_secret)
    body, content_type = multipart_body(fields, "file", asset.path)
    request = urllib.request.Request(
        f"https://api.cloudinary.com/v1_1/{cloud_name}/image/upload",
        data=body,
        headers={"Content-Type": content_type or mimetypes.guess_type(asset.path.name)[0] or "application/octet-stream"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=60) as response:
            payload = json.loads(response.read().decode("utf-8"))
        asset.hosted_url = payload.get("secure_url", "")
        asset.upload_status = "uploaded" if asset.hosted_url else "error"
        if not asset.hosted_url:
            asset.error = "Cloudinary response without secure_url"
    except Exception as exc:  # noqa: BLE001
        asset.upload_status = "error"
        asset.error = f"{type(exc).__name__}: {exc}"
    return asset


def workbook_headers(sheet) -> dict[str, int]:
    return {
        str(sheet.cell(row=1, column=col).value): col
        for col in range(1, sheet.max_column + 1)
        if sheet.cell(row=1, column=col).value
    }


def current_image_urls(sheet, headers: dict[str, int], row_idx: int) -> list[str]:
    urls = []
    for key in IMAGE_KEYS:
        col = headers.get(key)
        if not col:
            continue
        value = str(sheet.cell(row_idx, col).value or "").strip()
        if value.startswith(("http://", "https://")):
            urls.append(value)
    return list(dict.fromkeys(urls))


def build_assets_for_row(row_idx: int, ean: str, product, current_urls: list[str], output_dir: Path, canvas_size: int) -> tuple[list[LocalAsset], dict[str, str]]:
    detail_path = Path("libri_bulk_pages") / f"{ean}.html"
    reference_image = extract_reference_cover_image(detail_path) if detail_path.exists() else None
    reference_safe = bool(reference_image) and image_is_single_cover_like(reference_image)
    source_page = ""
    cover_image = None
    cover_source = ""
    cover_method = ""
    cover_confidence = ""
    secondary_images: list[tuple[Image.Image, str]] = []
    if product:
        page_text, source_page = try_prh_product_page(product)
        if page_text:
            prh_images = parse_prh_images(page_text)
            for source, url in prh_images["front"]:
                image, best_url = download_best_image(responsive_candidates(url))
                if image and image_is_single_cover_like(image):
                    cover_image = image
                    cover_source = f"{source} ({best_url})"
                    cover_method = "publisher_front_cover"
                    cover_confidence = "high"
                    break
            for source, url in prh_images["secondary"]:
                image, best_url = download_best_image(responsive_candidates(url))
                if image and image_is_single_cover_like(image):
                    secondary_images.append((image, f"{source} ({best_url})"))

    if cover_image is None and reference_image is not None and reference_safe:
        cover_image = reference_image
        cover_source = "Libri eingebettetes Detailcover"
        cover_method = "libri_embedded_detail_cover"
        cover_confidence = "medium"
    if cover_image is None and product:
        for url in product.image_urls:
            if librinet_media_role(url) == 2:
                continue
            try:
                image = download_image(url)
            except Exception:  # noqa: BLE001
                continue
            if image_is_single_cover_like(image):
                cover_image = image
                cover_source = f"Libri Medienbild ({url})"
                cover_method = "libri_media_cover"
                cover_confidence = "medium"
                break
    if cover_image is None:
        for url in current_urls:
            try:
                image = download_image(url)
            except Exception:  # noqa: BLE001
                continue
            if image_is_single_cover_like(image):
                cover_image = image
                cover_source = f"aktuelles TikTok-Bild ({url})"
                cover_method = "existing_tiktok_image"
                cover_confidence = "low"
                break
    if cover_image is None:
        return [], {
            "source_row": str(row_idx),
            "ean": ean,
            "cover_source": "",
            "publisher_page": source_page,
            "status": "no_cover_image_available",
            "asset_count": "0",
        }

    assets_dir = output_dir / "generated_gallery_images" / ean
    assets: list[LocalAsset] = []
    main_path = assets_dir / f"{ean}_01_front_cover.jpg"
    create_white_cover_image_from_image(cover_image, main_path, canvas_size)
    assets.append(LocalAsset(main_path, cover_source, "main_front_cover"))

    for idx, (image, source) in enumerate(secondary_images[:2], start=1):
        path = assets_dir / f"{ean}_secondary_{idx}.jpg"
        canvas_with_image(image, path, canvas_size, fill_ratio=0.86)
        assets.append(LocalAsset(path, source, "official_secondary"))

    derivative_specs = [
        ("02_angle_view", "cover_derived_angle"),
        ("03_full_cover_large", "cover_derived_large_front"),
        ("04_detail_top", "cover_derived_detail_top"),
        ("05_detail_bottom", "cover_derived_detail_bottom"),
    ]
    for name, role in derivative_specs:
        path = assets_dir / f"{ean}_{name}.jpg"
        if role == "cover_derived_angle":
            make_angle_view(cover_image, path, canvas_size)
        elif role == "cover_derived_large_front":
            canvas_with_image(cover_image, path, canvas_size, fill_ratio=0.94)
        elif role == "cover_derived_detail_top":
            make_detail_crop(cover_image, path, canvas_size, "top")
        else:
            make_detail_crop(cover_image, path, canvas_size, "bottom")
        assets.append(LocalAsset(path, "aus Frontcover erzeugt", role))

    return assets[:5], {
        "source_row": str(row_idx),
        "ean": ean,
        "cover_source": cover_source,
        "publisher_page": source_page,
            "status": "assets_created",
            "asset_count": str(min(5, len(assets))),
        "cover_selection_method": cover_method,
        "cover_confidence": cover_confidence,
    }


def write_csv(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def build(args: argparse.Namespace) -> int:
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    load_env(Path(args.env))
    cloud_name = require_env("CLOUDINARY_CLOUD_NAME")
    api_key = require_env("CLOUDINARY_API_KEY")
    api_secret = require_env("CLOUDINARY_API_SECRET")
    folder = os.environ.get("CLOUDINARY_FOLDER", "tiktokshop/books").strip() or "tiktokshop/books"

    workbook = openpyxl.load_workbook(args.workbook)
    sheet = workbook["Template"] if "Template" in workbook.sheetnames else workbook.active
    headers = workbook_headers(sheet)
    required = ["product_name", "gtin_code", "main_image"]
    missing = [key for key in required if key not in headers]
    if missing:
        raise SystemExit("Missing required headers: " + ", ".join(missing))

    data_start_row = DEFAULT_DATA_START_ROW
    while data_start_row <= sheet.max_row and not str(sheet.cell(data_start_row, headers["product_name"]).value or "").strip():
        data_start_row += 1
    rows = [
        row_idx
        for row_idx in range(data_start_row, sheet.max_row + 1)
        if str(sheet.cell(row_idx, headers["product_name"]).value or "").strip()
    ]
    if args.limit:
        rows = rows[: args.limit]
    report_rows: list[dict[str, str]] = []
    upload_rows: list[dict[str, str]] = []
    patch_rows: list[dict[str, str]] = []

    for index, row_idx in enumerate(rows, start=1):
        ean = re.sub(r"\D", "", str(sheet.cell(row_idx, headers["gtin_code"]).value or ""))
        if not ean:
            continue
        detail_path = Path(args.detail_dir) / f"{ean}.html"
        product = parse_libri_detail_html(detail_path) if detail_path.exists() else None
        current_urls = current_image_urls(sheet, headers, row_idx)
        main_good = bool(current_urls) and "res.cloudinary.com" in current_urls[0] and f"LIBRI-{ean}" in current_urls[0]
        if len(current_urls) >= 5 and main_good:
            report_rows.append(
                {
                    "source_row": str(row_idx),
                    "ean": ean,
                    "cover_source": "unveraendert",
                    "publisher_page": "",
                    "status": "already_has_5_images",
                    "asset_count": "0",
                    "cover_selection_method": "keep_existing",
                    "cover_confidence": "high",
                    "product_name": str(sheet.cell(row_idx, headers["product_name"]).value or ""),
                    "current_image_count": str(len(current_urls)),
                    "target_image_count": str(len(current_urls)),
                    "uploaded_image_count": "0",
                }
            )
            print(f"[{index}/{len(rows)}] {ean} -> bereits {len(current_urls)} Bilder")
            continue
        assets, report = build_assets_for_row(row_idx, ean, product, current_urls, output_dir, args.canvas_size)
        report["product_name"] = str(sheet.cell(row_idx, headers["product_name"]).value or "")
        report["current_image_count"] = str(len(current_urls))
        report["target_image_count"] = "0"
        report["uploaded_image_count"] = "0"
        if not assets:
            report_rows.append(report)
            continue

        final_urls: list[str] = []
        seen: set[str] = set()
        uploaded_count = 0
        asset_cursor = 0
        if main_good:
            final_urls.append(current_urls[0])
            seen.add(image_fingerprint(current_urls[0]))
            asset_cursor = 1
        else:
            asset = assets[0]
            if not args.no_upload:
                upload_asset(asset, f"LIBRI-{ean}-gallery-01", folder, cloud_name, api_key, api_secret)
            upload_rows.append(
                {
                    "ean": ean,
                    "product_name": report["product_name"],
                    "slot": "1",
                    "role": asset.role,
                    "source": asset.source,
                    "local_path": str(asset.path.resolve()),
                    "hosted_url": asset.hosted_url,
                    "upload_status": asset.upload_status,
                    "error": asset.error,
                }
            )
            if asset.hosted_url:
                final_urls.append(asset.hosted_url)
                seen.add(image_fingerprint(asset.hosted_url))
                uploaded_count += 1
            asset_cursor = 1

        for existing in current_urls[1:]:
            fp = image_fingerprint(existing)
            if fp not in seen and len(final_urls) < 5:
                final_urls.append(existing)
                seen.add(fp)

        for local_slot_idx, asset in enumerate(assets[asset_cursor:], start=asset_cursor + 1):
            if len(final_urls) >= 5:
                break
            if not args.no_upload:
                upload_asset(asset, f"LIBRI-{ean}-gallery-{local_slot_idx:02d}", folder, cloud_name, api_key, api_secret)
            upload_rows.append(
                {
                    "ean": ean,
                    "product_name": report["product_name"],
                    "slot": str(local_slot_idx),
                    "role": asset.role,
                    "source": asset.source,
                    "local_path": str(asset.path.resolve()),
                    "hosted_url": asset.hosted_url,
                    "upload_status": asset.upload_status,
                    "error": asset.error,
                }
            )
            if asset.hosted_url:
                fp = image_fingerprint(asset.hosted_url)
                if fp not in seen:
                    final_urls.append(asset.hosted_url)
                    seen.add(fp)
                    uploaded_count += 1

        if final_urls:
            for key, url in zip(PRIMARY_GALLERY_KEYS, final_urls):
                col = headers.get(key)
                if col:
                    sheet.cell(row_idx, col).value = url
            report["target_image_count"] = str(len(final_urls))
            report["uploaded_image_count"] = str(uploaded_count)
            patch_rows.append(
                {
                    "row": str(row_idx),
                    "ean": ean,
                    "seller_sku": str(sheet.cell(row_idx, headers.get("seller_sku", 1)).value or ""),
                    "image_count": str(len(final_urls)),
                    "changes": "|".join(PRIMARY_GALLERY_KEYS[: len(final_urls)]),
                }
            )
        report_rows.append(report)
        print(f"[{index}/{len(rows)}] {ean} -> {report.get('target_image_count', '0')} Bilder")

    output_workbook = output_dir / "tiktok_bulk_edit_5_images_hosted.xlsx"
    workbook.save(output_workbook)
    write_csv(output_dir / "image_source_report.csv", report_rows, list(report_rows[0].keys()) if report_rows else [])
    write_csv(output_dir / "gallery_upload_log.csv", upload_rows, ["ean", "product_name", "slot", "role", "source", "local_path", "hosted_url", "upload_status", "error"])
    write_csv(output_dir / "tiktok_bulk_edit_5_images_hosted.patch_log.csv", patch_rows, ["row", "ean", "seller_sku", "image_count", "changes"])
    print(f"Products: {len(rows)}")
    print(f"Patched rows: {len(patch_rows)}")
    print(f"Output: {output_workbook.resolve()}")
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Create a TikTok bulk-edit workbook with five images per product.")
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--detail-dir", default="libri_bulk_pages")
    parser.add_argument("--output-dir", default="outputs/current_listing_5_images_20260520")
    parser.add_argument("--env", default=".env")
    parser.add_argument("--canvas-size", type=int, default=1200)
    parser.add_argument("--no-upload", action="store_true")
    parser.add_argument("--limit", type=int, default=0)
    return parser


def main() -> int:
    return build(build_parser().parse_args())


if __name__ == "__main__":
    raise SystemExit(main())
