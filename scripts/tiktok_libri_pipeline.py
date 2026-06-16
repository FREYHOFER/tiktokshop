#!/usr/bin/env python3
"""Build TikTok Shop bulk-upload workbooks from saved Libri product data.

The first implementation intentionally works with local, auditable inputs:
- saved Mein.Libri product detail HTML files
- saved Mein.Libri bestseller PDF exports
- an optional manual CSV candidate list

It preserves the TikTok Shop category template structure and only writes data
rows into a copied workbook.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import html
import io
import json
import math
import re
import sys
import urllib.error
import urllib.request
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover - environment guard
    raise SystemExit("Missing dependency: openpyxl") from exc

try:
    from PIL import Image
except ImportError:  # pragma: no cover - optional but expected in Codex runtime
    Image = None

try:
    import pypdf
except ImportError:  # pragma: no cover - optional when no PDF input is used
    pypdf = None


CATEGORY = "Literatur und Kunst/Roman"
BRAND_NO_BRAND = "Keine Marke"
GTIN_TYPE = "ISBN"
DEFAULT_RP_ID = "69e8ebf9401201b6c3954a0f"
DEFAULT_MANUFACTURER_ID = ""
DEFAULT_WARNING_ON_PACKAGE = "Nein"
ENV_FILE_NAME = ".env"
MIN_TIKTOK_IMAGE_SIDE_PX = 300
MIN_LISTING_QUALITY_TITLE_CHARS = 40

TEMPLATE_DATA_START_ROW = 7
TEMPLATE_MAX_DATA_ROW = 5000

FORBIDDEN_TEMPLATE_KEYS = {
    "product_property/100530",  # JAHR
    "product_property/100532",  # Herausgeber
    "product_property/100534",  # ISBN/ISSN
    "product_property/100536",  # Übersetzer
    "product_property/100537",  # Editor
    "product_property/100538",  # Anzahl der Seiten
}

REQUIRED_UPLOAD_KEYS = {
    "category",
    "product_name",
    "product_description",
    "main_image",
    "parcel_weight",
    "parcel_length",
    "parcel_width",
    "parcel_height",
    "price",
    "quantity",
    "rp_ids",
    "product_property/102277",
}

REJECT_TERMS = {
    "digital_format": [
        "ebook",
        "e-book",
        "hörbuch",
        "hoerbuch",
        "audio download",
        "mp3",
        "epub",
    ],
    "used_or_open_box": [
        "gebraucht",
        "mängelexemplar",
        "maengelexemplar",
        "second hand",
        "used",
    ],
}

REVIEW_TERMS = {
    "adult_or_explicit": [
        r"\b18\+\b",
        r"\bab\s*18\b",
        "leseempfehlung ab 18",
        "dark romance",
        "spice",
        "spicy",
        "bdsm",
        "erotik",
        "sexuell",
        "sexual",
    ],
    "self_harm_or_suicide": [
        "suizid",
        "suicide",
        "selbstverletz",
        "self-harm",
        "self harm",
    ],
    "violent_or_shocking": [
        "tödlich",
        "toedlich",
        "grausam",
        "brutal",
        "folter",
        "mord",
        "messer",
        "knife",
    ],
}

DETAIL_LEAK_MARKERS = [
    "mein.libri",
    "herzlich willkommen",
    "logout",
    "gutschriften",
    "verlags-",
    "bs-sendungen",
    "konto",
    "in den warenkorb",
]

DESCRIPTION_REPLACEMENTS = [
    (r"\bstand\s+er\b", "war er"),
    (r"\bstand\s+sie\b", "war sie"),
    (r"\bstand\s+es\b", "war es"),
    (r"\bstand\s+ihm\b", "war ihm"),
    (r"\bstand\s+ihr\b", "war ihr"),
]


@dataclass
class ImageCheck:
    url: str
    ok: bool
    public_ok: bool = False
    authenticated_ok: bool = False
    requires_auth: bool = False
    width: int | None = None
    height: int | None = None
    content_type: str = ""
    reason: str = ""

    @property
    def ratio(self) -> float | None:
        if not self.width or not self.height:
            return None
        return self.width / self.height

    @property
    def tiktok_size_ok(self) -> bool:
        return bool(
            self.width
            and self.height
            and self.width >= MIN_TIKTOK_IMAGE_SIDE_PX
            and self.height >= MIN_TIKTOK_IMAGE_SIDE_PX
        )

    @property
    def template_ratio_ok(self) -> bool:
        ratio = self.ratio
        if ratio is None:
            return False
        allowed = (4 / 3, 3 / 4, 1.0)
        return any(abs(ratio - candidate) <= 0.04 for candidate in allowed)


@dataclass
class ProductRecord:
    title: str = ""
    author: str = ""
    subtitle: str = ""
    publisher: str = ""
    language: str = ""
    original_language: str = ""
    translator: str = ""
    binding: str = ""
    edition: str = ""
    release_date: str = ""
    pages: int | None = None
    weight_g: int | None = None
    stock: int | None = None
    price: float | None = None
    ean: str = ""
    libri_no: str = ""
    product_group: str = ""
    blurb: str = ""
    author_bio: str = ""
    image_urls: list[str] = field(default_factory=list)
    source_names: list[str] = field(default_factory=list)
    source_rank: int | None = None
    source_price: float | None = None
    status: str = "review"
    reasons: list[str] = field(default_factory=list)
    image_checks: list[ImageCheck] = field(default_factory=list)

    @property
    def normalized_title(self) -> str:
        return normalize_key(self.title)

    @property
    def uploadable_identity(self) -> str:
        return self.ean or self.normalized_title


def normalize_text(value: str | None) -> str:
    if not value:
        return ""
    value = html.unescape(str(value))
    value = value.replace("\xa0", " ")
    value = value.replace("▾", " ").replace("▴", " ")
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def normalize_key(value: str | None) -> str:
    return re.sub(r"[^a-z0-9]+", " ", normalize_text(value).lower()).strip()


def clean_detail_value(value: str | None, max_len: int = 220) -> str:
    cleaned = normalize_text(value)
    if not cleaned:
        return ""
    lowered = cleaned.lower()
    if any(marker in lowered for marker in DETAIL_LEAK_MARKERS):
        return ""
    if len(cleaned) > max_len:
        return ""
    return cleaned


def parse_price(value: str | None) -> float | None:
    if not value:
        return None
    match = re.search(r"(\d+(?:[.,]\d{1,2})?)", value)
    if not match:
        return None
    return float(match.group(1).replace(",", "."))


def load_env_file(path: Path) -> dict[str, str]:
    values: dict[str, str] = {}
    if not path.exists():
        return values
    for raw_line in path.read_text(encoding="utf-8", errors="replace").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        values[key.strip()] = value.strip().strip('"').strip("'")
    return values


def request_headers(cookie: str = "") -> dict[str, str]:
    headers = {"User-Agent": "Mozilla/5.0"}
    if cookie:
        headers["Cookie"] = cookie
    return headers


def parse_int(value: str | None) -> int | None:
    if not value:
        return None
    match = re.search(r"(\d+)", value.replace(".", ""))
    if not match:
        return None
    return int(match.group(1))


def html_to_visible_text(raw_html: str) -> str:
    raw_html = re.sub(r"<(script|style)[\s\S]*?</\1>", " ", raw_html, flags=re.I)
    raw_html = re.sub(r"<[^>]+>", " ", raw_html)
    return normalize_text(raw_html)


def extract_media_urls(raw_html: str) -> list[str]:
    decoded_html = html.unescape(raw_html)
    urls = re.findall(r"https://medias\.librinet\.de/[^\s\"'<>]+", decoded_html)
    return sorted(dict.fromkeys(urls))


def find_product_title(visible_text: str) -> str:
    match = re.search(
        r"Zurück zur Ergebnisseite\s+Nächster Eintrag\s+(.*?)\s+Ladenpreis",
        visible_text,
        flags=re.I,
    )
    if match:
        return normalize_text(match.group(1))
    match = re.search(r"Nächster Eintrag\s+(.*?)\s+Ladenpreis", visible_text, flags=re.I)
    if match:
        return normalize_text(match.group(1))
    match = re.search(
        r"(?:Willkommen\s+Herr\s+\S+|Logout)\s+(.+?)\s+Ladenpreis",
        visible_text,
        flags=re.I,
    )
    if match:
        value = normalize_text(match.group(1))
        value = re.sub(r"^.*?Herzlich Willkommen Herr \S+\s+", "", value, flags=re.I)
        value = re.sub(r"^.*?Logout\s+", "", value, flags=re.I)
        return value
    return ""


def extract_label_sections(visible_text: str) -> dict[str, str]:
    labels = [
        "Autor",
        "Untertitel",
        "Reihe",
        "Reihennr.",
        "Verlag",
        "Sprache",
        "Originalsprache",
        "Übersetzer",
        "Einband",
        "Auflage",
        "Erscheinungsdatum",
        "Seiten",
        "Gewicht",
        "Bestand",
        "Artikel-Nr./EAN",
        "Libri-Nr.",
        "Warengruppe",
        "Rabattgruppe",
        "Klappentext",
        "Innenansicht",
        "Über den Autor",
        "Ladenpreis",
    ]
    positions: list[tuple[int, str]] = []
    for label in labels:
        match = re.search(re.escape(label), visible_text)
        if match:
            positions.append((match.start(), label))
    positions.sort()

    sections: dict[str, str] = {}
    for idx, (start, label) in enumerate(positions):
        end = positions[idx + 1][0] if idx + 1 < len(positions) else len(visible_text)
        value = visible_text[start + len(label) : end]
        sections[label] = normalize_text(value)
    return sections


def section_before_labels(value: str, stop_labels: list[str]) -> str:
    stop_positions = []
    for label in stop_labels:
        match = re.search(rf"\b{re.escape(label)}\b", value)
        if match:
            stop_positions.append(match.start())
    if stop_positions:
        value = value[: min(stop_positions)]
    return normalize_text(value)


def parse_libri_detail_html(path: Path) -> ProductRecord:
    raw_html = path.read_text(encoding="utf-8", errors="replace")
    visible = html_to_visible_text(raw_html)
    price_match = re.search(r"Ladenpreis\s+([\d,.]+)\s+EUR", visible)
    detail_visible = visible[price_match.end():] if price_match else visible
    sections = extract_label_sections(detail_visible)

    record = ProductRecord()
    record.title = find_product_title(visible)
    record.author = section_before_labels(sections.get("Autor", ""), ["Untertitel", "Reihe", "Reihennr.", "Verlag"])
    record.subtitle = section_before_labels(sections.get("Untertitel", ""), ["Reihe", "Reihennr.", "Verlag"])
    record.publisher = clean_detail_value(sections.get("Verlag", ""))
    record.language = clean_detail_value(sections.get("Sprache", ""), max_len=80)
    record.original_language = clean_detail_value(sections.get("Originalsprache", ""), max_len=80)
    record.translator = sections.get("Übersetzer", "")
    record.binding = clean_detail_value(sections.get("Einband", ""), max_len=80)
    record.edition = clean_detail_value(sections.get("Auflage", ""), max_len=120)
    record.release_date = clean_detail_value(sections.get("Erscheinungsdatum", ""), max_len=40)
    record.pages = parse_int(sections.get("Seiten"))
    record.weight_g = parse_int(sections.get("Gewicht"))
    record.stock = parse_int(sections.get("Bestand"))
    record.ean = normalize_text(sections.get("Artikel-Nr./EAN", ""))
    record.libri_no = normalize_text(sections.get("Libri-Nr.", ""))
    record.product_group = normalize_text(sections.get("Warengruppe", ""))
    record.blurb = section_before_labels(sections.get("Klappentext", ""), ["Innenansicht", "Ãœber den Autor"])
    record.author_bio = sections.get("Über den Autor", "")
    record.author_bio = section_before_labels(
        record.author_bio,
        ["Rezension", "Schlagworte", "Produktinformationen", "Weitere Informationen", "In den Warenkorb"],
    )
    record.image_urls = extract_media_urls(raw_html)
    record.price = parse_price(price_match.group(1)) if price_match else None
    record.source_names.append(str(path))
    return record


def parse_bestseller_pdf(path: Path) -> list[ProductRecord]:
    if pypdf is None:
        return []
    reader = pypdf.PdfReader(str(path))
    products: list[ProductRecord] = []
    rank = 1
    buffer: list[str] = []

    ignored_exact = {
        "Bestseller",
        "Novitäten",
        "BookTok",
        "LITPROM",
        "SWR Bestenliste",
        "Deutschlandfunk Kultur",
        "First Choice",
        "Spiegel Bestseller",
        "BookTok Bestseller▼",
        "In den Warenkorb",
    }

    def is_ignored(line: str) -> bool:
        if not line:
            return True
        if line in ignored_exact:
            return True
        if line.startswith("https://"):
            return True
        if re.match(r"\d{2}\.\d{2}\.\d{2},", line):
            return True
        if line.startswith("Menge "):
            return True
        if "Mein.Libri - Bestellen - Bestseller" in line:
            return True
        if "In den Warenkorb" in line:
            return True
        if line in {
            "Die BookTok Bestseller für den November",
            "Englische Novitäten",
            "Internationale Literatur - Empfehlungen",
            "Internationale Bestseller",
            "Top-/Longseller Essen und Trinken, Backen",
            "Anzeige der Treffer 1 bis 20 von insgesamt 2025 pro Seite ▼",
            "Libri GmbH - Hamburg & Bad Hersfeld © 2026 Libri GmbH. Alle Rechte vorbehalten.",
        }:
            return True
        return False

    for page in reader.pages:
        for raw_line in (page.extract_text() or "").splitlines():
            line = normalize_text(raw_line)
            if is_ignored(line):
                continue
            if re.fullmatch(r"\d+(?:[.,]\d{2})\s+EUR", line):
                if len(buffer) >= 2:
                    author = buffer[-1]
                    title = normalize_text(" ".join(buffer[:-1]))
                    product = ProductRecord(
                        title=title,
                        author=author,
                        source_price=parse_price(line),
                        source_rank=rank,
                        source_names=[str(path)],
                        status="review",
                        reasons=["missing_libri_detail_page"],
                    )
                    products.append(product)
                    rank += 1
                buffer = []
                continue
            buffer.append(line)
    return products


def parse_manual_csv(path: Path) -> list[ProductRecord]:
    products: list[ProductRecord] = []
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        for row in csv.DictReader(fh):
            source_rank = parse_int(row.get("score") or row.get("source_rank") or row.get("rank"))
            product = ProductRecord(
                title=normalize_text(row.get("title") or row.get("product_name")),
                author=normalize_text(row.get("author")),
                subtitle=normalize_text(row.get("subtitle")),
                publisher=normalize_text(row.get("publisher")),
                language=normalize_text(row.get("language")),
                binding=normalize_text(row.get("binding")),
                release_date=normalize_text(row.get("release_date")),
                pages=parse_int(row.get("pages")),
                weight_g=parse_int(row.get("weight_g")),
                stock=parse_int(row.get("stock") or row.get("quantity")),
                price=parse_price(row.get("price")),
                ean=normalize_text(row.get("ean") or row.get("isbn")),
                product_group=normalize_text(row.get("product_group")),
                blurb=normalize_text(row.get("blurb") or row.get("description")),
                author_bio=normalize_text(row.get("author_bio")),
                image_urls=[normalize_text(v) for v in (row.get("images") or row.get("image_urls") or "").split("|") if normalize_text(v)],
                source_names=[str(path)],
                source_rank=source_rank,
                source_price=parse_price(row.get("source_price_text")),
                reasons=[normalize_text(row.get("notes"))] if normalize_text(row.get("notes")) else [],
            )
            source_label = normalize_text(row.get("sources") or row.get("source"))
            if source_label:
                product.source_names.append(source_label)
            products.append(product)
    return products


def merge_into(current: ProductRecord, product: ProductRecord) -> ProductRecord:
    for attr in [
        "title",
        "author",
        "subtitle",
        "publisher",
        "language",
        "original_language",
        "translator",
        "binding",
        "edition",
        "release_date",
        "ean",
        "libri_no",
        "product_group",
        "blurb",
        "author_bio",
    ]:
        if not getattr(current, attr) and getattr(product, attr):
            setattr(current, attr, getattr(product, attr))

    for attr in ["pages", "weight_g", "stock", "price", "source_rank", "source_price"]:
        if getattr(current, attr) is None and getattr(product, attr) is not None:
            setattr(current, attr, getattr(product, attr))

    current.image_urls = list(dict.fromkeys(current.image_urls + product.image_urls))
    current.source_names = list(dict.fromkeys(current.source_names + product.source_names))
    if "missing_libri_detail_page" in current.reasons and product.ean:
        current.reasons.remove("missing_libri_detail_page")
    return current


def merge_products(products: Iterable[ProductRecord]) -> list[ProductRecord]:
    """Merge sparse candidate rows with fuller Libri detail rows.

    Bestseller PDFs usually provide title/author/rank only, while saved detail
    pages provide the EAN and upload fields. The first pass joins by normalized
    title; the second pass collapses any remaining duplicate EAN rows.
    """

    by_title: dict[str, ProductRecord] = {}
    title_order: list[str] = []

    for product in products:
        title_key = product.normalized_title
        if not title_key and not product.ean:
            continue
        key = title_key or product.ean
        if key not in by_title:
            by_title[key] = product
            title_order.append(key)
        else:
            by_title[key] = merge_into(by_title[key], product)

    by_ean: dict[str, ProductRecord] = {}
    output: list[ProductRecord] = []
    for key in title_order:
        product = by_title[key]
        if product.ean and product.ean in by_ean:
            merge_into(by_ean[product.ean], product)
            continue
        if product.ean:
            by_ean[product.ean] = product
        output.append(product)

    return output


def fetch_image_metadata(url: str, timeout: int = 15, cookie: str = "") -> ImageCheck:
    if not url.startswith(("http://", "https://")):
        return ImageCheck(url=url, ok=False, reason="not_public_http_url")
    if Image is None:
        return ImageCheck(url=url, ok=True, reason="pillow_unavailable_dimensions_not_checked")
    try:
        request = urllib.request.Request(url, headers=request_headers(cookie))
        with urllib.request.urlopen(request, timeout=timeout) as response:
            data = response.read(2_500_000)
            content_type = response.headers.get("content-type", "")
        with Image.open(io.BytesIO(data)) as image:
            width, height = image.size
        return ImageCheck(url=url, ok=True, width=width, height=height, content_type=content_type)
    except (urllib.error.URLError, OSError, TimeoutError) as exc:
        return ImageCheck(url=url, ok=False, reason=f"{type(exc).__name__}: {exc}")


def validate_image_url(url: str, timeout: int = 15, auth_cookie: str = "") -> ImageCheck:
    public_check = fetch_image_metadata(url, timeout=timeout)
    if public_check.ok:
        public_check.public_ok = True
        return public_check

    if not auth_cookie:
        return public_check

    auth_check = fetch_image_metadata(url, timeout=timeout, cookie=auth_cookie)
    if auth_check.ok:
        auth_check.authenticated_ok = True
        auth_check.requires_auth = True
        auth_check.public_ok = False
        auth_check.reason = "requires_libri_authentication_not_suitable_as_tiktok_image_url"
        return auth_check

    return public_check


def contains_term(text: str, term: str) -> bool:
    if term.startswith(r"\b") or "\\" in term:
        return re.search(term, text, flags=re.I) is not None
    return term.lower() in text


def assess_product(product: ProductRecord, check_images: bool = True, auth_cookie: str = "") -> ProductRecord:
    reasons: list[str] = []
    full_text = " ".join(
        [
            product.title,
            product.subtitle,
            product.author,
            product.publisher,
            product.binding,
            product.product_group,
            product.blurb,
            product.author_bio,
        ]
    ).lower()

    allowed_group_terms = [
        "belletristik",
        "roman",
        "erzähl",
        "erzaehl",
        "fantasy",
        "science fiction",
        "krimi",
        "thriller",
        "jugendromane",
        "jugenderzählungen",
        "jugenderzaehlungen",
    ]
    if product.product_group and not any(term in product.product_group.lower() for term in allowed_group_terms):
        reasons.append("reject_category_not_supported_by_roman_template")

    for category, terms in REJECT_TERMS.items():
        for term in terms:
            if contains_term(full_text, term):
                reasons.append(f"reject_{category}:{term}")
                break

    hard_missing = []
    for field_name in ["title", "ean", "blurb"]:
        if not getattr(product, field_name):
            hard_missing.append(field_name)
    if product.price is None:
        hard_missing.append("price")
    if product.stock is None:
        hard_missing.append("stock")
    if product.stock is not None and product.stock <= 0:
        reasons.append("reject_out_of_stock")
    if product.weight_g is None:
        hard_missing.append("weight_g")
    if not product.image_urls:
        hard_missing.append("main_image")
    if hard_missing:
        reasons.append("missing_required_detail:" + ",".join(hard_missing))

    review_reasons = []
    for category, terms in REVIEW_TERMS.items():
        hits = [term for term in terms if contains_term(full_text, term)]
        if hits:
            review_reasons.append(f"review_{category}:{'|'.join(hits[:5])}")

    reject_before_image_checks = any(
        reason.startswith("reject_") or reason.startswith("missing_required_detail")
        for reason in reasons
    )
    if check_images and product.image_urls and not reject_before_image_checks:
        product.image_checks = [validate_image_url(url, auth_cookie=auth_cookie) for url in product.image_urls[:9]]
        if any(check.requires_auth for check in product.image_checks):
            review_reasons.append("review_image_requires_libri_login_public_hosting_or_media_center_needed")
        public_checks = [check for check in product.image_checks if check.ok and check.public_ok]
        uploadable_checks = [check for check in public_checks if check.tiktok_size_ok]
        if not public_checks:
            reasons.append("missing_reachable_image")
        elif not uploadable_checks:
            reasons.append(f"reject_image_below_{MIN_TIKTOK_IMAGE_SIDE_PX}x{MIN_TIKTOK_IMAGE_SIDE_PX}")

    if any(reason.startswith("reject_") or reason.startswith("missing_required_detail") for reason in reasons):
        product.status = "reject"
    elif review_reasons:
        product.status = "review"
    else:
        product.status = "green"

    product.reasons = list(dict.fromkeys(product.reasons + reasons + review_reasons))
    return product


def build_description(product: ProductRecord) -> str:
    parts = []
    if product.subtitle:
        parts.append(product.subtitle)
    if product.blurb:
        parts.append(product.blurb)

    details = []
    for label, value in [
        ("Autor", product.author),
        ("Verlag", product.publisher),
        ("Einband", product.binding),
        ("Sprache", product.language),
        ("Erscheinungsdatum", product.release_date),
        ("Seiten", str(product.pages) if product.pages else ""),
        ("ISBN/EAN", product.ean),
    ]:
        if value:
            details.append(f"{label}: {value}")
    if details:
        parts.append("Produktdetails:\n" + "\n".join(details))
    if product.author_bio:
        parts.append("Über den Autor:\n" + product.author_bio)

    description = "\n\n".join(parts)
    description = re.sub(r"https?://\S+", "", description)
    description = re.sub(r"\b(Bestseller|BookTok Bestseller|Trending Item)\b", "", description, flags=re.I)
    for pattern, replacement in DESCRIPTION_REPLACEMENTS:
        description = re.sub(pattern, replacement, description, flags=re.I)
    lines = [normalize_text(line) for line in description.splitlines()]
    description = "\n".join(line for line in lines if line)
    description = re.sub(r"\n{3,}", "\n\n", description)
    return description.strip()


def display_author(author: str) -> str:
    author = normalize_text(author)
    if "," not in author:
        return author
    last, first = [part.strip() for part in author.split(",", 1)]
    return normalize_text(f"{first} {last}") if first and last else author


def listing_genre(product: ProductRecord) -> str:
    text = normalize_text(product.product_group).lower()
    if "krimi" in text or "thriller" in text:
        return "Krimi Thriller"
    if "science fiction" in text and "fantasy" in text:
        return "Science-Fiction Fantasy Roman"
    if "fantasy" in text:
        return "Fantasy Roman"
    if "jugend" in text:
        return "Jugendroman"
    return "Roman"


def build_product_name(product: ProductRecord) -> str:
    pieces = [product.title]
    if product.author:
        pieces.append(product.author)
    name = " - ".join(piece for piece in pieces if piece)

    descriptors = [
        listing_genre(product),
        product.binding,
        product.language,
        f"{product.pages} Seiten" if product.pages else "",
    ]
    for descriptor in descriptors:
        descriptor = normalize_text(descriptor)
        if descriptor and descriptor.lower() not in name.lower():
            name = f"{name} - {descriptor}"
        if len(name) >= MIN_LISTING_QUALITY_TITLE_CHARS:
            break
    return name[:254]


def image_quality_score(check: ImageCheck) -> tuple[int, int, int]:
    width = check.width or 0
    height = check.height or 0
    area = width * height
    ratio = check.ratio or 0
    cover_ratio_bonus = 1 if 0.58 <= ratio <= 0.82 else 0
    square_bonus = 1 if 0.92 <= ratio <= 1.08 else 0
    return (cover_ratio_bonus, square_bonus, area)


def parcel_weight(product: ProductRecord, packing_weight_g: int) -> int:
    return max(1, int(product.weight_g or 0) + packing_weight_g)


def parcel_height(product: ProductRecord, min_height_cm: int) -> int:
    if not product.pages:
        return min_height_cm
    estimated_book_block = math.ceil(product.pages / 120)
    return max(min_height_cm, estimated_book_block + 1)


def load_template_map(workbook_path: Path) -> tuple[openpyxl.Workbook, dict[str, int]]:
    workbook = openpyxl.load_workbook(workbook_path)
    sheet = workbook["Template"]
    key_to_col = {
        str(sheet.cell(row=1, column=col).value): col
        for col in range(1, sheet.max_column + 1)
        if sheet.cell(row=1, column=col).value
    }
    missing = REQUIRED_UPLOAD_KEYS - set(key_to_col)
    if missing:
        raise ValueError(f"Template is missing required keys: {sorted(missing)}")
    return workbook, key_to_col


def clear_template_data(sheet) -> None:
    for row in range(TEMPLATE_DATA_START_ROW, min(TEMPLATE_MAX_DATA_ROW, sheet.max_row) + 1):
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row=row, column=col).value = None


def set_value(sheet, key_to_col: dict[str, int], row: int, key: str, value) -> None:
    if key in FORBIDDEN_TEMPLATE_KEYS:
        raise ValueError(f"Refusing to write forbidden template key: {key}")
    if key not in key_to_col:
        return
    sheet.cell(row=row, column=key_to_col[key]).value = value


def write_upload_workbook(
    template_path: Path,
    output_path: Path,
    products: list[ProductRecord],
    packing_weight_g: int,
    parcel_length_cm: int,
    parcel_width_cm: int,
    min_parcel_height_cm: int,
) -> None:
    workbook, key_to_col = load_template_map(template_path)
    sheet = workbook["Template"]
    clear_template_data(sheet)

    for offset, product in enumerate(products, start=0):
        row = TEMPLATE_DATA_START_ROW + offset
        public_image_checks = sorted(
            [
                check for check in product.image_checks
                if check.ok and check.public_ok and check.tiktok_size_ok
            ],
            key=image_quality_score,
            reverse=True,
        )
        main_image = next((check.url for check in public_image_checks), product.image_urls[0] if product.image_urls else "")
        extra_images = [check.url for check in public_image_checks if check.url != main_image]

        values = {
            "category": CATEGORY,
            "brand": BRAND_NO_BRAND,
            "product_name": build_product_name(product),
            "product_description": build_description(product),
            "main_image": main_image,
            "gtin_type": GTIN_TYPE,
            "gtin_code": product.ean,
            "parcel_weight": parcel_weight(product, packing_weight_g),
            "parcel_length": parcel_length_cm,
            "parcel_width": parcel_width_cm,
            "parcel_height": parcel_height(product, min_parcel_height_cm),
            "price": product.price,
            "quantity": product.stock,
            "seller_sku": f"LIBRI-{product.ean}",
            "manufacturer_ids": DEFAULT_MANUFACTURER_ID,
            "rp_ids": DEFAULT_RP_ID,
            "product_property/102277": DEFAULT_WARNING_ON_PACKAGE,
        }
        for image_index, url in enumerate(extra_images[:8], start=2):
            values[f"image_{image_index}"] = url

        for key, value in values.items():
            set_value(sheet, key_to_col, row, key, value)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def write_csv(path: Path, products: list[ProductRecord]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "status",
        "reasons",
        "rank",
        "title",
        "author",
        "ean",
        "price",
        "stock",
        "weight_g",
        "product_group",
        "main_image",
        "image_checks",
        "sources",
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        for product in products:
            writer.writerow(
                {
                    "status": product.status,
                    "reasons": "; ".join(product.reasons),
                    "rank": product.source_rank or "",
                    "title": product.title,
                    "author": product.author,
                    "ean": product.ean,
                    "price": product.price if product.price is not None else product.source_price or "",
                    "stock": product.stock if product.stock is not None else "",
                    "weight_g": product.weight_g if product.weight_g is not None else "",
                    "product_group": product.product_group,
                    "main_image": product.image_urls[0] if product.image_urls else "",
                    "image_checks": json.dumps([check.__dict__ for check in product.image_checks], ensure_ascii=False),
                    "sources": " | ".join(product.source_names),
                }
            )


def write_log(
    path: Path,
    products: list[ProductRecord],
    upload_path: Path,
    workbook_statuses: set[str],
    workbook_count: int,
) -> None:
    counts = {status: sum(1 for product in products if product.status == status) for status in ["green", "review", "reject"]}
    lines = [
        "# TikTok Shop Upload Log",
        "",
        f"Generated: {dt.datetime.now().isoformat(timespec='seconds')}",
        f"Upload workbook: {upload_path}",
        f"Workbook statuses: {', '.join(sorted(workbook_statuses))}",
        "",
        "## Counts",
        "",
        f"- green: {counts['green']}",
        f"- review: {counts['review']}",
        f"- reject: {counts['reject']}",
        f"- workbook rows: {workbook_count}",
        "",
        "## Next step",
        "",
        "Upload the generated XLSX in TikTok Seller Center > Products > Bulk listing only when the workbook contains publish-ready rows.",
        "Publish only after TikTok's pre-check reports no errors. If workbook statuses include review, treat the workbook as a draft for inspection.",
    ]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def discover_default_template(workspace: Path) -> Path:
    matches = sorted(workspace.glob("Tiktoksellercenter_*Roman_template.xlsx"))
    if not matches:
        raise FileNotFoundError("Could not find TikTok Roman template XLSX.")
    return matches[0]


def collect_inputs(args: argparse.Namespace) -> list[ProductRecord]:
    workspace = Path(args.workspace).resolve()
    products: list[ProductRecord] = []

    detail_globs = args.detail_glob or ["*.html"]
    bestseller_pdfs = args.bestseller_pdf or ["Mein.Libri - Bestellen - Bestseller.pdf"]

    for pattern in detail_globs:
        for path in sorted(workspace.glob(pattern)):
            products.append(parse_libri_detail_html(path))

    for pdf in bestseller_pdfs:
        path = (workspace / pdf).resolve() if not Path(pdf).is_absolute() else Path(pdf)
        if path.exists():
            products.extend(parse_bestseller_pdf(path))

    for csv_path in args.manual_csv:
        path = (workspace / csv_path).resolve() if not Path(csv_path).is_absolute() else Path(csv_path)
        if path.exists():
            products.extend(parse_manual_csv(path))

    return merge_products(products)


def run(args: argparse.Namespace) -> int:
    workspace = Path(args.workspace).resolve()
    env = load_env_file(workspace / ENV_FILE_NAME)
    auth_cookie = env.get("LIBRI_COOKIE", "")
    template_path = Path(args.template).resolve() if args.template else discover_default_template(workspace)
    timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = (Path(args.output_dir) if args.output_dir else workspace / "outputs" / timestamp).resolve()

    products = collect_inputs(args)
    if not products:
        raise SystemExit("No products found. Add saved Libri product HTML files, a bestseller PDF, or a manual CSV.")

    assessed = [
        assess_product(product, check_images=not args.skip_image_checks, auth_cookie=auth_cookie)
        for product in products
    ]
    assessed.sort(
        key=lambda product: (
            product.source_rank is None,
            product.source_rank or 999999,
            product.status != "green",
            product.title,
        )
    )

    workbook_statuses = {status.strip() for status in args.workbook_statuses.split(",") if status.strip()}
    workbook_candidates = [product for product in assessed if product.status in workbook_statuses]
    workbook_products = workbook_candidates[args.skip : args.skip + args.limit]
    review = [product for product in assessed if product.status == "review"]
    reject = [product for product in assessed if product.status == "reject"]

    workbook_name = "tiktok_upload_green.xlsx" if workbook_statuses == {"green"} else "tiktok_upload_draft.xlsx"
    upload_path = output_dir / workbook_name
    write_upload_workbook(
        template_path=template_path,
        output_path=upload_path,
        products=workbook_products,
        packing_weight_g=args.packing_weight_g,
        parcel_length_cm=args.parcel_length_cm,
        parcel_width_cm=args.parcel_width_cm,
        min_parcel_height_cm=args.min_parcel_height_cm,
    )
    write_csv(output_dir / "candidate_report.csv", assessed)
    write_csv(output_dir / "review_hold.csv", review)
    write_csv(output_dir / "rejects.csv", reject)
    write_log(output_dir / "upload_log.md", assessed, upload_path, workbook_statuses, len(workbook_products))

    print(f"Output directory: {output_dir}")
    print(f"Upload workbook: {upload_path}")
    green_count = sum(1 for product in assessed if product.status == "green")
    print(f"Products: green={green_count} review={len(review)} reject={len(reject)} workbook={len(workbook_products)} total={len(assessed)}")
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Create TikTok Shop bulk upload XLSX files from Libri data.")
    parser.add_argument("--workspace", default=".", help="Workspace folder containing the source files.")
    parser.add_argument("--template", default="", help="TikTok Shop Roman template XLSX. Defaults to Tiktoksellercenter_*Roman_template.xlsx.")
    parser.add_argument("--detail-glob", action="append", default=None, help="Glob for saved Libri product detail HTML files, relative to workspace. Can be repeated. Defaults to *.html.")
    parser.add_argument("--bestseller-pdf", action="append", default=None, help="Saved Libri bestseller PDF, relative to workspace. Can be repeated. Defaults to the local Mein.Libri bestseller PDF.")
    parser.add_argument("--manual-csv", action="append", default=[], help="Optional UTF-8 CSV with product fields. Can be repeated.")
    parser.add_argument("--output-dir", default="", help="Output directory. Defaults to outputs/<timestamp>.")
    parser.add_argument("--limit", type=int, default=100, help="Maximum green products to write to the upload workbook.")
    parser.add_argument("--skip", type=int, default=0, help="Skip this many eligible products before writing workbook rows.")
    parser.add_argument("--workbook-statuses", default="green", help="Comma-separated statuses to write into the workbook. Default: green. Use green,review for a non-publish draft.")
    parser.add_argument("--packing-weight-g", type=int, default=50, help="Packaging weight added to Libri product weight.")
    parser.add_argument("--parcel-length-cm", type=int, default=25, help="Default parcel length in cm.")
    parser.add_argument("--parcel-width-cm", type=int, default=18, help="Default parcel width in cm.")
    parser.add_argument("--min-parcel-height-cm", type=int, default=4, help="Minimum parcel height in cm; page count can increase it.")
    parser.add_argument("--skip-image-checks", action="store_true", help="Skip online image reachability and dimension checks.")
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    return run(args)


if __name__ == "__main__":
    raise SystemExit(main())
