#!/usr/bin/env python3
"""Build a controlled repair pack for TikTok listing-quality suggestions.

The pack is aimed at already-created products: it maps each LIBRI SKU/EAN to
safe title improvements, keeps the front-cover image as the primary image, and
creates local white-background cover images for manual/TikTok Media Center
upload when the Seller Center complains about image background quality.
"""

from __future__ import annotations

import argparse
import csv
import base64
import html
import io
import re
import sys
import urllib.request
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from PIL import Image, ImageFilter, ImageOps

sys.path.insert(0, str(Path(__file__).resolve().parent))
from tiktok_libri_pipeline import (  # noqa: E402
    MIN_LISTING_QUALITY_TITLE_CHARS,
    build_product_name,
    fetch_image_metadata,
    image_quality_score,
    parse_libri_detail_html,
)


DATA_START_ROW = 7
IMAGE_KEYS = ["main_image"] + [f"image_{idx}" for idx in range(2, 10)]
REFERENCE_MATCH_MAX_DISTANCE = 0.34
REFERENCE_MATCH_MAX_MAE = 0.32
MAX_IMAGES_TO_COMPARE = 8


@dataclass
class CoverSelection:
    url: str
    checks: list
    method: str = "no_public_image"
    confidence: str = "low"
    hash_distance: float | None = None
    pixel_mae: float | None = None
    likely_back_cover: bool = False
    generated_source: str = "selected_url"
    reference_image: Image.Image | None = None
    selected_librinet_role: int | None = None
    review_reason: str = ""


def load_headers(sheet) -> dict[str, int]:
    return {
        str(sheet.cell(row=1, column=col).value): col
        for col in range(1, sheet.max_column + 1)
        if sheet.cell(row=1, column=col).value
    }


def extract_ean_from_text(value: str) -> str:
    for match in re.findall(r"(?:97[89])[\d\-\s]{10,20}", value or ""):
        digits = re.sub(r"\D", "", match)
        if len(digits) == 13:
            return digits
    return ""


def read_upload_workbook(path: Path) -> list[dict[str, str]]:
    workbook = openpyxl.load_workbook(path, data_only=False, read_only=False)
    sheet = workbook["Template"]
    headers = load_headers(sheet)
    rows: list[dict[str, str]] = []
    for row_idx in range(DATA_START_ROW, sheet.max_row + 1):
        product_name = sheet.cell(row_idx, headers["product_name"]).value
        if not product_name:
            continue
        description = str(sheet.cell(row_idx, headers.get("product_description", 1)).value or "")
        gtin = str(sheet.cell(row_idx, headers.get("gtin_code", 1)).value or "")
        ean = gtin if gtin else extract_ean_from_text(f"{product_name} {description}")
        image_urls = []
        for key in IMAGE_KEYS:
            col = headers.get(key)
            value = sheet.cell(row_idx, col).value if col else ""
            if value:
                image_urls.append(str(value))
        rows.append(
            {
                "source_workbook": str(path),
                "source_row": str(row_idx),
                "product_id": str(sheet.cell(row_idx, headers.get("product_id", 1)).value or ""),
                "product_name": str(product_name),
                "ean": ean,
                "seller_sku": str(sheet.cell(row_idx, headers.get("seller_sku", 1)).value or ""),
                "main_image": str(sheet.cell(row_idx, headers.get("main_image", 1)).value or ""),
                "image_urls": "|".join(dict.fromkeys(image_urls)),
            }
        )
    return rows


def download_image(url: str) -> Image.Image:
    request = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(request, timeout=30) as response:
        data = response.read(5_000_000)
    image = Image.open(io.BytesIO(data))
    image.load()
    return image


def librinet_media_role(url: str) -> int | None:
    """Return Libri media role number from medias.librinet URLs.

    In the current Libri pages we have inspected, role 1 is the front-cover
    asset and role 2 is a back-cover/detail asset. Role 2 must never be used as
    TikTok's main image because TikTok requires the physical front view.
    """
    decoded = html.unescape(url)
    decoded = decoded.replace("\\/", "/")
    match = re.search(r"medias\.librinet\.de/dl/[^/]+/(\d+)/", decoded, flags=re.I)
    if not match:
        return None
    return int(match.group(1))


def is_librinet_back_url(url: str) -> bool:
    return librinet_media_role(url) == 2


def extract_reference_cover_image(detail_path: Path) -> Image.Image | None:
    if not detail_path.exists():
        return None
    raw_html = detail_path.read_text(encoding="utf-8", errors="replace")
    img_tags = re.findall(r"<img\b[^>]*>", raw_html, flags=re.I | re.S)
    ordered_tags = sorted(
        img_tags,
        key=lambda tag: 0 if re.search(r'class="[^"]*\bdetail-image\b[^"]*"', tag, flags=re.I) else 1,
    )
    for tag in ordered_tags:
        if "data:image/" not in tag:
            continue
        src_match = re.search(r'src="data:image/[^;]+;base64,([^"]+)"', tag, flags=re.I | re.S)
        if not src_match:
            continue
        try:
            data = base64.b64decode(re.sub(r"\s+", "", src_match.group(1)))
            image = Image.open(io.BytesIO(data))
            image.load()
            return ImageOps.exif_transpose(image).convert("RGB")
        except Exception:  # noqa: BLE001 - try next embedded image if Libri changes markup
            continue
    return None


def trim_plain_border(image: Image.Image) -> Image.Image:
    image = ImageOps.exif_transpose(image).convert("RGB")
    width, height = image.size
    if width < 20 or height < 20:
        return image
    corners = [
        image.getpixel((0, 0)),
        image.getpixel((width - 1, 0)),
        image.getpixel((0, height - 1)),
        image.getpixel((width - 1, height - 1)),
    ]
    background = tuple(sorted(channel)[len(channel) // 2] for channel in zip(*corners))
    threshold = 22
    mask = Image.new("L", image.size, 0)
    pixels = image.load()
    mask_pixels = mask.load()
    for y in range(height):
        for x in range(width):
            pixel = pixels[x, y]
            distance = sum(abs(pixel[idx] - background[idx]) for idx in range(3))
            if distance > threshold:
                mask_pixels[x, y] = 255
    bbox = mask.getbbox()
    if not bbox:
        return image
    left, top, right, bottom = bbox
    bbox_area = (right - left) * (bottom - top)
    image_area = width * height
    if bbox_area < image_area * 0.08:
        return image
    return image.crop(bbox)


def prepared_for_comparison(image: Image.Image) -> Image.Image:
    return trim_plain_border(image).resize((64, 64), Image.Resampling.LANCZOS).convert("RGB")


def difference_hash(image: Image.Image, hash_size: int = 16) -> int:
    resized = trim_plain_border(image).convert("L").resize((hash_size + 1, hash_size), Image.Resampling.LANCZOS)
    pixels = list(resized.getdata())
    value = 0
    bit = 0
    for row in range(hash_size):
        offset = row * (hash_size + 1)
        for col in range(hash_size):
            if pixels[offset + col] > pixels[offset + col + 1]:
                value |= 1 << bit
            bit += 1
    return value


def normalized_hash_distance(left: Image.Image, right: Image.Image) -> float:
    left_hash = difference_hash(left)
    right_hash = difference_hash(right)
    return (left_hash ^ right_hash).bit_count() / 256


def normalized_pixel_mae(left: Image.Image, right: Image.Image) -> float:
    left_prepared = prepared_for_comparison(left)
    right_prepared = prepared_for_comparison(right)
    left_data = list(left_prepared.getdata())
    right_data = list(right_prepared.getdata())
    diff = 0
    for left_pixel, right_pixel in zip(left_data, right_data):
        diff += sum(abs(left_pixel[idx] - right_pixel[idx]) for idx in range(3))
    return diff / (len(left_data) * 3 * 255)


def edge_density(image: Image.Image) -> float:
    gray = trim_plain_border(image).convert("L").resize((240, 240), Image.Resampling.LANCZOS)
    edges = gray.filter(ImageFilter.FIND_EDGES)
    values = list(edges.getdata())
    return sum(1 for value in values if value > 48) / len(values)


def likely_back_cover(
    image: Image.Image,
    reference_image: Image.Image | None,
    hash_distance: float | None,
    pixel_mae: float | None,
    source_url: str = "",
) -> bool:
    if is_librinet_back_url(source_url):
        return True
    if reference_image and hash_distance is not None and pixel_mae is not None:
        if hash_distance <= REFERENCE_MATCH_MAX_DISTANCE or pixel_mae <= REFERENCE_MATCH_MAX_MAE:
            return False
    ratio = image.width / image.height if image.height else 0
    return 0.55 <= ratio <= 0.82 and edge_density(image) > 0.245


def create_white_cover_image_from_image(image: Image.Image, output_path: Path, canvas_size: int) -> tuple[int, int, int]:
    image = trim_plain_border(ImageOps.exif_transpose(image)).convert("RGBA")
    canvas = Image.new("RGB", (canvas_size, canvas_size), "white")
    max_width = int(canvas_size * 0.86)
    max_height = int(canvas_size * 0.94)
    scale = min(max_width / image.width, max_height / image.height)
    new_size = (max(1, int(image.width * scale)), max(1, int(image.height * scale)))
    image = image.resize(new_size, Image.Resampling.LANCZOS)
    x = (canvas_size - image.width) // 2
    y = (canvas_size - image.height) // 2
    white_backing = Image.new("RGBA", image.size, "white")
    white_backing.alpha_composite(image)
    canvas.paste(white_backing.convert("RGB"), (x, y))
    output_path.parent.mkdir(parents=True, exist_ok=True)
    canvas.save(output_path, format="JPEG", quality=95, optimize=True)
    return canvas.width, canvas.height, output_path.stat().st_size


def create_white_cover_image(source_url: str, output_path: Path, canvas_size: int) -> tuple[int, int, int]:
    return create_white_cover_image_from_image(download_image(source_url), output_path, canvas_size)


def is_cover_like(check) -> bool:
    ratio = check.ratio or 0
    return 0.55 <= ratio <= 0.82


def image_is_single_cover_like(image: Image.Image) -> bool:
    trimmed = trim_plain_border(image)
    ratio = trimmed.width / trimmed.height if trimmed.height else 0
    area_share = (trimmed.width * trimmed.height) / (image.width * image.height)
    return 0.50 <= ratio <= 0.90 and area_share >= 0.08


def select_cover_url(
    urls: list[str],
    reference_image: Image.Image | None = None,
    current_main_image: str = "",
    reference_image_safe: bool = True,
) -> CoverSelection:
    reference_for_matching = reference_image if reference_image_safe else None
    checks = []
    metadata_candidates = []
    candidates = []
    for url in dict.fromkeys(urls):
        if not url.startswith(("http://", "https://")):
            continue
        check = fetch_image_metadata(url)
        if check.ok and check.tiktok_size_ok:
            checks.append(check)
            metadata_candidates.append((url, check))

    front_metadata = [(url, check) for url, check in metadata_candidates if librinet_media_role(url) == 1 and is_cover_like(check)]
    if front_metadata:
        selected_url, selected_check = sorted(
            front_metadata,
            key=lambda item: (
                item[0] == current_main_image,
                is_cover_like(item[1]),
                image_quality_score(item[1])[2],
            ),
            reverse=True,
        )[0]
        return CoverSelection(
            url=selected_url,
            checks=checks,
            method="libri_front_media_role",
            confidence="high",
            reference_image=reference_for_matching,
            selected_librinet_role=1,
        )

    prioritized_metadata = sorted(
        metadata_candidates,
        key=lambda item: (
            librinet_media_role(item[0]) == 1,
            item[0] == current_main_image,
            is_cover_like(item[1]),
            image_quality_score(item[1])[2],
        ),
        reverse=True,
    )[:MAX_IMAGES_TO_COMPARE]

    for url, check in prioritized_metadata:
        try:
            image = download_image(url)
        except Exception:  # noqa: BLE001 - metadata passed, but image body can still fail
            continue
        hash_distance = None
        pixel_mae = None
        if reference_for_matching:
            hash_distance = normalized_hash_distance(reference_for_matching, image)
            pixel_mae = normalized_pixel_mae(reference_for_matching, image)
        role = librinet_media_role(url)
        back = likely_back_cover(image, reference_for_matching, hash_distance, pixel_mae, url)
        cover_like = is_cover_like(check) or image_is_single_cover_like(image)
        reference_match = bool(
            reference_for_matching
            and role != 2
            and hash_distance is not None
            and pixel_mae is not None
            and (hash_distance <= REFERENCE_MATCH_MAX_DISTANCE or pixel_mae <= REFERENCE_MATCH_MAX_MAE)
        )
        public_quality = image_quality_score(check)
        score = public_quality[2] / 10_000_000
        if cover_like:
            score += 1.0
        if role == 1:
            score += 6.0
        elif role == 2:
            score -= 12.0
        if url == current_main_image:
            score += 0.35
        if reference_match:
            score += 10.0 + (1 - min(hash_distance or 1, 1)) * 2 + (1 - min(pixel_mae or 1, 1))
        elif reference_for_matching:
            score -= 2.5
        if back:
            score -= 3.5
        candidates.append(
            {
                "url": url,
                "check": check,
                "score": score,
                "hash_distance": hash_distance,
                "pixel_mae": pixel_mae,
                "likely_back_cover": back,
                "reference_match": reference_match,
                "cover_like": cover_like,
                "librinet_role": role,
                "image": image,
            }
        )

    if not candidates:
        if reference_image and reference_image_safe:
            return CoverSelection(
                url="",
                checks=checks,
                method="use_embedded_reference_cover",
                confidence="medium",
                reference_image=reference_image,
                generated_source="libri_reference_cover_low_res",
                review_reason="no_public_cover_url_use_embedded_cover",
            )
        return CoverSelection(url="", checks=checks, reference_image=reference_for_matching)

    role_2_candidates = [candidate for candidate in candidates if candidate["librinet_role"] == 2]
    if role_2_candidates and current_main_image:
        for candidate in candidates:
            if candidate["url"] != current_main_image or candidate["librinet_role"] is not None:
                continue
            matches_back_asset = False
            for back_candidate in role_2_candidates:
                try:
                    back_hash_distance = normalized_hash_distance(candidate["image"], back_candidate["image"])
                    back_pixel_mae = normalized_pixel_mae(candidate["image"], back_candidate["image"])
                except Exception:  # noqa: BLE001 - keep conservative if comparison fails
                    continue
                if back_hash_distance <= REFERENCE_MATCH_MAX_DISTANCE or back_pixel_mae <= REFERENCE_MATCH_MAX_MAE:
                    matches_back_asset = True
                    break
            if matches_back_asset:
                candidate["likely_back_cover"] = True
                candidate["score"] -= 12.0

    safe_candidates = [
        candidate
        for candidate in candidates
        if not candidate["likely_back_cover"] and candidate["librinet_role"] != 2
    ]
    front_role_candidates = [
        candidate for candidate in safe_candidates if candidate["librinet_role"] == 1 and candidate["cover_like"]
    ]
    if front_role_candidates:
        selected = sorted(front_role_candidates, key=lambda candidate: candidate["score"], reverse=True)[0]
        return CoverSelection(
            url=selected["url"],
            checks=checks,
            method="libri_front_media_role",
            confidence="high",
            hash_distance=selected["hash_distance"],
            pixel_mae=selected["pixel_mae"],
            likely_back_cover=selected["likely_back_cover"],
            reference_image=reference_for_matching,
            selected_librinet_role=selected["librinet_role"],
        )

    reference_matches = [candidate for candidate in safe_candidates if candidate["reference_match"] and candidate["cover_like"]]
    if reference_matches:
        selected = sorted(reference_matches, key=lambda candidate: candidate["score"], reverse=True)[0]
        return CoverSelection(
            url=selected["url"],
            checks=checks,
            method="libri_reference_match",
            confidence="high",
            hash_distance=selected["hash_distance"],
            pixel_mae=selected["pixel_mae"],
            likely_back_cover=selected["likely_back_cover"],
            reference_image=reference_for_matching,
            selected_librinet_role=selected["librinet_role"],
        )

    if reference_image and reference_image_safe:
        return CoverSelection(
            url="",
            checks=checks,
            method="use_embedded_reference_cover",
            confidence="medium",
            reference_image=reference_image,
            generated_source="libri_reference_cover_low_res",
            review_reason="no_public_flat_front_cover_use_embedded_cover",
        )

    if current_main_image:
        current = next((candidate for candidate in candidates if candidate["url"] == current_main_image), None)
        if current and not current["likely_back_cover"] and current["librinet_role"] != 2:
            return CoverSelection(
                url=current["url"],
                checks=checks,
                method="keep_current_no_reference_match",
                confidence="medium" if not current["likely_back_cover"] else "low",
                hash_distance=current["hash_distance"],
                pixel_mae=current["pixel_mae"],
                likely_back_cover=current["likely_back_cover"],
                reference_image=reference_for_matching,
                selected_librinet_role=current["librinet_role"],
            )

    if not safe_candidates:
        return CoverSelection(
            url="",
            checks=checks,
            method="no_safe_front_cover",
            confidence="low",
            likely_back_cover=any(candidate["likely_back_cover"] or candidate["librinet_role"] == 2 for candidate in candidates),
            reference_image=reference_for_matching,
            review_reason="only_back_cover_or_unsafe_images_available",
        )

    cover_like_candidates = [candidate for candidate in safe_candidates if candidate["cover_like"]]
    ranked = sorted(cover_like_candidates or safe_candidates, key=lambda candidate: candidate["score"], reverse=True)
    selected = ranked[0]
    return CoverSelection(
        url=selected["url"],
        checks=checks,
        method="fallback_cover_shape_quality",
        confidence="medium" if not reference_image and not selected["likely_back_cover"] else "low",
        hash_distance=selected["hash_distance"],
        pixel_mae=selected["pixel_mae"],
        likely_back_cover=selected["likely_back_cover"],
        reference_image=reference_for_matching,
        selected_librinet_role=selected["librinet_role"],
    )


def build_pack(args: argparse.Namespace) -> int:
    output_dir = Path(args.output_dir)
    image_dir = output_dir / "white_cover_main_images"
    workbook_rows: list[dict[str, str]] = []
    for workbook_path in args.workbook:
        workbook_rows.extend(read_upload_workbook(Path(workbook_path)))

    by_ean: dict[str, dict[str, str]] = {}
    for row in workbook_rows:
        ean = row["ean"]
        if ean and ean not in by_ean:
            by_ean[ean] = row

    report_rows = []
    for idx, row in enumerate(by_ean.values(), start=1):
        ean = row["ean"]
        detail_path = Path(args.detail_dir) / f"{ean}.html"
        product = parse_libri_detail_html(detail_path) if detail_path.exists() else None
        reference_image = extract_reference_cover_image(detail_path)
        # The embedded Libri detail image is often the only true front cover when
        # public media role 2 contains back covers or publisher marketing images.
        # Trust it only if its own geometry looks like one flat cover.
        reference_image_safe = bool(reference_image) and image_is_single_cover_like(reference_image)
        suggested_name = build_product_name(product) if product else row["product_name"]
        candidate_urls = []
        if product:
            candidate_urls.extend(product.image_urls)
        candidate_urls.extend(row["image_urls"].split("|"))
        cover_selection = select_cover_url(
            candidate_urls,
            reference_image=reference_image,
            current_main_image=row["main_image"],
            reference_image_safe=reference_image_safe,
        )
        cover_url = cover_selection.url
        image_checks = cover_selection.checks

        generated_path = ""
        generated_size = ""
        generated_width = ""
        generated_height = ""
        generated_source = cover_selection.generated_source
        image_action = "no_cover_url_found"
        if cover_url:
            image_action = "main_image_is_cover" if row["main_image"] == cover_url else "set_main_image_to_cover"
            if args.generate_images:
                output_path = image_dir / f"{ean}_main_cover_white.jpg"
                try:
                    width, height, size = create_white_cover_image(cover_url, output_path, args.canvas_size)
                    generated_path = str(output_path.resolve())
                    generated_width = str(width)
                    generated_height = str(height)
                    generated_size = str(size)
                    image_action = "upload_white_cover_as_main_image"
                except Exception as exc:  # noqa: BLE001 - report per product, keep batch running
                    image_action = f"cover_generation_failed:{type(exc).__name__}"
        elif reference_image and args.generate_images and cover_selection.method == "use_embedded_reference_cover":
            output_path = image_dir / f"{ean}_main_cover_white.jpg"
            try:
                width, height, size = create_white_cover_image_from_image(reference_image, output_path, args.canvas_size)
                generated_path = str(output_path.resolve())
                generated_width = str(width)
                generated_height = str(height)
                generated_size = str(size)
                generated_source = "libri_reference_cover_low_res"
                image_action = "upload_reference_cover_as_main_image"
            except Exception as exc:  # noqa: BLE001
                image_action = f"reference_cover_generation_failed:{type(exc).__name__}"

        available_public_images = len(image_checks)
        current_images = [url for url in row["image_urls"].split("|") if url]
        title_action = (
            "update_title_40_plus"
            if len(row["product_name"]) < MIN_LISTING_QUALITY_TITLE_CHARS and suggested_name != row["product_name"]
            else "keep_title"
        )
        report_rows.append(
            {
                "sequence": idx,
                "product_id": row["product_id"],
                "seller_sku": row["seller_sku"],
                "ean": ean,
                "current_product_name": row["product_name"],
                "current_title_chars": len(row["product_name"]),
                "suggested_product_name": suggested_name,
                "suggested_title_chars": len(suggested_name),
                "title_action": title_action,
                "current_main_image": row["main_image"],
                "selected_cover_url": cover_url,
                "main_image_action": image_action,
                "cover_selection_method": cover_selection.method,
                "cover_confidence": cover_selection.confidence,
                "cover_reference_hash_distance": f"{cover_selection.hash_distance:.4f}" if cover_selection.hash_distance is not None else "",
                "cover_reference_pixel_mae": f"{cover_selection.pixel_mae:.4f}" if cover_selection.pixel_mae is not None else "",
                "likely_back_cover": "yes" if cover_selection.likely_back_cover else "no",
                "selected_librinet_role": cover_selection.selected_librinet_role or "",
                "libri_media_roles": "|".join(str(role) for role in libri_media_roles if role),
                "reference_image_safe": "yes" if reference_image_safe else "no",
                "cover_review_reason": cover_selection.review_reason,
                "generated_main_image_local": generated_path,
                "generated_main_image_source": generated_source,
                "generated_image_width": generated_width,
                "generated_image_height": generated_height,
                "generated_image_bytes": generated_size,
                "current_image_count": len(current_images),
                "available_public_image_count": available_public_images,
                "needs_real_extra_photos": "yes" if available_public_images < 5 else "no",
                "source_workbook": row["source_workbook"],
                "source_row": row["source_row"],
            }
        )

    output_dir.mkdir(parents=True, exist_ok=True)
    report_path = output_dir / "listing_quality_update_plan.csv"
    with report_path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=list(report_rows[0].keys()))
        writer.writeheader()
        writer.writerows(report_rows)

    title_path = output_dir / "title_updates.csv"
    with title_path.open("w", encoding="utf-8-sig", newline="") as fh:
        fieldnames = ["seller_sku", "ean", "product_name"]
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        for row in report_rows:
            if row["title_action"] == "update_title_40_plus":
                writer.writerow(
                    {
                        "seller_sku": row["seller_sku"],
                        "ean": row["ean"],
                        "product_name": row["suggested_product_name"],
                    }
                )

    image_path = output_dir / "main_image_upload_queue.csv"
    with image_path.open("w", encoding="utf-8-sig", newline="") as fh:
        fieldnames = ["seller_sku", "ean", "product_name", "local_image_path", "cover_url", "generated_source", "cover_confidence", "action"]
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        for row in report_rows:
            if row["generated_main_image_local"]:
                writer.writerow(
                    {
                        "seller_sku": row["seller_sku"],
                        "ean": row["ean"],
                        "product_name": row["suggested_product_name"],
                        "local_image_path": row["generated_main_image_local"],
                        "cover_url": row["selected_cover_url"],
                        "generated_source": row["generated_main_image_source"],
                        "cover_confidence": row["cover_confidence"],
                        "action": "upload as primary image; keep as front-cover main image",
                    }
                )

    print(f"Products: {len(report_rows)}")
    print(f"Title updates: {sum(row['title_action'] == 'update_title_40_plus' for row in report_rows)}")
    print(f"Generated main images: {sum(bool(row['generated_main_image_local']) for row in report_rows)}")
    print(f"Need real extra photos for 5+ images: {sum(row['needs_real_extra_photos'] == 'yes' for row in report_rows)}")
    print(f"Report: {report_path.resolve()}")
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Build TikTok listing-quality update plan and cover image assets.")
    parser.add_argument("--workbook", action="append", required=True, help="Upload workbook used for current products.")
    parser.add_argument("--detail-dir", default="libri_bulk_pages")
    parser.add_argument("--output-dir", default="outputs/current_listing_quality_updates")
    parser.add_argument("--canvas-size", type=int, default=1200)
    parser.add_argument("--generate-images", action=argparse.BooleanOptionalAction, default=True)
    return parser


def main(argv: list[str] | None = None) -> int:
    return build_pack(build_parser().parse_args(argv))


if __name__ == "__main__":
    raise SystemExit(main())
