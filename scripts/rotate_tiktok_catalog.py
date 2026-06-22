#!/usr/bin/env python3
"""Rotate older TikTok Shop Libri titles into new prepared titles.

This script intentionally separates catalog rotation from daily inventory sync:
- new titles come from a prepared TikTok upload workbook,
- old titles are selected explicitly or by low current Libri stock,
- live writes require --live.

Default behavior is a dry run that writes an audit plan only.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import html
import json
import mimetypes
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from fetch_libri_product_pages import load_env_file, login  # noqa: E402
from tiktok_order_automation import TikTokApiError, clean, compact_json  # noqa: E402
from update_tiktok_quantities_from_libri import (  # noqa: E402
    DEFAULT_SELLER_SKU_PREFIX,
    TikTokInventoryClient,
    TikTokSku,
    effective_warehouse_id,
    extract_ean,
    fetch_libri_stock,
    product_skus,
    read_local_libri_stock,
    response_errors,
    seller_sku_for_ean,
)


DEFAULT_OUTPUT_ROOT = Path("outputs") / "catalog_rotation"
DEFAULT_CATEGORY_ID = "987016"
DEFAULT_RESPONSIBLE_PERSON_ID = "69e8ebf9401201b6c3954a0f"
DEFAULT_WARNING_ATTRIBUTE_ID = "102277"
DEFAULT_WARNING_NO_VALUE_ID = "1000059"
DEFAULT_CURRENCY = "EUR"
DATA_START_ROW = 7
LOG_FIELDS = [
    "timestamp_utc",
    "phase",
    "status",
    "ean",
    "seller_sku",
    "product_id",
    "sku_id",
    "warehouse_id",
    "title",
    "old_quantity",
    "libri_quantity",
    "new_quantity",
    "message",
]


@dataclass
class NewProductRow:
    row: int
    title: str
    description: str
    ean: str
    price: str
    quantity: str
    seller_sku: str
    parcel_weight_g: str
    parcel_length_cm: str
    parcel_width_cm: str
    parcel_height_cm: str
    image_urls: list[str]


@dataclass
class RetireCandidate:
    sku: TikTokSku
    libri_quantity: int | None = None
    reason: str = ""
    libri_page: str = ""


class CatalogRotationClient(TikTokInventoryClient):
    def upload_product_image(self, image_bytes: bytes, filename: str, use_case: str = "MAIN_IMAGE") -> dict[str, Any]:
        path = "/product/202309/images/upload"
        body, content_type = multipart_body(
            fields={"use_case": use_case},
            file_field="data",
            filename=filename,
            file_bytes=image_bytes,
            content_type=mimetypes.guess_type(filename)[0] or "image/jpeg",
        )
        query: dict[str, str] = {
            "app_key": self.app_key,
            "timestamp": str(int(time.time())),
        }
        if self.include_version:
            query["version"] = self.version
        if self.access_token_in_query:
            query["access_token"] = self.access_token
        query["sign"] = self.sign(path, query, body_text="", include_body=False)

        request = urllib.request.Request(
            self.base_url + path + "?" + urllib.parse.urlencode(query),
            data=body,
            method="POST",
            headers={
                "Content-Type": content_type,
                "x-tts-access-token": self.access_token,
                "User-Agent": "TikTokShop-Libri-Catalog-Rotation/1.0",
            },
        )
        try:
            with urllib.request.urlopen(request, timeout=90) as response:
                raw = response.read().decode("utf-8", errors="replace")
        except urllib.error.HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace")
            raise TikTokApiError(f"TikTok image upload HTTP {exc.code}: {detail[:1000]}") from exc
        except urllib.error.URLError as exc:
            raise TikTokApiError(f"TikTok image upload connection failed: {exc}") from exc

        result = json.loads(raw)
        code = clean(result.get("code"))
        if code and code not in {"0", "OK", "SUCCESS"}:
            raise TikTokApiError(f"TikTok image upload error {code}: {clean(result.get('message')) or result}")
        return result.get("data") or {}

    def create_product(self, payload: dict[str, Any]) -> dict[str, Any]:
        return self.request(
            "POST",
            "/product/202309/products",
            params={"shop_cipher": self.ensure_shop_cipher()},
            body=payload,
        )


def multipart_body(
    fields: dict[str, str],
    file_field: str,
    filename: str,
    file_bytes: bytes,
    content_type: str,
) -> tuple[bytes, str]:
    boundary = f"----codex-tiktok-{uuid.uuid4().hex}"
    chunks: list[bytes] = []
    for key, value in fields.items():
        chunks.append(f"--{boundary}\r\n".encode("utf-8"))
        chunks.append(f'Content-Disposition: form-data; name="{key}"\r\n\r\n'.encode("utf-8"))
        chunks.append(str(value).encode("utf-8"))
        chunks.append(b"\r\n")
    chunks.append(f"--{boundary}\r\n".encode("utf-8"))
    chunks.append(
        (
            f'Content-Disposition: form-data; name="{file_field}"; filename="{filename}"\r\n'
            f"Content-Type: {content_type}\r\n\r\n"
        ).encode("utf-8")
    )
    chunks.append(file_bytes)
    chunks.append(b"\r\n")
    chunks.append(f"--{boundary}--\r\n".encode("utf-8"))
    return b"".join(chunks), f"multipart/form-data; boundary={boundary}"


def now_utc() -> str:
    return dt.datetime.now(dt.timezone.utc).isoformat(timespec="seconds")


def dedupe(values: Iterable[str]) -> list[str]:
    return [value for value in dict.fromkeys(clean(value) for value in values if clean(value))]


def sniff_csv(text: str) -> csv.Dialect:
    try:
        return csv.Sniffer().sniff(text[:4096], delimiters=",;\t")
    except csv.Error:
        return csv.excel


def discover_latest_workbook() -> Path:
    matches = sorted(
        Path("outputs").glob("**/tiktok_upload_green.xlsx"),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    if not matches:
        raise FileNotFoundError("No tiktok_upload_green.xlsx found under outputs/. Pass --new-workbook.")
    return matches[0]


def detect_headers(sheet) -> dict[str, int]:
    headers = {
        clean(sheet.cell(row=1, column=col).value): col
        for col in range(1, sheet.max_column + 1)
        if clean(sheet.cell(row=1, column=col).value)
    }
    required = {"product_name", "product_description", "main_image", "gtin_code", "price", "quantity", "seller_sku"}
    missing = required - set(headers)
    if missing:
        raise ValueError(f"Workbook is missing required TikTok columns: {sorted(missing)}")
    return headers


def cell(sheet, headers: dict[str, int], row_idx: int, key: str) -> str:
    col = headers.get(key)
    return clean(sheet.cell(row=row_idx, column=col).value) if col else ""


def load_new_rows(path: Path, start_row: int, max_rows: int) -> list[NewProductRow]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Template"] if "Template" in workbook.sheetnames else workbook.active
    headers = detect_headers(sheet)
    rows: list[NewProductRow] = []
    for row_idx in range(start_row, min(sheet.max_row, 5000) + 1):
        title = cell(sheet, headers, row_idx, "product_name")
        if not title:
            continue
        ean = extract_ean(cell(sheet, headers, row_idx, "gtin_code"))
        seller_sku = cell(sheet, headers, row_idx, "seller_sku") or seller_sku_for_ean(ean, DEFAULT_SELLER_SKU_PREFIX)
        if not ean or not seller_sku:
            continue
        image_urls = [cell(sheet, headers, row_idx, "main_image")]
        for image_idx in range(2, 10):
            value = cell(sheet, headers, row_idx, f"image_{image_idx}")
            if value:
                image_urls.append(value)
        rows.append(
            NewProductRow(
                row=row_idx,
                title=title,
                description=cell(sheet, headers, row_idx, "product_description"),
                ean=ean,
                price=cell(sheet, headers, row_idx, "price"),
                quantity=cell(sheet, headers, row_idx, "quantity"),
                seller_sku=seller_sku,
                parcel_weight_g=cell(sheet, headers, row_idx, "parcel_weight"),
                parcel_length_cm=cell(sheet, headers, row_idx, "parcel_length"),
                parcel_width_cm=cell(sheet, headers, row_idx, "parcel_width"),
                parcel_height_cm=cell(sheet, headers, row_idx, "parcel_height"),
                image_urls=list(dict.fromkeys(url for url in image_urls if url)),
            )
        )
        if max_rows and len(rows) >= max_rows:
            break
    workbook.close()
    return rows


def decimal_string(value: str) -> str:
    number = float(clean(value).replace(",", "."))
    return f"{number:.2f}".rstrip("0").rstrip(".")


def int_string(value: str, default: int) -> int:
    text = clean(value).replace(",", ".")
    if not text:
        return default
    return max(0, int(float(text)))


def description_to_html(value: str) -> str:
    lines = [clean(line) for line in str(value or "").splitlines()]
    paragraphs = [line for line in lines if line]
    if not paragraphs:
        return "<p>Produktdetails siehe Artikeldaten.</p>"
    return "".join(f"<p>{html.escape(line)}</p>" for line in paragraphs)


def build_payload(row: NewProductRow, image_uris: list[str], args: argparse.Namespace, warehouse_id: str) -> dict[str, Any]:
    weight_kg = max(0.001, int_string(row.parcel_weight_g, 500) / 1000)
    return {
        "save_mode": args.save_mode,
        "description": description_to_html(row.description),
        "category_id": args.category_id,
        "main_images": [{"uri": uri} for uri in image_uris],
        "skus": [
            {
                "inventory": [{"warehouse_id": warehouse_id, "quantity": int_string(row.quantity, 1)}],
                "seller_sku": row.seller_sku,
                "price": {"amount": decimal_string(row.price), "currency": args.currency},
                "identifier_code": {"code": row.ean, "type": "ISBN"},
            }
        ],
        "title": row.title,
        "is_cod_allowed": False,
        "package_dimensions": {
            "length": str(int_string(row.parcel_length_cm, 25)),
            "width": str(int_string(row.parcel_width_cm, 18)),
            "height": str(int_string(row.parcel_height_cm, 4)),
            "unit": "CENTIMETER",
        },
        "product_attributes": [
            {
                "id": DEFAULT_WARNING_ATTRIBUTE_ID,
                "values": [{"id": DEFAULT_WARNING_NO_VALUE_ID}],
            }
        ],
        "package_weight": {"value": f"{weight_kg:.3f}".rstrip("0").rstrip("."), "unit": "KILOGRAM"},
        "responsible_person_ids": [args.responsible_person_id],
        "manufacturer_ids": [],
        "listing_platforms": ["TIKTOK_SHOP"],
        "shipping_insurance_requirement": "NOT_SUPPORTED",
        "minimum_order_quantity": 1,
        "is_pre_owned": False,
        "category_version": "v2",
        "idempotency_key": f"libri-{row.ean}-rotation-v1",
    }


def download_url(url: str) -> bytes:
    request = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(request, timeout=60) as response:
        return response.read()


def log_row(
    phase: str,
    status: str,
    ean: str = "",
    seller_sku: str = "",
    product_id: str = "",
    sku_id: str = "",
    warehouse_id: str = "",
    title: str = "",
    old_quantity: object = "",
    libri_quantity: object = "",
    new_quantity: object = "",
    message: str = "",
) -> dict[str, Any]:
    return {
        "timestamp_utc": now_utc(),
        "phase": phase,
        "status": status,
        "ean": ean,
        "seller_sku": seller_sku,
        "product_id": product_id,
        "sku_id": sku_id,
        "warehouse_id": warehouse_id,
        "title": title,
        "old_quantity": old_quantity,
        "libri_quantity": "" if libri_quantity is None else libri_quantity,
        "new_quantity": new_quantity,
        "message": message,
    }


def write_log(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=LOG_FIELDS, delimiter=";")
        writer.writeheader()
        writer.writerows(rows)


def write_summary(path: Path, args: argparse.Namespace, rows: list[dict[str, Any]], workbook: Path, log_path: Path) -> None:
    counts: dict[str, int] = {}
    for row in rows:
        key = f"{row['phase']}:{row['status']}"
        counts[key] = counts.get(key, 0) + 1
    lines = [
        "# TikTok Catalog Rotation",
        "",
        f"Generated: {dt.datetime.now().isoformat(timespec='seconds')}",
        f"Mode: {'live' if args.live else 'dry-run'}",
        f"New workbook: {workbook}",
        f"Log: {log_path}",
        "",
        "## Counts",
        "",
    ]
    for key, count in sorted(counts.items()):
        lines.append(f"- {key}: {count}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def read_retire_csv(path: Path, seller_sku_prefix: str) -> list[str]:
    text = path.read_text(encoding="utf-8-sig", errors="replace")
    rows = csv.DictReader(text.splitlines(), dialect=sniff_csv(text))
    values: list[str] = []
    for row in rows:
        seller_sku = clean(row.get("seller_sku") or row.get("Seller SKU") or row.get("sku") or row.get("SKU"))
        ean = extract_ean(row.get("ean") or row.get("isbn") or row.get("gtin_code") or seller_sku, seller_sku_prefix)
        if not seller_sku and ean:
            seller_sku = seller_sku_for_ean(ean, seller_sku_prefix)
        if seller_sku:
            values.append(seller_sku)
    return dedupe(values)


def explicit_retire_skus(args: argparse.Namespace) -> list[str]:
    values = list(args.retire_sku)
    values.extend(seller_sku_for_ean(extract_ean(ean, args.seller_sku_prefix), args.seller_sku_prefix) for ean in args.retire_ean)
    for csv_path in args.retire_csv:
        values.extend(read_retire_csv(Path(csv_path), args.seller_sku_prefix))
    return dedupe(values)


def find_skus_by_seller_sku(
    client: CatalogRotationClient,
    seller_skus: list[str],
    args: argparse.Namespace,
    warehouse_id: str,
) -> list[TikTokSku]:
    found: dict[str, TikTokSku] = {}
    batch_size = min(max(args.search_batch_size, 1), 10)
    for start in range(0, len(seller_skus), batch_size):
        batch = seller_skus[start : start + batch_size]
        products = client.search_products(seller_skus=batch, page_size=max(args.page_size, len(batch)))
        for product in products:
            for sku in product_skus(product, args.seller_sku_prefix, warehouse_id):
                found.setdefault(sku.seller_sku, sku)
        if args.tiktok_delay:
            time.sleep(args.tiktok_delay)
    return [found[key] for key in seller_skus if key in found]


def existing_seller_skus(
    client: CatalogRotationClient,
    rows: list[NewProductRow],
    args: argparse.Namespace,
    warehouse_id: str,
) -> set[str]:
    seller_skus = [row.seller_sku for row in rows]
    return {sku.seller_sku for sku in find_skus_by_seller_sku(client, seller_skus, args, warehouse_id)}


def discover_current_skus(
    client: CatalogRotationClient,
    args: argparse.Namespace,
    warehouse_id: str,
) -> list[TikTokSku]:
    products = client.search_products(page_size=args.page_size, max_pages=args.max_scan_pages)
    allowed_statuses = {status.strip() for status in args.retire_product_statuses.split(",") if status.strip()}
    skus: list[TikTokSku] = []
    for product in products:
        skus.extend(
            sku
            for sku in product_skus(product, args.seller_sku_prefix, warehouse_id)
            if (not allowed_statuses or sku.product_status in allowed_statuses)
            and sku.current_quantity != 0
        )
        if args.max_scan_skus and len(skus) >= args.max_scan_skus:
            break
    return skus[: args.max_scan_skus if args.max_scan_skus else None]


def auto_retire_candidates(
    client: CatalogRotationClient,
    args: argparse.Namespace,
    warehouse_id: str,
    exclude_eans: set[str],
    run_dir: Path,
) -> list[RetireCandidate]:
    skus = discover_current_skus(client, args, warehouse_id)
    skus = [sku for sku in skus if sku.ean not in exclude_eans]
    candidates: list[RetireCandidate] = []
    opener = login(Path(args.env)) if args.refresh_libri else None
    libri_page_dir = run_dir / "retire_libri_pages"

    for sku in skus:
        try:
            stock = (
                fetch_libri_stock(sku.ean, opener, libri_page_dir)
                if args.refresh_libri
                else read_local_libri_stock(sku.ean, args.local_libri_dir)
            )
            if stock.quantity is not None and stock.quantity <= args.retire_stock_threshold:
                candidates.append(
                    RetireCandidate(
                        sku=sku,
                        libri_quantity=stock.quantity,
                        reason=f"libri_stock_lte_{args.retire_stock_threshold}",
                        libri_page=str(stock.page_path),
                    )
                )
        except Exception as exc:
            if args.stop_on_error:
                raise
            print(f"Skipping retire candidate {sku.seller_sku}: Libri stock check failed: {str(exc)[:200]}")
        if args.libri_delay:
            time.sleep(args.libri_delay)

    candidates.sort(key=lambda item: (item.libri_quantity is None, item.libri_quantity or 999999, item.sku.seller_sku))
    return candidates


def collect_retire_candidates(
    client: CatalogRotationClient,
    args: argparse.Namespace,
    warehouse_id: str,
    exclude_eans: set[str],
    run_dir: Path,
) -> list[RetireCandidate]:
    explicit = explicit_retire_skus(args)
    if explicit:
        return [RetireCandidate(sku=sku, reason="explicit") for sku in find_skus_by_seller_sku(client, explicit, args, warehouse_id)]
    if args.auto_retire_low_libri_stock:
        return auto_retire_candidates(client, args, warehouse_id, exclude_eans, run_dir)
    return []


def retire_zero_inventory(
    client: CatalogRotationClient,
    candidate: RetireCandidate,
    args: argparse.Namespace,
    log_rows: list[dict[str, Any]],
    log_path: Path,
) -> bool:
    sku = candidate.sku
    row = log_row(
        phase="retire",
        status="planned_zero_inventory" if not args.live else "retiring_zero_inventory",
        ean=sku.ean,
        seller_sku=sku.seller_sku,
        product_id=sku.product_id,
        sku_id=sku.sku_id,
        warehouse_id=sku.warehouse_id,
        title=sku.title,
        old_quantity="" if sku.current_quantity is None else sku.current_quantity,
        libri_quantity=candidate.libri_quantity,
        new_quantity=0,
        message=candidate.reason,
    )
    if not args.live:
        log_rows.append(row)
        write_log(log_path, log_rows)
        return True
    try:
        response = client.update_inventory(sku.product_id, sku.sku_id, sku.warehouse_id, 0)
        errors = response_errors(response)
        row["status"] = "retired_zero_inventory_with_response_errors" if errors else "retired_zero_inventory"
        row["message"] = errors[:1500] if errors else row["message"]
    except Exception as exc:
        row["status"] = "failed"
        row["message"] = str(exc)[:1500]
        log_rows.append(row)
        write_log(log_path, log_rows)
        if args.stop_on_error:
            raise
        return False
    log_rows.append(row)
    write_log(log_path, log_rows)
    return row["status"] == "retired_zero_inventory"


def create_product(
    client: CatalogRotationClient,
    product: NewProductRow,
    args: argparse.Namespace,
    warehouse_id: str,
    output_dir: Path,
    log_rows: list[dict[str, Any]],
    log_path: Path,
    image_cache: dict[str, str],
) -> bool:
    row = log_row(
        phase="create",
        status="planned_create" if not args.live else "creating",
        ean=product.ean,
        seller_sku=product.seller_sku,
        warehouse_id=warehouse_id,
        title=product.title,
        new_quantity=int_string(product.quantity, 1),
        message=f"workbook_row={product.row}",
    )
    try:
        image_uris: list[str] = []
        for image_index, image_url in enumerate(product.image_urls[: args.max_images], start=1):
            if image_url in image_cache:
                image_uris.append(image_cache[image_url])
                continue
            if not args.live:
                image_uris.append(f"dry-run-image-{image_index}")
                continue
            image_bytes = download_url(image_url)
            uploaded = client.upload_product_image(image_bytes, f"{product.ean}_{image_index}.jpg")
            uri = clean(uploaded.get("uri"))
            if not uri:
                raise TikTokApiError(f"Image upload returned no uri: {uploaded}")
            image_cache[image_url] = uri
            image_uris.append(uri)

        payload = build_payload(product, image_uris, args, warehouse_id)
        payload_dir = output_dir / "payloads"
        payload_dir.mkdir(parents=True, exist_ok=True)
        (payload_dir / f"{product.ean}.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

        if args.live:
            response = client.create_product(payload)
            data = response.get("data") or {}
            skus = data.get("skus") or []
            row.update(
                status="created",
                product_id=clean(data.get("product_id")),
                sku_id=clean((skus[0] or {}).get("id")) if skus else "",
                message=compact_json(
                    {
                        "request_id": response.get("request_id"),
                        "warnings": data.get("warnings") or [],
                    }
                ),
            )
        else:
            row["message"] = "Payload written, product not created."
    except Exception as exc:
        row["status"] = "failed"
        row["message"] = str(exc)[:1500]
        log_rows.append(row)
        write_log(log_path, log_rows)
        if args.stop_on_error:
            raise
        return False
    log_rows.append(row)
    write_log(log_path, log_rows)
    return row["status"] in {"created", "planned_create"}


def pair_rotation(new_rows: list[NewProductRow], retire_candidates: list[RetireCandidate], args: argparse.Namespace) -> tuple[list[NewProductRow], list[RetireCandidate]]:
    if args.replace_count:
        new_rows = new_rows[: args.replace_count]
    if args.allow_create_without_retire:
        retire_limit = min(len(retire_candidates), len(new_rows))
        return new_rows, retire_candidates[:retire_limit]
    pair_count = min(len(new_rows), len(retire_candidates))
    return new_rows[:pair_count], retire_candidates[:pair_count]


def run(args: argparse.Namespace) -> int:
    env_path = Path(args.env)
    env = load_env_file(env_path)
    warehouse_id = effective_warehouse_id(args, env)
    client = CatalogRotationClient(env, env_path)
    workbook = Path(args.new_workbook).resolve() if args.new_workbook else discover_latest_workbook().resolve()
    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    run_dir = Path(args.output_root) / stamp
    log_path = run_dir / "catalog_rotation_log.csv"
    summary_path = run_dir / "summary.md"
    log_rows: list[dict[str, Any]] = []

    all_new_rows = load_new_rows(workbook, args.start_row, args.new_scan_limit)
    existing = existing_seller_skus(client, all_new_rows, args, warehouse_id) if all_new_rows else set()
    new_rows = [row for row in all_new_rows if row.seller_sku not in existing]
    for row in all_new_rows:
        if row.seller_sku in existing:
            log_rows.append(
                log_row(
                    phase="create",
                    status="skipped_existing_sku",
                    ean=row.ean,
                    seller_sku=row.seller_sku,
                    warehouse_id=warehouse_id,
                    title=row.title,
                    message="New candidate already exists in TikTok.",
                )
            )

    exclude_eans = {row.ean for row in all_new_rows}
    retire_candidates = collect_retire_candidates(client, args, warehouse_id, exclude_eans, run_dir)
    new_rows, retire_candidates = pair_rotation(new_rows, retire_candidates, args)

    if not new_rows:
        log_rows.append(log_row(phase="plan", status="nothing_to_create", message="No non-existing new product rows selected."))
    if not retire_candidates and not args.allow_create_without_retire:
        log_rows.append(log_row(phase="plan", status="nothing_to_retire", message="No retire candidates selected."))

    write_log(log_path, log_rows)
    image_cache: dict[str, str] = {}
    created_successes = 0
    retired_successes = 0

    if args.retire_before_create:
        for candidate in retire_candidates:
            if retire_zero_inventory(client, candidate, args, log_rows, log_path):
                retired_successes += 1
        create_target_count = len(retire_candidates) if not args.allow_create_without_retire else len(new_rows)
        for product in new_rows[:create_target_count]:
            if create_product(client, product, args, warehouse_id, run_dir, log_rows, log_path, image_cache):
                created_successes += 1
    else:
        for product in new_rows:
            if create_product(client, product, args, warehouse_id, run_dir, log_rows, log_path, image_cache):
                created_successes += 1
        retire_target_count = len(retire_candidates) if not args.live else min(len(retire_candidates), created_successes)
        for candidate in retire_candidates[:retire_target_count]:
            if retire_zero_inventory(client, candidate, args, log_rows, log_path):
                retired_successes += 1

    write_summary(summary_path, args, log_rows, workbook, log_path)
    failed = sum(1 for row in log_rows if row["status"] == "failed")
    print(f"Mode: {'live' if args.live else 'dry-run'}")
    print(f"New workbook: {workbook}")
    print(f"Selected new products: {len(new_rows)}")
    print(f"Selected retire candidates: {len(retire_candidates)}")
    print(f"Created/planned: {created_successes}")
    print(f"Retired/planned: {retired_successes}")
    print(f"Failed: {failed}")
    print(f"Log: {log_path.resolve()}")
    return 1 if failed else 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Replace older TikTok Libri titles with prepared new titles.")
    parser.add_argument("--env", default=".env")
    parser.add_argument("--output-root", default=str(DEFAULT_OUTPUT_ROOT))
    parser.add_argument("--new-workbook", default="", help="Prepared TikTok tiktok_upload_green.xlsx. Defaults to latest under outputs/.")
    parser.add_argument("--start-row", type=int, default=DATA_START_ROW)
    parser.add_argument("--new-scan-limit", type=int, default=200, help="Max workbook rows to inspect.")
    parser.add_argument("--replace-count", type=int, default=10, help="Max titles to rotate.")
    parser.add_argument("--live", action="store_true", help="Perform TikTok writes. Without this flag only an audit plan is written.")
    parser.add_argument("--save-mode", choices=["AS_DRAFT", "LISTING"], default="AS_DRAFT")
    parser.add_argument("--allow-create-without-retire", action="store_true")
    parser.add_argument("--retire-before-create", action="store_true", help="Set old selected SKUs to 0 before creating new listings.")
    parser.add_argument("--retire-sku", action="append", default=[], help="Explicit old Seller SKU to retire. Can be repeated.")
    parser.add_argument("--retire-ean", action="append", default=[], help="Explicit old ISBN/EAN to retire. Can be repeated.")
    parser.add_argument("--retire-csv", action="append", default=[], help="CSV containing old seller_sku/ean values to retire.")
    parser.add_argument("--auto-retire-low-libri-stock", action=argparse.BooleanOptionalAction, default=True)
    parser.add_argument("--retire-stock-threshold", type=int, default=0, help="Auto-retire when current Libri stock is <= this value.")
    parser.add_argument("--retire-product-statuses", default="ACTIVATE", help="Comma-separated TikTok product statuses eligible for automatic retirement.")
    parser.add_argument("--refresh-libri", action=argparse.BooleanOptionalAction, default=True)
    parser.add_argument("--local-libri-dir", action="append", default=["libri_product_pages", "libri_bulk_pages"])
    parser.add_argument("--seller-sku-prefix", default=DEFAULT_SELLER_SKU_PREFIX)
    parser.add_argument("--warehouse-id", default="", help="TikTok warehouse ID. Defaults to TIKTOK_WAREHOUSE_ID or project default.")
    parser.add_argument("--page-size", type=int, default=50)
    parser.add_argument("--max-scan-pages", type=int, default=4, help="Max TikTok product-search pages to scan for auto-retire.")
    parser.add_argument("--max-scan-skus", type=int, default=200)
    parser.add_argument("--search-batch-size", type=int, default=10)
    parser.add_argument("--max-images", type=int, default=5)
    parser.add_argument("--category-id", default=DEFAULT_CATEGORY_ID)
    parser.add_argument("--responsible-person-id", default=DEFAULT_RESPONSIBLE_PERSON_ID)
    parser.add_argument("--currency", default=DEFAULT_CURRENCY)
    parser.add_argument("--libri-delay", type=float, default=0.3)
    parser.add_argument("--tiktok-delay", type=float, default=0.2)
    parser.add_argument("--stop-on-error", action=argparse.BooleanOptionalAction, default=False)
    return parser


def main(argv: list[str] | None = None) -> int:
    return run(build_parser().parse_args(argv))


if __name__ == "__main__":
    raise SystemExit(main())
