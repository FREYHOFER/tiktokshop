#!/usr/bin/env python3
"""Patch a TikTok bulk-edit workbook from a listing-quality update plan.

Download TikTok's bulk-edit template for active products first, then run this
script. It preserves all template structure and only edits matching product rows.
Local image files cannot be written as upload URLs; for image fixes the script
keeps the Libri front-cover URL as primary unless the plan contains a hosted URL
column filled after uploading assets to TikTok Media Center or another approved
host.
"""

from __future__ import annotations

import argparse
import csv
import re
from pathlib import Path

import openpyxl


PREFERRED_HOSTED_IMAGE_COLUMNS = ["hosted_main_image_url", "media_center_url", "uploaded_main_image_url"]


def extract_ean_from_text(value: str) -> str:
    for match in re.findall(r"(?:97[89])[\d\-\s]{10,20}", value or ""):
        digits = re.sub(r"\D", "", match)
        if len(digits) == 13:
            return digits
    return ""


def load_plan(path: Path) -> dict[str, dict[str, str]]:
    rows_by_key: dict[str, dict[str, str]] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        for row in csv.DictReader(fh):
            for key in [row.get("product_id", ""), row.get("seller_sku", ""), row.get("ean", "")]:
                if key:
                    rows_by_key[key] = row
    return rows_by_key


def detect_header(sheet) -> tuple[int, dict[str, int]]:
    for row_idx in range(1, 11):
        headers = {
            str(sheet.cell(row=row_idx, column=col).value): col
            for col in range(1, sheet.max_column + 1)
            if sheet.cell(row=row_idx, column=col).value
        }
        if {"product_name", "main_image"} & set(headers) and {"seller_sku", "gtin_code"} & set(headers):
            return row_idx, headers
    raise ValueError("Could not find TikTok machine-key header row in the workbook.")


def first_hosted_image(plan_row: dict[str, str]) -> str:
    for column in PREFERRED_HOSTED_IMAGE_COLUMNS:
        value = plan_row.get(column, "").strip()
        if value:
            return value
    return plan_row.get("selected_cover_url", "").strip()


def patch_workbook(args: argparse.Namespace) -> int:
    plan = load_plan(Path(args.plan))
    workbook = openpyxl.load_workbook(args.template)
    sheet = workbook[args.sheet] if args.sheet else workbook.active
    header_row, headers = detect_header(sheet)
    data_start_row = 7 if header_row == 1 else header_row + 1

    patched = []
    for row_idx in range(data_start_row, sheet.max_row + 1):
        seller_sku_col = headers.get("seller_sku")
        gtin_col = headers.get("gtin_code")
        product_id_col = headers.get("product_id")
        product_name_col = headers.get("product_name")
        description_col = headers.get("product_description")
        product_id = str(sheet.cell(row_idx, product_id_col).value or "") if product_id_col else ""
        seller_sku = str(sheet.cell(row_idx, seller_sku_col).value or "") if seller_sku_col else ""
        ean = str(sheet.cell(row_idx, gtin_col).value or "") if gtin_col else ""
        if not ean:
            product_name = str(sheet.cell(row_idx, product_name_col).value or "") if product_name_col else ""
            description = str(sheet.cell(row_idx, description_col).value or "") if description_col else ""
            ean = extract_ean_from_text(f"{product_name} {description}")
        plan_row = plan.get(product_id) or plan.get(seller_sku) or plan.get(ean)
        if not plan_row:
            continue

        changes = []
        if product_name_col and args.patch_titles and plan_row.get("title_action") == "update_title_40_plus":
            new_name = plan_row.get("suggested_product_name", "").strip()
            if new_name:
                sheet.cell(row_idx, product_name_col).value = new_name
                changes.append("product_name")

        main_image_col = headers.get("main_image")
        if main_image_col and args.patch_main_image:
            image_value = first_hosted_image(plan_row)
            if image_value:
                sheet.cell(row_idx, main_image_col).value = image_value
                changes.append("main_image")

        if changes:
            patched.append({"row": row_idx, "seller_sku": seller_sku, "ean": ean, "changes": "|".join(changes)})

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    log_path = output_path.with_suffix(".patch_log.csv")
    with log_path.open("w", encoding="utf-8-sig", newline="") as fh:
        fieldnames = ["row", "seller_sku", "ean", "changes"]
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(patched)

    print(f"Patched rows: {len(patched)}")
    print(f"Output: {output_path.resolve()}")
    print(f"Log: {log_path.resolve()}")
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Patch TikTok bulk-edit template from quality update plan.")
    parser.add_argument("--template", required=True, help="Downloaded TikTok bulk-edit XLSX for existing products.")
    parser.add_argument("--plan", default="outputs/current_listing_quality_updates/listing_quality_update_plan.csv")
    parser.add_argument("--output", default="outputs/current_listing_quality_updates/tiktok_bulk_edit_patched.xlsx")
    parser.add_argument("--sheet", default="", help="Sheet name. Defaults to workbook active sheet.")
    parser.add_argument("--patch-titles", action=argparse.BooleanOptionalAction, default=True)
    parser.add_argument("--patch-main-image", action=argparse.BooleanOptionalAction, default=True)
    return parser


def main(argv: list[str] | None = None) -> int:
    return patch_workbook(build_parser().parse_args(argv))


if __name__ == "__main__":
    raise SystemExit(main())
