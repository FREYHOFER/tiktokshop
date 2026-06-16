#!/usr/bin/env python3
"""Build TikTok upload pack for the Sprache und Wörterbücher bestseller list."""

from __future__ import annotations

import argparse
import csv
import html
import re
import sys
import unicodedata
import urllib.parse
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from build_tiktok_quality_update_pack import (  # noqa: E402
    create_white_cover_image_from_image,
    download_image,
    extract_reference_cover_image,
    image_is_single_cover_like,
    is_cover_like,
    librinet_media_role,
)
from collect_libri_candidates import Source, parse_article_blocks  # noqa: E402
from fetch_libri_product_pages import PRODUCT_URL, fetch, login  # noqa: E402
from tiktok_libri_pipeline import (  # noqa: E402
    BRAND_NO_BRAND,
    DESCRIPTION_REPLACEMENTS,
    GTIN_TYPE,
    MIN_LISTING_QUALITY_TITLE_CHARS,
    ProductRecord,
    build_description,
    fetch_image_metadata,
    image_quality_score,
    normalize_text,
    parcel_height,
    parcel_weight,
    parse_int,
    parse_libri_detail_html,
    parse_price,
)
from upload_tiktok_images_cloudinary import load_env as load_cloudinary_env, upload_image  # noqa: E402


DATA_START_ROW = 7
CATEGORY = "Bildung und Schulbildung/Sprache und Wörterbücher"
MIN_MATCH_SCORE = 0.72
DEFAULT_RP_ID = "69e8ebf9401201b6c3954a0f"
DEFAULT_PACKAGING_WARNING = "Nein"

STOPWORDS = {
    "der",
    "die",
    "das",
    "den",
    "dem",
    "des",
    "ein",
    "eine",
    "einen",
    "einem",
    "einer",
    "und",
    "oder",
    "mit",
    "ohne",
    "für",
    "fuer",
    "von",
    "vom",
    "zum",
    "zur",
    "als",
    "in",
    "im",
    "am",
    "an",
    "auf",
    "aus",
    "zu",
}

LANGUAGE_TERMS = [
    "deutsch",
    "englisch",
    "persisch",
    "türkisch",
    "tuerkisch",
    "arabisch",
    "italienisch",
    "französisch",
    "franzoesisch",
    "spanisch",
    "latein",
    "sprache",
    "sprach",
    "wörterbuch",
    "woerterbuch",
    "dictionary",
    "grammatik",
    "rechtschreiben",
    "aufsatz",
    "lesen",
    "schreiben",
    "erstlesen",
    "wörter",
    "woerter",
    "wortschatz",
    "verb",
    "daf",
    "zuwanderer",
    "telc",
    "zertifikat",
    "kursbuch",
    "arbeitsbuch",
    "übungsgrammatik",
    "uebungsgrammatik",
    "prüfungstraining",
    "pruefungstraining",
    "sprachführer",
    "sprachfuehrer",
    "silbenfibel",
]

WRONG_CATEGORY_TERMS = [
    "mathe",
    "rechnen",
    "pflege",
    "pflegen",
    "pflegeexamen",
    "orakel",
    "tarot",
    "lenormand",
    "rauhnacht",
    "zaubersprüche",
    "zaubersprueche",
    "bibel",
    "roman",
    "soundbuch",
    "hör mal",
    "hoer mal",
    "babybücher",
    "babybuecher",
    "selbstvertrauen",
    "freunde gewinnt",
    "rich dad",
    "mountain is you",
    "secret",
    "alchimist",
    "toxische menschen",
    "48 laws",
    "übernatürlich",
    "uebernatuerlich",
    "traumdeutung",
    "hand",
    "ratgeber",
    "lebenshilfe",
    "psychologie",
    "wirtschaftsratgeber",
    "religi",
    "bilderbücher",
    "bilderbuecher",
]

REVIEW_TERMS = [
    "fick",
    "böse nachrichten",
    "boese nachrichten",
    "dunkle psychologie",
    "manipulation",
]

HARD_EXCLUDE_TERMS = [
    "pflege",
    "pflegeausbildung",
    "pflegeexamen",
    "psychologie",
    "wirtschaftsratgeber",
    "lebenshilfe",
    "bibel",
    "orakel",
    "tarot",
    "lenormand",
    "rauhnacht",
    "zaubersprueche",
]


@dataclass
class RankedItem:
    rank: int
    tiktok_title: str
    tiktok_price: float | None
    tiktok_sales: int | None


@dataclass
class CatalogItem:
    ean: str
    title: str
    author: str = ""
    source_price: float | None = None
    source: str = ""
    detail_path: Path | None = None
    product: ProductRecord | None = None


def ascii_fold(value: str) -> str:
    value = value.replace("ß", "ss").replace("ẞ", "SS")
    value = value.replace("ä", "ae").replace("ö", "oe").replace("ü", "ue")
    value = value.replace("Ä", "Ae").replace("Ö", "Oe").replace("Ü", "Ue")
    value = unicodedata.normalize("NFKD", value)
    return "".join(ch for ch in value if not unicodedata.combining(ch))


def norm(value: str) -> str:
    value = html.unescape(value or "")
    value = ascii_fold(value)
    value = value.lower().replace("...", " ")
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def tokens(value: str) -> list[str]:
    return [tok for tok in norm(value).split() if len(tok) > 2 and tok not in STOPWORDS]


def parse_ranked_list(path: Path) -> list[RankedItem]:
    items: list[RankedItem] = []
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.lower().startswith("rang "):
            continue
        parts = [part.strip() for part in line.split("|")]
        if len(parts) < 4 or not parts[0].isdigit():
            continue
        price_text = parts[2].replace("€", "").strip()
        sales_text = parts[3].strip()
        items.append(
            RankedItem(
                rank=int(parts[0]),
                tiktok_title=parts[1],
                tiktok_price=parse_price(price_text),
                tiktok_sales=parse_int(sales_text),
            )
        )
    return items


def read_candidate_csvs(paths: list[Path]) -> dict[str, CatalogItem]:
    catalog: dict[str, CatalogItem] = {}
    for path in paths:
        if not path.exists():
            continue
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            for row in csv.DictReader(fh):
                ean = re.sub(r"\D", "", row.get("ean", ""))
                if len(ean) != 13:
                    continue
                title = normalize_text(row.get("title", ""))
                if not title:
                    continue
                current = catalog.get(ean)
                item = CatalogItem(
                    ean=ean,
                    title=title,
                    author=normalize_text(row.get("author", "")),
                    source_price=parse_price(row.get("source_price_text", "")),
                    source=str(path),
                )
                if current is None or len(item.title) > len(current.title):
                    catalog[ean] = item
    return catalog


def add_detail_pages(catalog: dict[str, CatalogItem], detail_dir: Path) -> None:
    for path in detail_dir.glob("*.html"):
        ean = path.stem
        if len(ean) != 13:
            continue
        try:
            product = parse_libri_detail_html(path)
        except Exception:  # noqa: BLE001 - skip malformed cached page
            continue
        if not product.title:
            continue
        product.ean = product.ean or ean
        current = catalog.get(ean)
        item = CatalogItem(
            ean=ean,
            title=product.title,
            author=product.author,
            source_price=product.price,
            source=str(path),
            detail_path=path,
            product=product,
        )
        if current is None or current.product is None:
            catalog[ean] = item


def candidate_text(item: CatalogItem) -> str:
    product = item.product
    pieces = [item.title, item.author]
    if product:
        pieces.extend([product.subtitle, product.publisher, product.product_group, product.binding])
    return " ".join(piece for piece in pieces if piece)


def match_score(ranked: RankedItem, item: CatalogItem) -> float:
    query = norm(ranked.tiktok_title)
    candidate = norm(candidate_text(item))
    if not query or not candidate:
        return 0.0
    query_tokens = tokens(ranked.tiktok_title)
    candidate_tokens = set(tokens(candidate_text(item)))
    overlap = len([tok for tok in query_tokens if tok in candidate_tokens]) / max(1, len(query_tokens))
    seq = SequenceMatcher(None, query, candidate[: max(len(query) + 30, 80)]).ratio()
    prefix = 1.0 if candidate.startswith(query[: min(len(query), 28)]) else 0.0
    score = (overlap * 0.58) + (seq * 0.32) + (prefix * 0.10)
    price = item.source_price or (item.product.price if item.product else None)
    if ranked.tiktok_price is not None and price is not None:
        diff = abs(ranked.tiktok_price - price)
        if diff <= 0.11:
            score += 0.08
        elif diff <= 1.0:
            score += 0.04
        elif diff > 5.0:
            score -= 0.05
    return min(score, 1.0)


def clean_query(title: str) -> str:
    cleaned = title.split("...")[0].strip()
    words = cleaned.split()
    return " ".join(words[:10])


def search_libri(opener, query: str, page_size: int = 8) -> list[str]:
    params = {"searchInitiated": "1", "query": query, "ps": str(page_size)}
    url = "https://mein.libri.de/Bestellen/Suchen.html?" + urllib.parse.urlencode(params)
    _, body = fetch(opener, url)
    decoded = html.unescape(body)
    rows = parse_article_blocks(decoded, Source("tiktok_language_query", "Search", "", 1, 1), 1, page_size)
    eans = [str(row.get("ean", "")) for row in rows if re.fullmatch(r"\d{10,13}", str(row.get("ean", "")))]
    fallback = re.findall(r"/produkt/(\d{10,13})/", decoded)
    return list(dict.fromkeys(eans + fallback))[:page_size]


def fetch_detail(opener, ean: str, detail_dir: Path) -> Path | None:
    path = detail_dir / f"{ean}.html"
    if path.exists() and path.stat().st_size > 0:
        return path
    final_url, body = fetch(opener, PRODUCT_URL.format(ean=ean))
    if "Login.html" in final_url or "<title>Mein.Libri - Login</title>" in body:
        return None
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(body, encoding="utf-8")
    return path


def best_match(ranked: RankedItem, catalog: dict[str, CatalogItem]) -> tuple[CatalogItem | None, float]:
    best_item = None
    best = 0.0
    for item in catalog.values():
        score = match_score(ranked, item)
        if score > best:
            best = score
            best_item = item
    return best_item, best


def ensure_product(item: CatalogItem, detail_dir: Path) -> ProductRecord | None:
    if item.product:
        return item.product
    path = item.detail_path or (detail_dir / f"{item.ean}.html")
    if not path.exists():
        return None
    product = parse_libri_detail_html(path)
    product.ean = product.ean or item.ean
    item.product = product
    item.detail_path = path
    return product


def category_fit(product: ProductRecord, ranked_title: str) -> tuple[bool, str]:
    text = norm(" ".join([ranked_title, product.title, product.subtitle, product.product_group, product.blurb, product.binding]))
    title_text = norm(" ".join([ranked_title, product.title, product.subtitle]))
    group = norm(product.product_group)
    if any(term in text for term in HARD_EXCLUDE_TERMS):
        return False, "hard_exclude_category"
    strong_terms = [
        "woerterbuch",
        "dictionary",
        "grammatik",
        "uebungsgrammatik",
        "rechtschreiben",
        "aufsatz",
        "erstlesen",
        "silbenfibel",
        "wortschatz",
        "verbtabellen",
        "zuwanderer",
        "telc",
        "zertifikat",
        "pruefungstraining",
        "pruefung",
        "daf",
        "sprachfuehrer",
        "sprachkurs",
        "kursbuch",
        "arbeitsbuch",
    ]
    weak_deutsch_terms = ["deutsch", "englisch", "persisch", "tuerkisch", "arabisch", "franzoesisch", "spanisch", "latein"]
    language_signal = any(term in title_text for term in strong_terms)
    if not language_signal and any(term in title_text for term in weak_deutsch_terms):
        language_signal = any(term in title_text for term in ["lernen", "lesen", "schreiben", "test", "klasse", "schule"])
    school_language_signal = (
        any(term in group for term in ["lernhilfen", "schulbuch", "schule"])
        and any(term in title_text for term in ["deutsch", "lesen", "schreiben", "rechtschreiben", "aufsatz", "erstlesen", "silbenfibel"])
    )
    if any(term in text for term in map(norm, WRONG_CATEGORY_TERMS)):
        if not (language_signal or school_language_signal):
            return False, "wrong_category_terms"
        if any(term in text for term in ["soundbuch", "babybuecher", "bilderbuecher"]) and "woerterbuch" not in title_text:
            return False, "wrong_category_terms"
    if language_signal or school_language_signal:
        return True, "language_terms"
    return False, "no_language_dictionary_signal"


def compliance_status(product: ProductRecord, ranked_title: str) -> tuple[str, str]:
    text = norm(" ".join([ranked_title, product.title, product.subtitle, product.blurb, product.author_bio]))
    hits = [term for term in REVIEW_TERMS if norm(term) in text]
    if hits:
        return "review", "review_terms:" + "|".join(hits)
    return "ok", ""


def language_mismatch(product: ProductRecord, ranked_title: str) -> str:
    languages = {
        "deutsch": ["deutsch", "german"],
        "englisch": ["englisch", "english"],
        "italienisch": ["italienisch", "italian"],
        "franzoesisch": ["franzoesisch", "french"],
        "spanisch": ["spanisch", "spanish"],
        "persisch": ["persisch", "farsi"],
        "tuerkisch": ["tuerkisch", "turkish"],
        "arabisch": ["arabisch", "arabic"],
        "latein": ["latein", "latin"],
    }
    requested = {
        key
        for key, variants in languages.items()
        if any(variant in norm(ranked_title) for variant in variants)
    }
    if not requested:
        return ""
    product_text = norm(" ".join([product.title, product.subtitle, product.product_group]))
    present = {
        key
        for key, variants in languages.items()
        if any(variant in product_text for variant in variants)
    }
    if present and requested.isdisjoint(present):
        return f"language_mismatch:requested_{'|'.join(sorted(requested))}_matched_{'|'.join(sorted(present))}"
    return ""


def build_product_name(product: ProductRecord) -> str:
    name = product.title
    descriptor = ""
    text = norm(" ".join([product.title, product.subtitle, product.product_group]))
    if "woerterbuch" in text or "dictionary" in text:
        descriptor = "Wörterbuch"
    elif "grammatik" in text:
        descriptor = "Grammatik Übungsbuch"
    elif "zuwanderer" in text or "telc" in text or "zertifikat" in text:
        descriptor = "Deutsch Prüfungstraining"
    elif "kursbuch" in text or "arbeitsbuch" in text:
        descriptor = "Deutsch Kursbuch Arbeitsbuch"
    elif "lesen" in text or "rechtschreiben" in text or "aufsatz" in text:
        descriptor = "Deutsch Lernhilfe"
    else:
        descriptor = "Sprache und Wörterbücher"
    author = normalize_text(product.author)
    if len(author) > 80 or re.search(r"[.!?]", author):
        author = ""
    if author and norm(author) not in norm(name):
        name = f"{name} - {author}"
    if descriptor and norm(descriptor) not in norm(name):
        name = f"{name} - {descriptor}"
    if product.binding and len(name) < MIN_LISTING_QUALITY_TITLE_CHARS and norm(product.binding) not in norm(name):
        name = f"{name} - {product.binding}"
    return normalize_text(name)[:254]


def sanitize_description(product: ProductRecord) -> str:
    description = build_description(product)
    for pattern, replacement in DESCRIPTION_REPLACEMENTS:
        description = re.sub(pattern, replacement, description, flags=re.I)
    return description[:9900]


def select_cover_image(product: ProductRecord, detail_path: Path) -> tuple[str, object | None, str]:
    reference = extract_reference_cover_image(detail_path)
    roles = [librinet_media_role(url) for url in product.image_urls]
    reference_safe = bool(reference) and not (2 in roles and 1 not in roles)
    best_url = ""
    best_score = -1
    best_image = None
    for url in product.image_urls:
        if librinet_media_role(url) == 2:
            continue
        check = fetch_image_metadata(url)
        if not check.ok or not check.tiktok_size_ok:
            continue
        try:
            image = download_image(url)
        except Exception:  # noqa: BLE001
            continue
        if not (is_cover_like(check) or image_is_single_cover_like(image)):
            continue
        score = image_quality_score(check)[2]
        if librinet_media_role(url) == 1:
            score += 5_000_000
        if score > best_score:
            best_url = url
            best_score = score
            best_image = image
    if best_image is not None:
        return best_url, best_image, "public_front_cover"
    if reference_safe and reference is not None:
        return "", reference, "embedded_reference_cover"
    return "", None, "no_safe_cover"


def upload_cover(image_path: Path, product: ProductRecord, folder: str, cloud_name: str, api_key: str, api_secret: str) -> str:
    row = {
        "ean": product.ean,
        "seller_sku": f"LIBRI-{product.ean}",
        "local_image_path": str(image_path),
    }
    return upload_image(row, folder, cloud_name, api_key, api_secret).get("secure_url", "")


def clear_template(sheet) -> None:
    for row in range(DATA_START_ROW, min(sheet.max_row, 5000) + 1):
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row=row, column=col).value = None


def template_path_from_arg(value: str) -> Path:
    if value:
        path = Path(value)
        if path.exists():
            return path
    matches = [
        p
        for p in Path(r"C:\Users\User\Downloads").glob("Tiktoksellercenter_*.xlsx")
        if "20260516" in p.name and "Sprache" in p.name
    ]
    if not matches:
        raise FileNotFoundError("Sprache und Wörterbücher template not found in Downloads.")
    return matches[0]


def write_workbook(template_path: Path, output_path: Path, report_rows: list[dict[str, str]]) -> None:
    workbook = openpyxl.load_workbook(template_path)
    sheet = workbook["Template"]
    headers = {
        str(sheet.cell(row=1, column=col).value): col
        for col in range(1, sheet.max_column + 1)
        if sheet.cell(row=1, column=col).value
    }
    clear_template(sheet)

    for offset, row_data in enumerate([row for row in report_rows if row["status"] == "green"], start=0):
        row_idx = DATA_START_ROW + offset
        values = {
            "category": CATEGORY,
            "brand": BRAND_NO_BRAND,
            "product_name": row_data["product_name"],
            "product_description": row_data["product_description"],
            "main_image": row_data["hosted_main_image_url"],
            "gtin_type": GTIN_TYPE,
            "gtin_code": row_data["ean"],
            "parcel_weight": int(row_data["parcel_weight"]),
            "parcel_length": 24,
            "parcel_width": 18,
            "parcel_height": int(row_data["parcel_height"]),
            "price": float(row_data["price"]),
            "quantity": int(row_data["quantity"]),
            "seller_sku": f"LIBRI-{row_data['ean']}",
            "rp_ids": DEFAULT_RP_ID,
            "product_property/102277": DEFAULT_PACKAGING_WARNING,
        }
        extra_images = [url for url in row_data.get("extra_image_urls", "").split("|") if url]
        for idx, url in enumerate(extra_images[:8], start=2):
            values[f"image_{idx}"] = url
        for key, value in values.items():
            col = headers.get(key)
            if col:
                sheet.cell(row=row_idx, column=col).value = value
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def write_csv(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def build(args: argparse.Namespace) -> int:
    output_dir = Path(args.output_dir)
    detail_dir = Path(args.detail_dir)
    image_dir = output_dir / "white_cover_main_images"
    ranked_items = parse_ranked_list(Path(args.input))
    candidate_paths = [
        Path("outputs/research_1000/libri_candidate_1000.csv"),
        Path("outputs/research_next_400/libri_candidate_bestseller_expanded.csv"),
        Path("outputs/research_next_400/libri_candidate_combined_full.csv"),
        Path("outputs/research_next_400/dictionary_research.csv"),
    ]
    catalog = read_candidate_csvs(candidate_paths)
    add_detail_pages(catalog, detail_dir)

    opener = login(Path(args.env)) if args.live_search or args.upload_images else None
    report_rows: list[dict[str, str]] = []
    seen_eans: set[str] = set()

    cloud_name = api_key = api_secret = folder = ""
    if args.upload_images:
        load_cloudinary_env(Path(args.env))
        import os

        cloud_name = os.environ.get("CLOUDINARY_CLOUD_NAME", "")
        api_key = os.environ.get("CLOUDINARY_API_KEY", "")
        api_secret = os.environ.get("CLOUDINARY_API_SECRET", "")
        folder = os.environ.get("CLOUDINARY_FOLDER", "tiktokshop/books")
        missing = [name for name, value in [("CLOUDINARY_CLOUD_NAME", cloud_name), ("CLOUDINARY_API_KEY", api_key), ("CLOUDINARY_API_SECRET", api_secret)] if not value]
        if missing:
            raise SystemExit("Missing Cloudinary environment values: " + ", ".join(missing))

    for ranked in ranked_items:
        item, score = best_match(ranked, catalog)
        match_source = "local_catalog"
        if args.live_search and (item is None or score < MIN_MATCH_SCORE):
            for query in [clean_query(ranked.tiktok_title), " ".join(tokens(ranked.tiktok_title)[:6])]:
                if not query:
                    continue
                for ean in search_libri(opener, query, page_size=args.search_results):
                    path = fetch_detail(opener, ean, detail_dir)
                    if path:
                        product = parse_libri_detail_html(path)
                        product.ean = product.ean or ean
                        catalog[ean] = CatalogItem(
                            ean=ean,
                            title=product.title,
                            author=product.author,
                            source_price=product.price,
                            source=f"live_query:{query}",
                            detail_path=path,
                            product=product,
                        )
            item, score = best_match(ranked, catalog)
            match_source = "live_search" if item and score >= MIN_MATCH_SCORE else match_source

        row = {
            "rank": str(ranked.rank),
            "tiktok_title": ranked.tiktok_title,
            "tiktok_price": "" if ranked.tiktok_price is None else f"{ranked.tiktok_price:.2f}",
            "tiktok_sales": "" if ranked.tiktok_sales is None else str(ranked.tiktok_sales),
            "status": "review",
            "reasons": "",
            "match_score": f"{score:.4f}",
            "match_source": match_source,
            "ean": item.ean if item else "",
            "libri_title": item.title if item else "",
            "libri_author": item.author if item else "",
            "product_group": "",
            "price": "",
            "quantity": "",
            "parcel_weight": "",
            "parcel_height": "",
            "product_name": "",
            "product_description": "",
            "cover_source": "",
            "local_main_image": "",
            "hosted_main_image_url": "",
            "extra_image_urls": "",
        }
        reasons = []
        if item is None:
            reasons.append("no_libri_match")
            row["reasons"] = ";".join(reasons)
            report_rows.append(row)
            continue
        if item.ean in seen_eans:
            reasons.append("duplicate_ean")
            row["status"] = "duplicate"
            row["reasons"] = ";".join(reasons)
            report_rows.append(row)
            continue
        product = ensure_product(item, detail_dir)
        if product is None:
            reasons.append("missing_libri_detail_page")
            row["reasons"] = ";".join(reasons)
            report_rows.append(row)
            continue
        product.ean = product.ean or item.ean
        row["product_group"] = product.product_group
        row["price"] = "" if product.price is None else f"{product.price:.2f}"
        row["quantity"] = "" if product.stock is None else str(product.stock)
        row["parcel_weight"] = str(parcel_weight(product, 50))
        row["parcel_height"] = str(parcel_height(product, 2))
        row["product_name"] = build_product_name(product)
        row["product_description"] = sanitize_description(product)

        if score < MIN_MATCH_SCORE:
            reasons.append("low_match_score")
        fit, fit_reason = category_fit(product, ranked.tiktok_title)
        if not fit:
            reasons.append(f"category_review:{fit_reason}")
        comp_status, comp_reason = compliance_status(product, ranked.tiktok_title)
        if comp_status != "ok":
            reasons.append(comp_reason)
        mismatch = language_mismatch(product, ranked.tiktok_title)
        if mismatch:
            reasons.append(mismatch)
        for field_name, value in [
            ("ean", product.ean),
            ("title", product.title),
            ("description", row["product_description"]),
            ("price", product.price),
            ("stock", product.stock),
            ("weight_g", product.weight_g),
        ]:
            if value in ("", None):
                reasons.append(f"missing_{field_name}")
        if product.stock is not None and product.stock <= 0:
            reasons.append("out_of_stock")

        detail_path = item.detail_path or (detail_dir / f"{item.ean}.html")
        cover_url, cover_image, cover_source = select_cover_image(product, detail_path)
        row["cover_source"] = cover_source
        if cover_image is None:
            reasons.append("no_safe_cover_image")
        else:
            image_path = image_dir / f"{product.ean}_main_cover_white.jpg"
            create_white_cover_image_from_image(cover_image, image_path, args.canvas_size)
            row["local_main_image"] = str(image_path.resolve())
            if args.upload_images:
                try:
                    row["hosted_main_image_url"] = upload_cover(image_path, product, folder, cloud_name, api_key, api_secret)
                except Exception as exc:  # noqa: BLE001
                    reasons.append(f"cloudinary_upload_failed:{type(exc).__name__}")
            elif cover_url:
                row["hosted_main_image_url"] = cover_url

        extra_urls = []
        for url in product.image_urls:
            if url == cover_url:
                continue
            check = fetch_image_metadata(url)
            if check.ok and check.tiktok_size_ok:
                extra_urls.append(url)
        row["extra_image_urls"] = "|".join(dict.fromkeys(extra_urls[:8]))

        if reasons:
            row["status"] = "review"
        else:
            row["status"] = "green"
            seen_eans.add(product.ean)
        row["reasons"] = ";".join(reasons)
        report_rows.append(row)

    fieldnames = list(report_rows[0].keys()) if report_rows else []
    write_csv(output_dir / "candidate_report.csv", report_rows, fieldnames)
    write_csv(output_dir / "review_hold.csv", [row for row in report_rows if row["status"] == "review"], fieldnames)
    write_csv(output_dir / "duplicates.csv", [row for row in report_rows if row["status"] == "duplicate"], fieldnames)
    write_csv(output_dir / "green_upload_rows.csv", [row for row in report_rows if row["status"] == "green"], fieldnames)

    template_path = template_path_from_arg(args.template)
    workbook_path = output_dir / "tiktok_upload_sprache_woerterbuecher.xlsx"
    write_workbook(template_path, workbook_path, report_rows)

    print(f"Input rows: {len(ranked_items)}")
    print(f"Green rows: {sum(row['status'] == 'green' for row in report_rows)}")
    print(f"Review rows: {sum(row['status'] == 'review' for row in report_rows)}")
    print(f"Duplicate rows: {sum(row['status'] == 'duplicate' for row in report_rows)}")
    print(f"Output: {output_dir.resolve()}")
    print(f"Workbook: {workbook_path.resolve()}")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description="Build TikTok Sprache/Wörterbücher upload pack from ranked TikTok list.")
    parser.add_argument("--input", default="inputs/tiktok_language_bestsellers_20260516.txt")
    parser.add_argument("--template", default="")
    parser.add_argument("--detail-dir", default="libri_bulk_pages")
    parser.add_argument("--output-dir", default="outputs/upload_ready_language_dictionaries_20260516")
    parser.add_argument("--env", default=".env")
    parser.add_argument("--live-search", action=argparse.BooleanOptionalAction, default=True)
    parser.add_argument("--search-results", type=int, default=8)
    parser.add_argument("--upload-images", action=argparse.BooleanOptionalAction, default=True)
    parser.add_argument("--canvas-size", type=int, default=1200)
    args = parser.parse_args()
    return build(args)


if __name__ == "__main__":
    raise SystemExit(main())
